import os
import time
import pandas as pd
from math import isclose, sqrt
from copy import copy
import pyomo.opt as po
import pyomo.environ as pe
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from network import Network
from energy_storage import EnergyStorage
from shared_energy_storage_data import SharedEnergyStorageData
from operational_planning_parameters import ADMMParameters
from helper_functions import *


# ======================================================================================================================
#   Class OPERATIONAL PLANNING
# ======================================================================================================================
class OperationalPlanning:

    def __init__(self, data_dir, filename):
        self.name = filename.replace('.json', '')
        self.data_dir = data_dir
        self.filename = filename
        self.results_dir = os.path.join(data_dir, 'Results')
        self.diagrams_dir = os.path.join(data_dir, 'Diagrams')
        self.cost_energy_p = 0.00
        self.cost_flex = 0.00
        self.distribution_networks = dict()
        self.transmission_network = Network()
        self.shared_ess_data = SharedEnergyStorageData()
        self.active_distribution_network_nodes = list()
        self.params = ADMMParameters()

    def run_hierarchical_coordination(self, t=0, num_steps=8, filename=str(), print_pq_map=False):
        results, models = _run_hierarchical_coordination(self, t, num_steps, print_pq_map)
        if not filename:
            filename = self.name
        self.write_operational_planning_results_to_excel(models, results, t=t, filename=filename)

    def run_distributed_coordination(self, t=0, consider_shared_ess=False, filename=str(), debug_flag=False):
        convergence, results, models, primal_evolution = _run_distributed_coordination(self, t, consider_shared_ess, debug_flag=debug_flag)
        if not filename:
            filename = self.name
        self.write_operational_planning_results_to_excel(models, results, t=t, filename=filename, primal_evolution=primal_evolution)
        return convergence, results, models, primal_evolution

    def write_operational_planning_results_to_excel(self, optimization_models, results, t=0, filename=str(), primal_evolution=list()):
        if not filename:
            filename = 'operational_planning_results'
        processed_results = _process_operational_planning_results(self, optimization_models['tso'], optimization_models['dso'], results, t=t)
        _write_operational_planning_results_to_excel(self, processed_results, t, primal_evolution=primal_evolution, filename=filename)

    def get_primal_value(self, tso_model, dso_models):
        return _get_primal_value(self, tso_model, dso_models)

    def update_admm_consensus_variables(self, tso_model, dso_models, consensus_vars, dual_vars, results, params, consider_shared_ess=False, update_tn=False, update_dns=False):
        self.update_interface_power_flow_variables(tso_model, dso_models, consensus_vars, dual_vars, results, params, update_tn=update_tn, update_dns=update_dns)
        if consider_shared_ess:
            self.update_shared_energy_storage_variables(tso_model, dso_models, consensus_vars['ess'], dual_vars['ess'], results, params, update_tn=update_tn, update_dns=update_dns)

    def update_interface_power_flow_variables(self, tso_model, dso_models, interface_vars, dual_vars, results, params, update_tn=True, update_dns=True):
        _update_interface_power_flow_variables(self, tso_model, dso_models, interface_vars, dual_vars, results, params, update_tn=update_tn, update_dns=update_dns)

    def update_shared_energy_storage_variables(self, tso_model, dso_models, consensus_vars, dual_vars, results, params, update_tn=True, update_dns=True):
        _update_shared_energy_storage_variables(self, tso_model, dso_models, consensus_vars, dual_vars, results, params, update_tn=update_tn, update_dns=update_dns)

    def read_case_study(self):
        _read_case_study(self)

    def read_market_data_from_file(self, filename):
        _read_market_data_from_file(self, filename)

    def read_operational_planning_parameters_from_file(self, params_file):
        filename = os.path.join(self.data_dir, params_file)
        self.params.read_parameters_from_file(filename)


# ======================================================================================================================
#  OPERATIONAL PLANNING PROBLEM read functions
# ======================================================================================================================
def _read_case_study(operational_planning):

    # Create results folder
    if not os.path.exists(operational_planning.results_dir):
        os.makedirs(operational_planning.results_dir)

    # Create diagrams folder
    if not os.path.exists(operational_planning.diagrams_dir):
        os.makedirs(operational_planning.diagrams_dir)

    # Read specification file
    filename = os.path.join(operational_planning.data_dir, operational_planning.filename)
    operational_planning_data = convert_json_to_dict(read_json_file(filename))

    # Market data
    market_data_filename = operational_planning_data['MarketData']['market_data_filename']
    operational_planning.read_market_data_from_file(market_data_filename)

    # Distribution Networks
    for distribution_network_data in operational_planning_data['DistributionNetworks']:

        print('[INFO] Reading DISTRIBUTION NETWORK DATA from file(s)...')

        network_name = distribution_network_data['name']                                    # Network filename
        params_filename = distribution_network_data['params_filename']                      # Params filename
        operational_data_filename = distribution_network_data['operational_data_filename']  # Operational data filename
        connection_nodeid = distribution_network_data['connection_node_id']                 # Connection node ID

        distribution_network = Network()
        distribution_network.name = network_name
        distribution_network.is_transmission = False
        distribution_network.data_dir = operational_planning.data_dir
        distribution_network.results_dir = operational_planning.results_dir
        distribution_network.diagrams_dir = operational_planning.diagrams_dir
        distribution_network.cost_energy_p = operational_planning.cost_energy_p
        distribution_network.cost_flex = operational_planning.cost_flex
        distribution_network.read_network_parameters(params_filename)
        distribution_network.read_network_data(operational_data_filename)
        distribution_network.tn_connection_nodeid = connection_nodeid
        operational_planning.distribution_networks[connection_nodeid] = distribution_network
    operational_planning.active_distribution_network_nodes = [node_id for node_id in operational_planning.distribution_networks]

    # Transmission Network
    print('[INFO] Reading TRANSMISSION NETWORK DATA from file(s)...')
    transmission_network = Network()
    transmission_network.name = operational_planning_data['TransmissionNetwork']['name']
    transmission_network.is_transmission = True
    transmission_network.data_dir = operational_planning.data_dir
    transmission_network.results_dir = operational_planning.results_dir
    transmission_network.diagrams_dir = operational_planning.diagrams_dir
    transmission_network.cost_energy_p = operational_planning.cost_energy_p
    transmission_network.cost_flex = operational_planning.cost_flex
    params_filename = operational_planning_data['TransmissionNetwork']['params_filename']
    transmission_network.read_network_parameters(params_filename)
    operational_data_filename = operational_planning_data['TransmissionNetwork']['operational_data_filename']
    transmission_network.read_network_data(operational_data_filename)
    transmission_network.active_distribution_network_nodes = [node_id for node_id in operational_planning.distribution_networks]
    operational_planning.transmission_network = transmission_network

    # Shared ESS
    print('[INFO] Reading SHARED ESS DATA from file(s)...')
    shared_ess_data = SharedEnergyStorageData()
    shared_ess_data.data_dir = operational_planning.data_dir
    shared_ess_data_filename = operational_planning_data['SharedEnergyStorages']['shared_ess_filename']
    shared_ess_data.read_shared_energy_storage_data_from_file(shared_ess_data_filename)
    operational_planning.shared_ess_data = shared_ess_data

    # Planning Parameters
    print(f'[INFO] Reading PLANNING PARAMETERS from file...')
    params_filename = operational_planning_data['PlanningParameters']['params_filename']
    operational_planning.read_operational_planning_parameters_from_file(params_filename)

    _check_interface_nodes_base_voltage_consistency(operational_planning)

    # Add Shared Energy Storages to Transmission and Distribution Networks
    #_add_shared_energy_storage_to_transmission_network(operational_planning)
    #_add_shared_energy_storage_to_distribution_network(operational_planning)


# ======================================================================================================================
#  MARKET DATA read functions
# ======================================================================================================================
def _read_market_data_from_file(operational_planning, market_data_filename):
    try:
        filename = os.path.join(operational_planning.data_dir, 'Market Data', market_data_filename)
        market_costs = _get_market_costs_from_excel_file(filename)
        operational_planning.cost_energy_p = market_costs['Cp']
        operational_planning.cost_flex = market_costs['Cflex']
    except:
        print(f'[ERROR] Reading market data from file(s). Exiting...')
        exit(ERROR_SPECIFICATION_FILE)


def _get_market_costs_from_excel_file(filename):
    data = pd.read_excel(filename)
    num_rows, num_cols = data.shape
    market_costs = dict()
    for i in range(num_rows):
        cost_type = data.iloc[i, 0]
        values = list()
        for j in range(num_cols - 1):
            values.append(data.iloc[i, j + 1])
        market_costs[cost_type] = values
    return market_costs


# ======================================================================================================================
#  HIERARCHICAL OPERATIONAL PLANNING functions
# ======================================================================================================================
def _run_hierarchical_coordination(operational_planning, t, num_steps, print_pq_map):

    print('[INFO] Running HIERARCHICAL OPERATIONAL PLANNING...')

    transmission_network = operational_planning.transmission_network
    distribution_networks = operational_planning.distribution_networks
    results = {'tso': dict(), 'dso': dict()}

    start = time.time()

    # Get DN models representation (PQ maps)
    dn_models = dict()
    for node_id in distribution_networks:
        distribution_network = distribution_networks[node_id]
        init_solution, ineqs = distribution_network.determine_pq_map(t=t, num_steps=num_steps, print_pq_map=print_pq_map)
        dn_models[node_id] = {
            'initial_solution': init_solution,
            'inequalities': ineqs
        }

    # TN model
    tn_model = transmission_network.build_model(t=t)
    tn_model.active_distribution_networks = range(len(transmission_network.active_distribution_network_nodes))

    # TN, Fix Pc, Qc at the interface nodes, free flexibility
    for dn in tn_model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        adn_load_idx = transmission_network.get_adn_load_idx(adn_node_id)
        init_solution = dn_models[adn_node_id]['initial_solution']
        for s_o in tn_model.scenarios_operation:
            tn_model.pc[adn_load_idx, s_o].setub(init_solution['Pg'] / transmission_network.baseMVA + EQUALITY_TOLERANCE)
            tn_model.pc[adn_load_idx, s_o].setlb(init_solution['Pg'] / transmission_network.baseMVA - EQUALITY_TOLERANCE)
            tn_model.qc[adn_load_idx, s_o].setub(init_solution['Qg'] / transmission_network.baseMVA + EQUALITY_TOLERANCE)
            tn_model.qc[adn_load_idx, s_o].setlb(init_solution['Qg'] / transmission_network.baseMVA - EQUALITY_TOLERANCE)
            tn_model.flex_p_up[adn_load_idx, s_o].setub(None)
            tn_model.flex_p_down[adn_load_idx, s_o].setub(None)
            tn_model.flex_q_up[adn_load_idx, s_o].setub(None)
            tn_model.flex_q_down[adn_load_idx, s_o].setub(None)
            if transmission_network.params.l_curt:
                tn_model.pc_curt_down[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
                tn_model.pc_curt_up[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
                tn_model.qc_curt_down[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
                tn_model.qc_curt_up[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)

    # TN, Add expected interface values
    tn_model.expected_interface_pf_p = pe.Var(tn_model.active_distribution_networks, domain=pe.Reals, initialize=0.00)
    tn_model.expected_interface_pf_q = pe.Var(tn_model.active_distribution_networks, domain=pe.Reals, initialize=0.00)
    tn_model.interface_expected_values = pe.ConstraintList()
    for dn in tn_model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        adn_load_idx = transmission_network.get_adn_load_idx(adn_node_id)
        expected_pf_p = 0.00
        expected_pf_q = 0.00
        for s_o in tn_model.scenarios_operation:
            omega_oper = transmission_network.prob_operation_scenarios[s_o]
            adn_pc = tn_model.pc[adn_load_idx, s_o] + tn_model.flex_p_up[adn_load_idx, s_o] - tn_model.flex_p_down[adn_load_idx, s_o]
            adn_qc = tn_model.qc[adn_load_idx, s_o] + tn_model.flex_q_up[adn_load_idx, s_o] - tn_model.flex_q_down[adn_load_idx, s_o]
            expected_pf_p += omega_oper * adn_pc
            expected_pf_q += omega_oper * adn_qc
        tn_model.interface_expected_values.add(tn_model.expected_interface_pf_p[dn] <= expected_pf_p + SMALL_TOLERANCE)
        tn_model.interface_expected_values.add(tn_model.expected_interface_pf_p[dn] >= expected_pf_p - SMALL_TOLERANCE)
        tn_model.interface_expected_values.add(tn_model.expected_interface_pf_q[dn] <= expected_pf_q + SMALL_TOLERANCE)
        tn_model.interface_expected_values.add(tn_model.expected_interface_pf_q[dn] >= expected_pf_q - SMALL_TOLERANCE)

    # TN, Add ADNs' PQ maps constraints
    tn_model.pq_maps = pe.ConstraintList()
    for dn in tn_model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        adn_pq_map = dn_models[adn_node_id]
        # initial_solution = adn_pq_map['initial_solution']
        for ineq in adn_pq_map['inequalities']:
            a = ineq['Pg']
            b = ineq['Qg']
            c = ineq['c'] / transmission_network.baseMVA
            tn_model.pq_maps.add(a * tn_model.expected_interface_pf_p[dn] + b * tn_model.expected_interface_pf_q[dn] <= c)
            # tn_model.pq_maps.add(tn_model.expected_interface_pf_p[dn] <= initial_solution['Pg'] / transmission_network.baseMVA * 1.10)
            # tn_model.pq_maps.add(tn_model.expected_interface_pf_p[dn] >= initial_solution['Pg'] / transmission_network.baseMVA * 0.90)
            # tn_model.pq_maps.add(tn_model.expected_interface_pf_q[dn] <= initial_solution['Qg'] / transmission_network.baseMVA * 1.10)
            # tn_model.pq_maps.add(tn_model.expected_interface_pf_q[dn] >= initial_solution['Qg'] / transmission_network.baseMVA * 0.90)

    # Regularization -- Added to OF to minimize deviations from scenarios to expected values
    obj = copy(tn_model.objective.expr)
    tn_model.penalty_regularization = pe.Var(domain=pe.NonNegativeReals)
    tn_model.penalty_regularization.fix(PENALTY_REGULARIZATION)
    for dn in tn_model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        adn_load_idx = transmission_network.get_adn_load_idx(adn_node_id)
        for s_o in tn_model.scenarios_operation:
            obj += tn_model.penalty_regularization * transmission_network.baseMVA * (tn_model.pc[adn_load_idx, s_o] - tn_model.expected_interface_pf_p[dn]) ** 2
            obj += tn_model.penalty_regularization * transmission_network.baseMVA * (tn_model.qc[adn_load_idx, s_o] - tn_model.expected_interface_pf_q[dn]) ** 2
    tn_model.objective.expr = obj

    # Optimize TN, Get resulting interface PFs
    results['tso'] = transmission_network.optimize(tn_model)
    pf_requested = dict()
    for dn in tn_model.active_distribution_networks:

        adn_node_id = transmission_network.active_distribution_network_nodes[dn]

        pc = pe.value(tn_model.expected_interface_pf_p[dn]) * transmission_network.baseMVA
        qc = pe.value(tn_model.expected_interface_pf_q[dn]) * transmission_network.baseMVA
        pf_requested[adn_node_id] = {'p': pc, 'q': qc}

    # Run OPF on DNs, considering established power flow (settlement)
    dn_models = dict()
    for node_id in distribution_networks:

        distribution_network = distribution_networks[node_id]
        dn_model = distribution_network.build_model(t=t)
        distribution_network.update_of_to_settlement(dn_model)

        dn_model.interface_pf_p_req.fix(pf_requested[node_id]['p'] / distribution_network.baseMVA)
        dn_model.interface_pf_q_req.fix(pf_requested[node_id]['q'] / distribution_network.baseMVA)
        results['dso'][node_id] = distribution_network.optimize(dn_model)

        dn_models[node_id] = dn_model

    end = time.time()
    total_execution_time = end - start
    print('[INFO] \t - Execution time: {:.2f} s'.format(total_execution_time))

    optim_models = {'tso': tn_model, 'dso': dn_models}

    return results, optim_models


# ======================================================================================================================
#  DISTRIBUTED OPERATIONAL PLANNING functions
# ======================================================================================================================
def _run_distributed_coordination(operational_planning, t, consider_shared_ess, debug_flag=False):

    print('[INFO] Running DISTRIBUTED OPERATIONAL PLANNING...')

    transmission_network = operational_planning.transmission_network
    distribution_networks = operational_planning.distribution_networks
    admm_parameters = operational_planning.params
    results = {'tso': dict(), 'dso': dict()}

    # ------------------------------------------------------------------------------------------------------------------
    # 0. Initialization

    print('[INFO]\t\t - Initializing...')

    start = time.time()
    from_warm_start = False
    primal_evolution = list()

    # Create ADMM variables
    consensus_vars, dual_vars = create_admm_variables(operational_planning)

    # Create ADN models, get initial power flows
    dso_models, results['dso'] = create_distribution_networks_models(distribution_networks, consensus_vars, t, consider_shared_ess)
    tso_model, results['tso'] = create_transmission_network_model(transmission_network, consensus_vars, t, consider_shared_ess)

    # Update models to ADMM
    update_distribution_models_to_admm(operational_planning, dso_models, admm_parameters, consider_shared_ess)
    update_transmission_model_to_admm(operational_planning, tso_model, admm_parameters, consider_shared_ess)

    # Update consensus variables
    operational_planning.update_admm_consensus_variables(tso_model, dso_models, consensus_vars, dual_vars, results, admm_parameters, consider_shared_ess, update_tn=True, update_dns=True)

    # ------------------------------------------------------------------------------------------------------------------
    # ADMM -- Main cycle
    # ------------------------------------------------------------------------------------------------------------------
    convergence, iter = False, 1
    for iter in range(iter, admm_parameters.num_max_iters + 1):

        print(f'[INFO]\t - ADMM. Iter {iter}...')

        iter_start = time.time()

        # --------------------------------------------------------------------------------------------------------------
        # 1. Solve DSOs problems
        results['dso'] = update_distribution_coordination_models_and_solve(distribution_networks, dso_models,
                                                                           consensus_vars['v_sqr'], dual_vars['v_sqr']['dso'],
                                                                           consensus_vars['pf'], dual_vars['pf']['dso'],
                                                                           consensus_vars['ess'], dual_vars['ess']['dso'],
                                                                           admm_parameters, consider_shared_ess,
                                                                           from_warm_start=from_warm_start)

        # 1.1 Update ADMM CONSENSUS variables
        operational_planning.update_admm_consensus_variables(tso_model, dso_models,
                                                             consensus_vars, dual_vars, results, admm_parameters,
                                                             consider_shared_ess=consider_shared_ess, update_dns=True)

        # 1.2 Update primal evolution
        primal_evolution.append(operational_planning.get_primal_value(tso_model, dso_models))

        # 1.3 STOPPING CRITERIA evaluation
        if iter > 1:
            convergence = check_admm_convergence(operational_planning, consensus_vars, admm_parameters, consider_shared_ess, debug_flag=debug_flag)
            if convergence:
                iter_end = time.time()
                print('[INFO] \t - Iter {}: {:.2f} s'.format(iter, iter_end - iter_start))
                break

        # --------------------------------------------------------------------------------------------------------------
        # 2. Solve TSO problem
        results['tso'] = update_transmission_coordination_model_and_solve(transmission_network, tso_model,
                                                                          consensus_vars['v_sqr'], dual_vars['v_sqr']['tso'],
                                                                          consensus_vars['pf'], dual_vars['pf']['tso'],
                                                                          consensus_vars['ess'], dual_vars['ess']['tso'],
                                                                          admm_parameters, consider_shared_ess,
                                                                          from_warm_start=from_warm_start)

        # 2.1 Update ADMM CONSENSUS variables
        operational_planning.update_admm_consensus_variables(tso_model, dso_models,
                                                             consensus_vars, dual_vars, results, admm_parameters,
                                                             consider_shared_ess=consider_shared_ess, update_tn=True)

        # 2.2 Update primal evolution
        primal_evolution.append(operational_planning.get_primal_value(tso_model, dso_models))

        # 2.3 STOPPING CRITERIA evaluation
        convergence = check_admm_convergence(operational_planning, consensus_vars, admm_parameters, consider_shared_ess, debug_flag=debug_flag)
        if convergence:
            iter_end = time.time()
            print('[INFO] \t - Iter {}: {:.2f} s'.format(iter, iter_end - iter_start))
            break

        iter_end = time.time()
        print('[INFO] \t - Iter {}: {:.2f} s'.format(iter, iter_end - iter_start))

        from_warm_start = True

    if not convergence:
        print(f'[WARNING] ADMM did NOT converge in {admm_parameters.num_max_iters} iterations!')
    else:
        print(f'[INFO] \t - ADMM converged in {iter} iterations.')

    end = time.time()
    total_execution_time = end - start
    print('[INFO] \t - Execution time: {:.2f} s'.format(total_execution_time))

    optim_models = {'tso': tso_model, 'dso': dso_models}

    return convergence, results, optim_models, primal_evolution


def create_transmission_network_model(transmission_network, consensus_vars, t, consider_shared_ess):

    # Build model
    tso_model = transmission_network.build_model(t)
    s_base = transmission_network.baseMVA

    # Update model with expected interface values
    tso_model.active_distribution_networks = range(len(transmission_network.active_distribution_network_nodes))

    # Free Vmag, Pc, Qc at the interface nodes
    for dn in tso_model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        adn_node_idx = transmission_network.get_node_idx(adn_node_id)
        adn_load_idx = transmission_network.get_adn_load_idx(adn_node_id)
        _, v_max = transmission_network.get_node_voltage_limits(adn_node_id)
        for s_o in tso_model.scenarios_operation:

            tso_model.e[adn_node_idx, s_o].fixed = False
            tso_model.e[adn_node_idx, s_o].setub(v_max + SMALL_TOLERANCE)
            tso_model.e[adn_node_idx, s_o].setlb(-v_max - SMALL_TOLERANCE)
            tso_model.f[adn_node_idx, s_o].fixed = False
            tso_model.f[adn_node_idx, s_o].setub(v_max + SMALL_TOLERANCE)
            tso_model.f[adn_node_idx, s_o].setlb(-v_max - SMALL_TOLERANCE)
            if transmission_network.params.slacks.grid_operation.voltage:
                tso_model.slack_e[adn_node_idx, s_o].setub(EQUALITY_TOLERANCE)
                tso_model.slack_e[adn_node_idx, s_o].setlb(-EQUALITY_TOLERANCE)
                tso_model.slack_f[adn_node_idx, s_o].setub(EQUALITY_TOLERANCE)
                tso_model.slack_f[adn_node_idx, s_o].setlb(-EQUALITY_TOLERANCE)

            tso_model.pc[adn_load_idx, s_o].fixed = False
            tso_model.pc[adn_load_idx, s_o].setub(None)
            tso_model.pc[adn_load_idx, s_o].setlb(None)
            tso_model.qc[adn_load_idx, s_o].fixed = False
            tso_model.qc[adn_load_idx, s_o].setub(None)
            tso_model.qc[adn_load_idx, s_o].setlb(None)
            if transmission_network.params.fl_reg:
                tso_model.flex_p_up[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
                tso_model.flex_p_down[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
                tso_model.flex_q_up[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
                tso_model.flex_q_down[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
            if transmission_network.params.l_curt:
                tso_model.pc_curt_down[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
                tso_model.pc_curt_up[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
                tso_model.qc_curt_down[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)
                tso_model.qc_curt_up[adn_load_idx, s_o].setub(EQUALITY_TOLERANCE)

    # Add expected interface and shared ESS values
    tso_model.interface_expected_values = pe.ConstraintList()
    tso_model.expected_interface_vmag_sqr = pe.Var(tso_model.active_distribution_networks, domain=pe.NonNegativeReals, initialize=1.00)
    tso_model.expected_interface_pf_p = pe.Var(tso_model.active_distribution_networks, domain=pe.Reals, initialize=0.00)
    tso_model.expected_interface_pf_q = pe.Var(tso_model.active_distribution_networks, domain=pe.Reals, initialize=0.00)
    for dn in tso_model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        adn_node_idx = transmission_network.get_node_idx(adn_node_id)
        adn_load_idx = transmission_network.get_adn_load_idx(adn_node_id)
        expected_vmag_sqr = 0.00
        expected_pf_p = 0.00
        expected_pf_q = 0.00
        for s_o in tso_model.scenarios_operation:
            omega_oper = transmission_network.prob_operation_scenarios[s_o]
            expected_vmag_sqr += omega_oper * (tso_model.e[adn_node_idx, s_o] ** 2 + tso_model.f[adn_node_idx, s_o] ** 2)
            expected_pf_p += omega_oper * tso_model.pc[adn_load_idx, s_o]
            expected_pf_q += omega_oper * tso_model.qc[adn_load_idx, s_o]
        tso_model.interface_expected_values.add(tso_model.expected_interface_vmag_sqr[dn] <= expected_vmag_sqr + SMALL_TOLERANCE)
        tso_model.interface_expected_values.add(tso_model.expected_interface_vmag_sqr[dn] >= expected_vmag_sqr - SMALL_TOLERANCE)
        tso_model.interface_expected_values.add(tso_model.expected_interface_pf_p[dn] <= expected_pf_p + SMALL_TOLERANCE)
        tso_model.interface_expected_values.add(tso_model.expected_interface_pf_p[dn] >= expected_pf_p - SMALL_TOLERANCE)
        tso_model.interface_expected_values.add(tso_model.expected_interface_pf_q[dn] <= expected_pf_q + SMALL_TOLERANCE)
        tso_model.interface_expected_values.add(tso_model.expected_interface_pf_q[dn] >= expected_pf_q - SMALL_TOLERANCE)
    if consider_shared_ess:
        tso_model.expected_shared_ess_p = pe.Var(tso_model.shared_energy_storages, domain=pe.Reals, initialize=0.00)
        tso_model.expected_shared_ess_q = pe.Var(tso_model.shared_energy_storages, domain=pe.Reals, initialize=0.00)
        for e in tso_model.shared_energy_storages:
            expected_ess_p = 0.00
            expected_ess_q = 0.00
            for s_o in tso_model.scenarios_operation:
                omega_oper = transmission_network.prob_operation_scenarios[s_o]
                expected_ess_p += omega_oper * tso_model.shared_es_pnet[e, s_o]
                expected_ess_q += omega_oper * tso_model.shared_es_qnet[e, s_o]
            tso_model.interface_expected_values.add(tso_model.expected_shared_ess_p[e] <= expected_ess_p + SMALL_TOLERANCE)
            tso_model.interface_expected_values.add(tso_model.expected_shared_ess_p[e] >= expected_ess_p - SMALL_TOLERANCE)
            tso_model.interface_expected_values.add(tso_model.expected_shared_ess_q[e] <= expected_ess_q + SMALL_TOLERANCE)
            tso_model.interface_expected_values.add(tso_model.expected_shared_ess_q[e] >= expected_ess_q - SMALL_TOLERANCE)

    # Regularization -- Added to OF to minimize deviations from scenarios to expected values
    obj = copy(tso_model.objective.expr)
    tso_model.penalty_regularization = pe.Var(domain=pe.NonNegativeReals)
    tso_model.penalty_regularization.fix(PENALTY_REGULARIZATION)
    for dn in tso_model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        adn_node_idx = transmission_network.get_node_idx(adn_node_id)
        adn_load_idx = transmission_network.get_adn_load_idx(adn_node_id)
        for s_o in tso_model.scenarios_operation:
            obj += tso_model.penalty_regularization * ((tso_model.e[adn_node_idx, s_o] ** 2 + tso_model.f[adn_node_idx, s_o] ** 2) - tso_model.expected_interface_vmag_sqr[dn]) ** 2
            obj += tso_model.penalty_regularization * s_base * (tso_model.pc[adn_load_idx, s_o] - tso_model.expected_interface_pf_p[dn]) ** 2
            obj += tso_model.penalty_regularization * s_base * (tso_model.qc[adn_load_idx, s_o] - tso_model.expected_interface_pf_q[dn]) ** 2
    if consider_shared_ess:
        for e in tso_model.shared_energy_storages:
            for s_o in tso_model.scenarios_operation:
                obj += tso_model.penalty_regularization * s_base * (tso_model.shared_es_pnet[e, s_o] - tso_model.expected_shared_ess_p[e]) ** 2
                obj += tso_model.penalty_regularization * s_base * (tso_model.shared_es_qnet[e, s_o] - tso_model.expected_shared_ess_q[e]) ** 2
    tso_model.objective.expr = obj

    # Fix initial values, run OPF
    for dn in tso_model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        v_base = transmission_network.get_node_base_kv(adn_node_id)
        interface_v_sqr = consensus_vars['v_sqr']['dso']['current'][adn_node_id] / (v_base ** 2)
        interface_pf_p = consensus_vars['pf']['dso']['current'][adn_node_id]['p'] / s_base
        interface_pf_q = consensus_vars['pf']['dso']['current'][adn_node_id]['q'] / s_base
        tso_model.expected_interface_vmag_sqr[dn].fix(interface_v_sqr)
        tso_model.expected_interface_pf_p[dn].fix(interface_pf_p)
        tso_model.expected_interface_pf_q[dn].fix(interface_pf_q)

    # Run SMOPF
    results = transmission_network.optimize(tso_model)

    # Get initial interface and shared ESS values
    for dn in tso_model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        v_base = transmission_network.get_node_base_kv(adn_node_id)
        interface_v_sqr = pe.value(tso_model.expected_interface_vmag_sqr[dn]) * (v_base ** 2)
        interface_pf_p = pe.value(tso_model.expected_interface_pf_p[dn]) * s_base
        interface_pf_q = pe.value(tso_model.expected_interface_pf_q[dn]) * s_base
        consensus_vars['v_sqr']['tso']['current'][adn_node_id] = interface_v_sqr
        consensus_vars['pf']['tso']['current'][adn_node_id]['p'] = interface_pf_p
        consensus_vars['pf']['tso']['current'][adn_node_id]['q'] = interface_pf_q
        if consider_shared_ess:
            shared_ess_idx = transmission_network.get_shared_energy_storage_idx(adn_node_id)
            p_ess = pe.value(tso_model.expected_shared_ess_p[shared_ess_idx]) * s_base
            q_ess = pe.value(tso_model.expected_shared_ess_q[shared_ess_idx]) * s_base
            consensus_vars['ess']['tso']['current'][adn_node_id]['p'] = p_ess
            consensus_vars['ess']['tso']['current'][adn_node_id]['q'] = q_ess

    return tso_model, results


def create_distribution_networks_models(distribution_networks, consensus_vars, t, consider_shared_ess):

    dso_models = dict()
    results = dict()

    for node_id in distribution_networks:

        distribution_network = distribution_networks[node_id]
        s_base = distribution_network.baseMVA
        ref_node_id = distribution_network.get_reference_node_id()
        ref_node_idx = distribution_network.get_node_idx(ref_node_id)
        ref_gen_idx = distribution_network.get_reference_gen_idx()

        # Build model
        dso_model = distribution_network.build_model(t)

        # Update model with expected interface values
        dso_model.interface_expected_values = pe.ConstraintList()

        dso_model.expected_interface_vmag_sqr = pe.Var(domain=pe.NonNegativeReals, initialize=1.00)
        dso_model.expected_interface_pf_p = pe.Var(domain=pe.Reals, initialize=0.00)
        dso_model.expected_interface_pf_q = pe.Var(domain=pe.Reals, initialize=0.00)
        expected_vmag_sqr = 0.00
        expected_pf_p = 0.00
        expected_pf_q = 0.00
        for s_o in dso_model.scenarios_operation:
            omega_oper = distribution_network.prob_operation_scenarios[s_o]
            expected_vmag_sqr += omega_oper * dso_model.e[ref_node_idx, s_o] ** 2
            expected_pf_p += omega_oper * dso_model.pg[ref_gen_idx, s_o]
            expected_pf_q += omega_oper * dso_model.qg[ref_gen_idx, s_o]
        dso_model.interface_expected_values.add(dso_model.expected_interface_vmag_sqr <= expected_vmag_sqr + SMALL_TOLERANCE)
        dso_model.interface_expected_values.add(dso_model.expected_interface_vmag_sqr >= expected_vmag_sqr - SMALL_TOLERANCE)
        dso_model.interface_expected_values.add(dso_model.expected_interface_pf_p <= expected_pf_p + SMALL_TOLERANCE)
        dso_model.interface_expected_values.add(dso_model.expected_interface_pf_p >= expected_pf_p - SMALL_TOLERANCE)
        dso_model.interface_expected_values.add(dso_model.expected_interface_pf_q <= expected_pf_q + SMALL_TOLERANCE)
        dso_model.interface_expected_values.add(dso_model.expected_interface_pf_q >= expected_pf_q - SMALL_TOLERANCE)

        # Regularization -- Added to OF to minimize deviations from scenarios to expected values
        obj = copy(dso_model.objective.expr)
        dso_model.penalty_regularization = pe.Var(domain=pe.NonNegativeReals)
        dso_model.penalty_regularization.fix(PENALTY_REGULARIZATION)
        for s_o in dso_model.scenarios_operation:
            obj += dso_model.penalty_regularization * (dso_model.e[ref_node_idx, s_o] ** 2 - dso_model.expected_interface_vmag_sqr) ** 2
            obj += dso_model.penalty_regularization * s_base * (dso_model.pg[ref_gen_idx, s_o] - dso_model.expected_interface_pf_p) ** 2
            obj += dso_model.penalty_regularization * s_base * (dso_model.qg[ref_gen_idx, s_o] - dso_model.expected_interface_pf_q) ** 2
        dso_model.objective.expr = obj

        if consider_shared_ess:
            shared_ess_idx = distribution_network.get_shared_energy_storage_idx(ref_node_id)
            dso_model.expected_shared_ess_p = pe.Var(domain=pe.Reals, initialize=0.00)
            dso_model.expected_shared_ess_q = pe.Var(domain=pe.Reals, initialize=0.00)
            expected_ess_p = 0.00
            expected_ess_q = 0.00
            for s_o in dso_model.scenarios_operation:
                omega_oper = distribution_network.prob_operation_scenarios[s_o]
                expected_ess_p += omega_oper * dso_model.shared_es_pnet[shared_ess_idx, s_o]
                expected_ess_q += omega_oper * dso_model.shared_es_qnet[shared_ess_idx, s_o]
            dso_model.interface_expected_values.add(dso_model.expected_shared_ess_p <= expected_ess_p + SMALL_TOLERANCE)
            dso_model.interface_expected_values.add(dso_model.expected_shared_ess_p >= expected_ess_p - SMALL_TOLERANCE)
            dso_model.interface_expected_values.add(dso_model.expected_shared_ess_q <= expected_ess_q + SMALL_TOLERANCE)
            dso_model.interface_expected_values.add(dso_model.expected_shared_ess_q >= expected_ess_q - SMALL_TOLERANCE)

            for s_o in dso_model.scenarios_operation:
                obj += dso_model.penalty_regularization * s_base * (dso_model.shared_es_pnet[shared_ess_idx, s_o] - dso_model.expected_shared_ess_p) ** 2
                obj += dso_model.penalty_regularization * s_base * (dso_model.shared_es_qnet[shared_ess_idx, s_o] - dso_model.expected_shared_ess_q) ** 2

        dso_model.objective.expr = obj

        # Run SMOPF
        results[node_id] = distribution_network.optimize(dso_model)

        # Get initial interface and shared ESS values
        v_base = distribution_network.get_node_base_kv(ref_node_id)
        interface_v_sqr = pe.value(dso_model.expected_interface_vmag_sqr) * (v_base ** 2)
        interface_pf_p = pe.value(dso_model.expected_interface_pf_p) * s_base
        interface_pf_q = pe.value(dso_model.expected_interface_pf_q) * s_base
        consensus_vars['v_sqr']['dso']['current'][node_id] = interface_v_sqr
        consensus_vars['pf']['dso']['current'][node_id]['p'] = interface_pf_p
        consensus_vars['pf']['dso']['current'][node_id]['q'] = interface_pf_q
        if consider_shared_ess:
            p_ess = pe.value(dso_model.expected_shared_ess_p) * s_base
            q_ess = pe.value(dso_model.expected_shared_ess_q) * s_base
            consensus_vars['ess']['dso']['current'][node_id]['p'] = p_ess
            consensus_vars['ess']['dso']['current'][node_id]['q'] = q_ess

        dso_models[node_id] = dso_model

    return dso_models, results


def update_transmission_model_to_admm(operational_planning, model, params, consider_shared_ess):

    transmission_network = operational_planning.transmission_network
    distribution_networks = operational_planning.distribution_networks

    s_base = transmission_network.baseMVA

    # Free expected values
    for dn in model.active_distribution_networks:
        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        v_min, v_max = transmission_network.get_node_voltage_limits(adn_node_id)
        model.expected_interface_vmag_sqr[dn].fixed = False
        model.expected_interface_vmag_sqr[dn].setub(v_max ** 2 + SMALL_TOLERANCE)
        model.expected_interface_vmag_sqr[dn].setlb(v_min ** 2 - SMALL_TOLERANCE)
        model.expected_interface_pf_p[dn].fixed = False
        model.expected_interface_pf_p[dn].setub(None)
        model.expected_interface_pf_p[dn].setlb(None)
        model.expected_interface_pf_q[dn].fixed = False
        model.expected_interface_pf_q[dn].setub(None)
        model.expected_interface_pf_q[dn].setlb(None)
        if consider_shared_ess:
            shared_ess_idx = transmission_network.get_shared_energy_storage_idx(adn_node_id)
            model.expected_shared_ess_p[shared_ess_idx].setub(None)
            model.expected_shared_ess_p[shared_ess_idx].setlb(None)
            model.expected_shared_ess_q[shared_ess_idx].setub(None)
            model.expected_shared_ess_q[shared_ess_idx].setlb(None)

    # Update costs (penalties) for the coordination procedure
    model.penalty_ess_usage.fix(1e-6)
    if transmission_network.params.obj_type == OBJ_MIN_COST:
        model.cost_res_curtailment.fix(COST_GENERATION_CURTAILMENT)
        model.cost_load_curtailment.fix(COST_CONSUMPTION_CURTAILMENT)
    elif transmission_network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
        model.penalty_gen_curtailment.fix(1e-2)
        model.penalty_load_curtailment.fix(PENALTY_LOAD_CURTAILMENT)
        model.penalty_flex_usage.fix(1e-3)

    # Add ADMM variables
    model.rho_v = pe.Var(domain=pe.NonNegativeReals)
    model.rho_v.fix(params.rho['v'][transmission_network.name])
    model.v_sqr_req = pe.Var(model.active_distribution_networks, domain=pe.NonNegativeReals)    # Square of voltage magnitude
    model.dual_v_sqr_req = pe.Var(model.active_distribution_networks, domain=pe.Reals)          # Dual variable - voltage magnitude requested

    model.rho_pf = pe.Var(domain=pe.NonNegativeReals)
    model.rho_pf.fix(params.rho['pf'][transmission_network.name])
    model.p_pf_req = pe.Var(model.active_distribution_networks, domain=pe.Reals)                # Active power - requested by distribution networks
    model.q_pf_req = pe.Var(model.active_distribution_networks, domain=pe.Reals)                # Reactive power - requested by distribution networks
    model.dual_pf_p_req = pe.Var(model.active_distribution_networks, domain=pe.Reals)           # Dual variable - active power requested
    model.dual_pf_q_req = pe.Var(model.active_distribution_networks, domain=pe.Reals)           # Dual variable - reactive power requested

    if consider_shared_ess:
        model.rho_ess = pe.Var(domain=pe.NonNegativeReals)
        model.rho_ess.fix(params.rho['ess'][transmission_network.name])
        model.p_ess_req = pe.Var(model.shared_energy_storages, domain=pe.Reals)                     # Shared ESS - Active power requested (DSO)
        model.q_ess_req = pe.Var(model.shared_energy_storages, domain=pe.Reals)                     # Shared ESS - Reactive power requested (DSO)
        model.dual_ess_p_req = pe.Var(model.shared_energy_storages, domain=pe.Reals)                # Dual variable - Shared ESS active power
        model.dual_ess_q_req = pe.Var(model.shared_energy_storages, domain=pe.Reals)                # Dual variable - Shared ESS reactive power
        if params.previous_iter['ess']['tso']:
            model.rho_ess_prev = pe.Var(domain=pe.NonNegativeReals)
            model.rho_ess_prev.fix(params.rho_previous_iter['ess'][transmission_network.name])
            model.p_ess_prev = pe.Var(model.shared_energy_storages, domain=pe.Reals)                # Shared ESS - previous iteration active power
            model.q_ess_prev = pe.Var(model.shared_energy_storages, domain=pe.Reals)                # Shared ESS - previous iteration reactive power
            model.dual_ess_p_prev = pe.Var(model.shared_energy_storages, domain=pe.Reals)           # Dual variable - previous iteration shared ESS active power
            model.dual_ess_q_prev = pe.Var(model.shared_energy_storages, domain=pe.Reals)           # Dual variable - previous iteration shared ESS reactive power

    # Objective function - augmented Lagrangian
    init_of_value = 1.00
    if transmission_network.params.obj_type == OBJ_MIN_COST:
        init_of_value = abs(pe.value(model.objective))
    if isclose(init_of_value, 0.00, abs_tol=SMALL_TOLERANCE):
        init_of_value = 0.01
    obj = copy(model.objective.expr) / init_of_value

    for dn in model.active_distribution_networks:

        adn_node_id = transmission_network.active_distribution_network_nodes[dn]
        distribution_network = distribution_networks[adn_node_id]
        interface_transf_rating = distribution_network.get_interface_branch_rating() / s_base

        constraint_v_req = (model.expected_interface_vmag_sqr[dn] - model.v_sqr_req[dn])
        obj += model.dual_v_sqr_req[dn] * constraint_v_req
        obj += (model.rho_v / 2) * (constraint_v_req ** 2)

        constraint_p_req = (model.expected_interface_pf_p[dn] - model.p_pf_req[dn]) / interface_transf_rating
        constraint_q_req = (model.expected_interface_pf_q[dn] - model.q_pf_req[dn]) / interface_transf_rating
        obj += model.dual_pf_p_req[dn] * constraint_p_req
        obj += model.dual_pf_q_req[dn] * constraint_q_req
        obj += (model.rho_pf / 2) * (constraint_p_req ** 2)
        obj += (model.rho_pf / 2) * (constraint_q_req ** 2)

    if consider_shared_ess:
        for e in model.shared_energy_storages:

            shared_ess_rating = abs(transmission_network.network.shared_energy_storages[e].s)
            if isclose(shared_ess_rating, 0.00, abs_tol=SMALL_TOLERANCE):
                shared_ess_rating = 0.01

            constraint_ess_p_req = (model.expected_shared_ess_p[e] - model.p_ess_req[e]) / (2 * shared_ess_rating)
            constraint_ess_q_req = (model.expected_shared_ess_q[e] - model.q_ess_req[e]) / (2 * shared_ess_rating)
            obj += (model.dual_ess_p_req[e]) * constraint_ess_p_req
            obj += (model.dual_ess_q_req[e]) * constraint_ess_q_req
            obj += (model.rho_ess / 2) * constraint_ess_p_req ** 2
            obj += (model.rho_ess / 2) * constraint_ess_q_req ** 2
            if params.previous_iter['ess']['tso']:
                constraint_ess_p_prev = (model.expected_shared_ess_p[e] - model.p_ess_prev[e]) / (2 * shared_ess_rating)
                constraint_ess_q_prev = (model.expected_shared_ess_q[e] - model.q_ess_prev[e]) / (2 * shared_ess_rating)
                obj += (model.rho_ess_prev / 2) * constraint_ess_p_prev ** 2
                obj += (model.rho_ess_prev / 2) * constraint_ess_q_prev ** 2

    # Add ADMM OF, deactivate original OF
    model.objective.deactivate()
    model.admm_objective = pe.Objective(sense=pe.minimize, expr=obj)


def update_distribution_models_to_admm(operational_planning, models, params, consider_shared_ess):

    distribution_networks = operational_planning.distribution_networks

    for node_id in distribution_networks:

        dso_model = models[node_id]
        distribution_network = distribution_networks[node_id]
        s_base = distribution_network.baseMVA
        ref_node_id = distribution_network.get_reference_node_id()
        ref_node_idx = distribution_network.get_node_idx(ref_node_id)
        ref_gen_idx = distribution_network.get_reference_gen_idx()
        _, v_max = distribution_network.get_node_voltage_limits(ref_node_id)

        # Update Vmag, Pg, Qg limits at the interface node
        for s_o in dso_model.scenarios_operation:
            dso_model.e[ref_node_idx, s_o].fixed = False
            dso_model.e[ref_node_idx, s_o].setub(v_max + SMALL_TOLERANCE)
            dso_model.e[ref_node_idx, s_o].setlb(-v_max - SMALL_TOLERANCE)
            dso_model.f[ref_node_idx, s_o].setub(SMALL_TOLERANCE)
            dso_model.f[ref_node_idx, s_o].setlb(-SMALL_TOLERANCE)
            if distribution_network.params.slacks.grid_operation.voltage:
                dso_model.slack_e[ref_node_idx, s_o].setub(SMALL_TOLERANCE)
                dso_model.slack_e[ref_node_idx, s_o].setlb(-SMALL_TOLERANCE)
                dso_model.slack_f[ref_node_idx, s_o].setub(SMALL_TOLERANCE)
                dso_model.slack_f[ref_node_idx, s_o].setlb(-SMALL_TOLERANCE)
            dso_model.pg[ref_gen_idx, s_o].fixed = False
            dso_model.qg[ref_gen_idx, s_o].fixed = False
            if distribution_network.params.rg_curt:
                dso_model.sg_curt[ref_gen_idx, s_o].setub(SMALL_TOLERANCE)

        # Update expected interface values limits
        dso_model.expected_interface_vmag_sqr.fixed = False
        dso_model.expected_interface_vmag_sqr.setub(None)
        dso_model.expected_interface_vmag_sqr.setlb(None)
        dso_model.expected_interface_pf_p.fixed = False
        dso_model.expected_interface_pf_p.setub(None)
        dso_model.expected_interface_pf_p.setlb(None)
        dso_model.expected_interface_pf_q.fixed = False
        dso_model.expected_interface_pf_q.setub(None)
        dso_model.expected_interface_pf_q.setlb(None)
        if consider_shared_ess:
            dso_model.expected_shared_ess_p.fixed = False
            dso_model.expected_shared_ess_p.setub(None)
            dso_model.expected_shared_ess_p.setlb(None)
            dso_model.expected_shared_ess_q.fixed = False
            dso_model.expected_shared_ess_q.setub(None)
            dso_model.expected_shared_ess_q.setlb(None)

        # Update costs (penalties) for the coordination procedure
        dso_model.penalty_ess_usage.fix(0.00)
        if distribution_network.params.obj_type == OBJ_MIN_COST:
            dso_model.cost_res_curtailment.fix(0.00)
            dso_model.cost_load_curtailment.fix(COST_CONSUMPTION_CURTAILMENT)
        elif distribution_network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
            dso_model.penalty_gen_curtailment.fix(0.00)
            dso_model.penalty_load_curtailment.fix(PENALTY_LOAD_CURTAILMENT)
            dso_model.penalty_flex_usage.fix(0.00)

        # Add ADMM variables
        dso_model.rho_v = pe.Var(domain=pe.NonNegativeReals)
        dso_model.rho_v.fix(params.rho['v'][distribution_network.name])
        dso_model.v_sqr_req = pe.Var(domain=pe.NonNegativeReals)       # Voltage magnitude - requested by TSO
        dso_model.dual_v_sqr_req = pe.Var(domain=pe.Reals)             # Dual variable - voltage magnitude

        dso_model.rho_pf = pe.Var(domain=pe.NonNegativeReals)
        dso_model.rho_pf.fix(params.rho['pf'][distribution_network.name])
        dso_model.p_pf_req = pe.Var(domain=pe.Reals)                   # Active power - requested by TSO
        dso_model.q_pf_req = pe.Var(domain=pe.Reals)                   # Reactive power - requested by TSO
        dso_model.dual_pf_p_req = pe.Var(domain=pe.Reals)              # Dual variable - active power
        dso_model.dual_pf_q_req = pe.Var(domain=pe.Reals)              # Dual variable - reactive power

        if consider_shared_ess:
            dso_model.rho_ess = pe.Var(domain=pe.NonNegativeReals)
            dso_model.rho_ess.fix(params.rho['ess'][distribution_network.name])
            dso_model.p_ess_req = pe.Var(domain=pe.Reals)                  # Shared ESS - active power requested (TSO)
            dso_model.q_ess_req = pe.Var(domain=pe.Reals)                  # Shared ESS - reactive power requested (TSO)
            dso_model.dual_ess_p_req = pe.Var(domain=pe.Reals)             # Dual variable - Shared ESS active power
            dso_model.dual_ess_q_req = pe.Var(domain=pe.Reals)             # Dual variable - Shared ESS reactive power
            if params.previous_iter['ess']['dso']:
                dso_model.rho_ess_prev = pe.Var(domain=pe.NonNegativeReals)
                dso_model.rho_ess_prev.fix(params.rho_previous_iter['ess'][distribution_network.name])
                dso_model.p_ess_prev = pe.Var(domain=pe.Reals)             # Shared ESS - previous iteration active power
                dso_model.q_ess_prev = pe.Var(domain=pe.Reals)             # Shared ESS - previous iteration reactive power
                dso_model.dual_ess_p_prev = pe.Var(domain=pe.Reals)        # Dual variable - Shared ESS previous iteration active power
                dso_model.dual_ess_q_prev = pe.Var(domain=pe.Reals)        # Dual variable - Shared ESS previous iteration reactive power

        # Objective function - augmented Lagrangian
        init_of_value = 1.00
        if distribution_network.params.obj_type == OBJ_MIN_COST:
            init_of_value = abs(pe.value(dso_model.objective))
        if isclose(init_of_value, 0.00, abs_tol=SMALL_TOLERANCE):
            init_of_value = 0.01
        obj = copy(dso_model.objective.expr) / init_of_value

        if consider_shared_ess:
            shared_ess_idx = distribution_network.get_shared_energy_storage_idx(ref_node_id)
            shared_ess_rating = abs(distribution_network.shared_energy_storages[shared_ess_idx].s)
            if isclose(shared_ess_rating, 0.00, abs_tol=SMALL_TOLERANCE):
                shared_ess_rating = 0.01

        interface_transf_rating = distribution_network.get_interface_branch_rating() / s_base

        # Augmented Lagrangian -- Interface power flow (residual balancing)

        # Voltage magnitude
        constraint_vmag_req = (dso_model.expected_interface_vmag_sqr - dso_model.v_sqr_req)
        obj += (dso_model.dual_v_sqr_req) * constraint_vmag_req
        obj += (dso_model.rho_v / 2) * (constraint_vmag_req ** 2)

        # Interface power flow
        constraint_p_req = (dso_model.expected_interface_pf_p - dso_model.p_pf_req) / interface_transf_rating
        constraint_q_req = (dso_model.expected_interface_pf_q - dso_model.q_pf_req) / interface_transf_rating
        obj += (dso_model.dual_pf_p_req) * constraint_p_req
        obj += (dso_model.dual_pf_q_req) * constraint_q_req
        obj += (dso_model.rho_pf / 2) * (constraint_p_req ** 2)
        obj += (dso_model.rho_pf / 2) * (constraint_q_req ** 2)

        # Shared ESS
        if consider_shared_ess:
            constraint_ess_p_req = (dso_model.expected_shared_ess_p - dso_model.p_ess_req) / (2 * shared_ess_rating)
            constraint_ess_q_req = (dso_model.expected_shared_ess_q - dso_model.q_ess_req) / (2 * shared_ess_rating)
            obj += (dso_model.dual_ess_p_req) * constraint_ess_p_req
            obj += (dso_model.dual_ess_q_req) * constraint_ess_q_req
            obj += (dso_model.rho_ess / 2) * constraint_ess_p_req ** 2
            obj += (dso_model.rho_ess / 2) * constraint_ess_q_req ** 2
            if params.previous_iter['ess']['dso']:
                constraint_ess_p_prev = (dso_model.expected_shared_ess_p - dso_model.p_ess_prev) / (2 * shared_ess_rating)
                constraint_ess_q_prev = (dso_model.expected_shared_ess_q - dso_model.q_ess_prev) / (2 * shared_ess_rating)
                obj += (dso_model.dual_ess_p_prev) * constraint_ess_p_prev
                obj += (dso_model.dual_ess_q_prev) * constraint_ess_q_prev
                obj += (dso_model.rho_ess_prev / 2) * constraint_ess_p_prev ** 2
                obj += (dso_model.rho_ess_prev / 2) * constraint_ess_q_prev ** 2

        # Add ADMM OF, deactivate original OF
        dso_model.objective.deactivate()
        dso_model.admm_objective = pe.Objective(sense=pe.minimize, expr=obj)


def update_transmission_coordination_model_and_solve(transmission_network, model, vsqr_req, dual_vsqr, pf_req, dual_pf, ess_req, dual_ess, params, consider_shared_ess, from_warm_start=False):

    print('[INFO] \t\t - Updating transmission network...')

    s_base = transmission_network.baseMVA

    rho_v = params.rho['v'][transmission_network.name]
    rho_pf = params.rho['pf'][transmission_network.name]
    if params.adaptive_penalty:
        rho_v = pe.value(model.rho_v) * (1 + ADMM_ADAPTIVE_PENALTY_FACTOR)
        rho_pf = pe.value(model.rho_pf) * (1 + ADMM_ADAPTIVE_PENALTY_FACTOR)

    if consider_shared_ess:
        rho_ess = params.rho['ess'][transmission_network.name]
        if params.adaptive_penalty:
            rho_ess = pe.value(model.rho_pf) * (1 + ADMM_ADAPTIVE_PENALTY_FACTOR)
        if params.previous_iter['ess']['tso']:
            rho_ess_prev = params.rho_previous_iter['ess'][transmission_network.name]
            if params.adaptive_penalty:
                rho_ess_prev = pe.value(model.rho_ess_prev) * (1 + ADMM_ADAPTIVE_PENALTY_FACTOR)

    # Update Rho parameter
    model.rho_v.fix(rho_v)
    model.rho_pf.fix(rho_pf)
    if consider_shared_ess:
        model.rho_ess.fix(rho_ess)
        if params.previous_iter['ess']['tso']:
            model.rho_ess_prev.fix(rho_ess_prev)

    for dn in model.active_distribution_networks:

        node_id = transmission_network.active_distribution_network_nodes[dn]
        v_base = transmission_network.get_node_base_kv(node_id)

        # Update VOLTAGE and POWER FLOW variables at connection point
        model.dual_v_sqr_req[dn].fix(dual_vsqr['current'][node_id] / (v_base ** 2))
        model.v_sqr_req[dn].fix(vsqr_req['dso']['current'][node_id] / (v_base ** 2))
        model.dual_pf_p_req[dn].fix(dual_pf['current'][node_id]['p'] / s_base)
        model.dual_pf_q_req[dn].fix(dual_pf['current'][node_id]['q'] / s_base)
        model.p_pf_req[dn].fix(pf_req['dso']['current'][node_id]['p'] / s_base)
        model.q_pf_req[dn].fix(pf_req['dso']['current'][node_id]['q'] / s_base)

        # Update shared ESS capacity and power requests
        if consider_shared_ess:
            shared_ess_idx = transmission_network.network.get_shared_energy_storage_idx(node_id)
            model.dual_ess_p_req[shared_ess_idx].fix(dual_ess['current'][node_id]['p'] / s_base)
            model.dual_ess_q_req[shared_ess_idx].fix(dual_ess['current'][node_id]['q'] / s_base)
            model.p_ess_req[shared_ess_idx].fix(ess_req['dso']['current'][node_id]['p'] / s_base)
            model.q_ess_req[shared_ess_idx].fix(ess_req['dso']['current'][node_id]['q'] / s_base)
            if params.previous_iter['ess']['tso']:
                model.dual_ess_p_prev[shared_ess_idx].fix(dual_ess['prev'][node_id]['p'] / s_base)
                model.dual_ess_q_prev[shared_ess_idx].fix(dual_ess['prev'][node_id]['q'] / s_base)
                model.p_ess_prev[shared_ess_idx].fix(ess_req['tso']['prev'][node_id]['p'] / s_base)
                model.q_ess_prev[shared_ess_idx].fix(ess_req['tso']['prev'][node_id]['q'] / s_base)

    # Solve!
    res = transmission_network.optimize(model, from_warm_start=from_warm_start)
    if res.solver.status == po.SolverStatus.error:
        print(f'[ERROR] Network {model.name} did not converge!')
        # exit(ERROR_NETWORK_OPTIMIZATION)

    return res


def update_distribution_coordination_models_and_solve(distribution_networks, models, vsqr_req, dual_vsqr, pf_req, dual_pf, ess_req, dual_ess, params, consider_shared_ess, from_warm_start=False):

    print('[INFO] \t\t - Updating distribution networks:')
    res = dict()

    for node_id in distribution_networks:

        model = models[node_id]
        distribution_network = distribution_networks[node_id]

        print(f'[INFO] \t\t\t - Updating active distribution network connected to node {node_id}...')

        ref_node_id = distribution_network.get_reference_node_id()
        v_base = distribution_network.get_node_base_kv(ref_node_id)
        s_base = distribution_network.baseMVA

        rho_v = params.rho['v'][distribution_network.name]
        rho_pf = params.rho['pf'][distribution_network.name]
        if consider_shared_ess:
            rho_ess = params.rho['ess'][distribution_network.name]
        if params.adaptive_penalty:
            rho_v = pe.value(model.rho_v) * (1 + ADMM_ADAPTIVE_PENALTY_FACTOR)
            rho_pf = pe.value(model.rho_pf) * (1 + ADMM_ADAPTIVE_PENALTY_FACTOR)
            if consider_shared_ess:
                rho_ess = pe.value(model.rho_ess) * (1 + ADMM_ADAPTIVE_PENALTY_FACTOR)
                if params.previous_iter['ess']['dso']:
                    rho_ess_prev = params.rho_previous_iter['ess'][distribution_network.name]
                    if params.adaptive_penalty:
                        rho_ess_prev = pe.value(model.rho_ess_prev) * (1 + ADMM_ADAPTIVE_PENALTY_FACTOR)

        # Update Rho parameter
        model.rho_v.fix(rho_v)
        model.rho_pf.fix(rho_pf)
        if consider_shared_ess:
            model.rho_ess.fix(rho_ess)
            if consider_shared_ess:
                if params.previous_iter['ess']['dso']:
                    model.rho_ess_prev.fix(rho_ess_prev)

        # Update VOLTAGE and POWER FLOW variables at connection point
        model.dual_v_sqr_req.fix(dual_vsqr['current'][node_id] / (v_base ** 2))
        model.v_sqr_req.fix(vsqr_req['tso']['current'][node_id] / (v_base ** 2))
        model.dual_pf_p_req.fix(dual_pf['current'][node_id]['p'] / s_base)
        model.dual_pf_q_req.fix(dual_pf['current'][node_id]['q'] / s_base)
        model.p_pf_req.fix(pf_req['tso']['current'][node_id]['p'] / s_base)
        model.q_pf_req.fix(pf_req['tso']['current'][node_id]['q'] / s_base)

        # Update SHARED ENERGY STORAGE variables (if existent)
        if consider_shared_ess:
            model.dual_ess_p_req.fix(dual_ess['current'][node_id]['p'] / s_base)
            model.dual_ess_q_req.fix(dual_ess['current'][node_id]['q'] / s_base)
            model.p_ess_req.fix(ess_req['esso']['current'][node_id]['p'] / s_base)
            model.q_ess_req.fix(ess_req['esso']['current'][node_id]['q'] / s_base)
            if params.previous_iter['ess']['dso']:
                model.dual_ess_p_prev.fix(dual_ess['prev'][node_id]['p'] / s_base)
                model.dual_ess_q_prev.fix(dual_ess['prev'][node_id]['q'] / s_base)
                model.p_ess_prev.fix(ess_req['dso']['prev'][node_id]['p'] / s_base)
                model.q_ess_prev.fix(ess_req['dso']['prev'][node_id]['q'] / s_base)

        # Solve!
        res[node_id] = distribution_network.optimize(model, from_warm_start=from_warm_start)
        if res[node_id].solver.status != po.SolverStatus.ok:
            print(f'[WARNING] Network {model.name} did not converge!')
            #exit(ERROR_NETWORK_OPTIMIZATION)

    return res


def create_admm_variables(operational_planning):

    consensus_variables = {
        'v_sqr': {'tso': {'current': dict(), 'prev': dict()},
                  'dso': {'current': dict(), 'prev': dict()}},
        'pf': {'tso': {'current': dict(), 'prev': dict()},
               'dso': {'current': dict(), 'prev': dict()}},
        'ess': {'tso': {'current': dict(), 'prev': dict()},
                'dso': {'current': dict(), 'prev': dict()},
                'esso': {'current': dict(), 'prev': dict()}}
    }

    dual_variables = {
        'v_sqr': {'tso': {'current': dict()}, 'dso': {'current': dict()}},
        'pf': {'tso': {'current': dict()}, 'dso': {'current': dict()}},
        'ess': {'tso': {'current': dict()}, 'dso': {'current': dict()}, 'esso': {'current': dict()}}
    }

    if operational_planning.params.previous_iter['ess']:
        dual_variables['ess']['tso']['prev'] = dict()
        dual_variables['ess']['dso']['prev'] = dict()

    for dn in range(len(operational_planning.active_distribution_network_nodes)):

        node_id = operational_planning.active_distribution_network_nodes[dn]
        node_base_kv = operational_planning.transmission_network.get_node_base_kv(node_id)

        consensus_variables['v_sqr']['tso']['current'][node_id] = node_base_kv
        consensus_variables['v_sqr']['dso']['current'][node_id] = node_base_kv
        consensus_variables['pf']['tso']['current'][node_id] = {'p': 0.00, 'q': 0.00}
        consensus_variables['pf']['dso']['current'][node_id] = {'p': 0.00, 'q': 0.00}
        consensus_variables['ess']['tso']['current'][node_id] = {'p': 0.00, 'q': 0.00}
        consensus_variables['ess']['dso']['current'][node_id] = {'p': 0.00, 'q': 0.00}
        consensus_variables['ess']['esso']['current'][node_id] = {'p': 0.00, 'q': 0.00}

        consensus_variables['v_sqr']['tso']['prev'][node_id] = node_base_kv
        consensus_variables['v_sqr']['dso']['prev'][node_id] = node_base_kv
        consensus_variables['pf']['tso']['prev'][node_id] = {'p': 0.00, 'q': 0.00}
        consensus_variables['pf']['dso']['prev'][node_id] = {'p': 0.00, 'q': 0.00}
        consensus_variables['ess']['tso']['prev'][node_id] = {'p': 0.00, 'q': 0.00}
        consensus_variables['ess']['dso']['prev'][node_id] = {'p': 0.00, 'q': 0.00}
        consensus_variables['ess']['esso']['prev'][node_id] = {'p': 0.00, 'q': 0.00}

        dual_variables['v_sqr']['tso']['current'][node_id] = 0.00
        dual_variables['v_sqr']['dso']['current'][node_id] = 0.00
        dual_variables['pf']['tso']['current'][node_id] = {'p': 0.00, 'q': 0.00}
        dual_variables['pf']['dso']['current'][node_id] = {'p': 0.00, 'q': 0.00}
        dual_variables['ess']['tso']['current'][node_id] = {'p': 0.00, 'q': 0.00}
        dual_variables['ess']['dso']['current'][node_id] = {'p': 0.00, 'q': 0.00}
        dual_variables['ess']['esso']['current'][node_id] = {'p': 0.00, 'q': 0.00}

        if operational_planning.params.previous_iter['ess']:
            dual_variables['ess']['tso']['prev'][node_id] = {'p': 0.00, 'q': 0.00}
            dual_variables['ess']['dso']['prev'][node_id] = {'p': 0.00, 'q': 0.00}

    return consensus_variables, dual_variables


def _update_interface_power_flow_variables(operational_planning, tso_model, dso_models, interface_vars, dual_vars, results, params, update_tn=True, update_dns=True):

    transmission_network = operational_planning.transmission_network
    distribution_networks = operational_planning.distribution_networks

    # Transmission network - Update Vmag and PF at the TN-DN interface
    if update_tn:
        for dn in range(len(operational_planning.active_distribution_network_nodes)):
            node_id = operational_planning.active_distribution_network_nodes[dn]
            v_base = transmission_network.get_node_base_kv(node_id)
            s_base = transmission_network.baseMVA
            if results['tso'].solver.status == po.SolverStatus.ok:
                interface_vars['v_sqr']['tso']['prev'][node_id] = copy(interface_vars['v_sqr']['tso']['current'][node_id])
                interface_vars['pf']['tso']['prev'][node_id]['p'] = copy(interface_vars['pf']['tso']['current'][node_id]['p'])
                interface_vars['pf']['tso']['prev'][node_id]['q'] = copy(interface_vars['pf']['tso']['current'][node_id]['q'])

                vsqr_req = pe.value(tso_model.expected_interface_vmag_sqr[dn]) * (v_base ** 2)
                p_req = pe.value(tso_model.expected_interface_pf_p[dn]) * s_base
                q_req = pe.value(tso_model.expected_interface_pf_q[dn]) * s_base
                interface_vars['v_sqr']['tso']['current'][node_id] = vsqr_req
                interface_vars['pf']['tso']['current'][node_id]['p'] = p_req
                interface_vars['pf']['tso']['current'][node_id]['q'] = q_req

    # Distribution Network - Update PF at the TN-DN interface
    if update_dns:
        for node_id in distribution_networks:
            distribution_network = distribution_networks[node_id]
            dso_model = dso_models[node_id]
            ref_node_id = distribution_network.get_reference_node_id()
            v_base = distribution_network.get_node_base_kv(ref_node_id)
            s_base = distribution_network.baseMVA
            if results['dso'][node_id].solver.status == po.SolverStatus.ok:
                interface_vars['v_sqr']['dso']['prev'][node_id] = copy(interface_vars['v_sqr']['dso']['current'][node_id])
                interface_vars['pf']['dso']['prev'][node_id]['p'] = copy(interface_vars['pf']['dso']['current'][node_id]['p'])
                interface_vars['pf']['dso']['prev'][node_id]['q'] = copy(interface_vars['pf']['dso']['current'][node_id]['q'])

                vsqr_req = pe.value(dso_model.expected_interface_vmag_sqr) * (v_base ** 2)
                p_req = pe.value(dso_model.expected_interface_pf_p) * s_base
                q_req = pe.value(dso_model.expected_interface_pf_q) * s_base
                interface_vars['v_sqr']['dso']['current'][node_id] = vsqr_req
                interface_vars['pf']['dso']['current'][node_id]['p'] = p_req
                interface_vars['pf']['dso']['current'][node_id]['q'] = q_req

    # Update Lambdas
    for node_id in distribution_networks:
        if update_tn:
            rho_v_tso = pe.value(tso_model.rho_v)
            rho_pf_tso = pe.value(tso_model.rho_pf)
            error_v_req_tso = interface_vars['v_sqr']['tso']['current'][node_id] - interface_vars['v_sqr']['dso']['current'][node_id]
            error_p_pf_req_tso = interface_vars['pf']['tso']['current'][node_id]['p'] - interface_vars['pf']['dso']['current'][node_id]['p']
            error_q_pf_req_tso = interface_vars['pf']['tso']['current'][node_id]['q'] - interface_vars['pf']['dso']['current'][node_id]['q']
            dual_vars['v_sqr']['tso']['current'][node_id] += rho_v_tso * error_v_req_tso
            dual_vars['pf']['tso']['current'][node_id]['p'] += rho_pf_tso * error_p_pf_req_tso
            dual_vars['pf']['tso']['current'][node_id]['q'] += rho_pf_tso * error_q_pf_req_tso

        if update_dns:
            rho_v_dso = pe.value(dso_models[node_id].rho_v)
            rho_pf_dso = pe.value(dso_models[node_id].rho_pf)
            error_v_req_dso = interface_vars['v_sqr']['dso']['current'][node_id] - interface_vars['v_sqr']['tso']['current'][node_id]
            error_p_pf_req_dso = interface_vars['pf']['dso']['current'][node_id]['p'] - interface_vars['pf']['tso']['current'][node_id]['p']
            error_q_pf_req_dso = interface_vars['pf']['dso']['current'][node_id]['q'] - interface_vars['pf']['tso']['current'][node_id]['q']
            dual_vars['v_sqr']['dso']['current'][node_id] += rho_v_dso * error_v_req_dso
            dual_vars['pf']['dso']['current'][node_id]['p'] += rho_pf_dso * error_p_pf_req_dso
            dual_vars['pf']['dso']['current'][node_id]['q'] += rho_pf_dso * error_q_pf_req_dso


def _update_shared_energy_storage_variables(operational_planning, tso_model, dso_models, shared_ess_vars, dual_vars, results, params, update_tn=True, update_dns=True):

    transmission_network = operational_planning.transmission_network
    distribution_networks = operational_planning.distribution_networks

    for node_id in operational_planning.active_distribution_network_nodes:

        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]

        # Power requested by TSO
        if update_tn:
            if results['tso'].solver.status == po.SolverStatus.ok:
                s_base = transmission_network.baseMVA
                shared_ess_idx = transmission_network.get_shared_energy_storage_idx(node_id)

                shared_ess_vars['tso']['prev'][node_id]['p'] = copy(shared_ess_vars['tso']['current'][node_id]['p'])
                shared_ess_vars['tso']['prev'][node_id]['q'] = copy(shared_ess_vars['tso']['current'][node_id]['q'])

                p_req = pe.value(tso_model.expected_shared_ess_p[shared_ess_idx]) * s_base
                q_req = pe.value(tso_model.expected_shared_ess_q[shared_ess_idx]) * s_base
                shared_ess_vars['tso']['current'][node_id]['p'] = p_req
                shared_ess_vars['tso']['current'][node_id]['q'] = q_req

        # Power requested by DSO
        if update_dns:
            if results['dso'][node_id].solver.status == po.SolverStatus.ok:
                s_base = distribution_network.baseMVA
                shared_ess_vars['dso']['prev'][node_id]['p'] = copy(shared_ess_vars['dso']['current'][node_id]['p'])
                shared_ess_vars['dso']['prev'][node_id]['q'] = copy(shared_ess_vars['dso']['current'][node_id]['q'])

                p_req = pe.value(dso_model.expected_shared_ess_p) * s_base
                q_req = pe.value(dso_model.expected_shared_ess_q) * s_base
                shared_ess_vars['dso']['current'][node_id]['p'] = p_req
                shared_ess_vars['dso']['current'][node_id]['q'] = q_req

        # Update dual variables Shared ESS
        if update_tn:
            rho_ess_tso = pe.value(tso_model.rho_ess)
            error_p_tso_dso = shared_ess_vars['tso']['current'][node_id]['p'] - shared_ess_vars['dso']['current'][node_id]['p']
            error_q_tso_dso = shared_ess_vars['tso']['current'][node_id]['q'] - shared_ess_vars['dso']['current'][node_id]['q']
            dual_vars['tso']['current'][node_id]['p'] += rho_ess_tso * error_p_tso_dso
            dual_vars['tso']['current'][node_id]['q'] += rho_ess_tso * error_q_tso_dso
            if params.previous_iter['ess']['tso']:
                error_p_tso_prev = shared_ess_vars['tso']['current'][node_id]['p'] - shared_ess_vars['tso']['prev'][node_id]['p']
                error_q_tso_prev = shared_ess_vars['tso']['current'][node_id]['q'] - shared_ess_vars['tso']['prev'][node_id]['q']
                dual_vars['tso']['prev'][node_id]['p'] += rho_ess_tso * error_p_tso_prev
                dual_vars['tso']['prev'][node_id]['q'] += rho_ess_tso * error_q_tso_prev

        if update_dns:
            rho_ess_dso = pe.value(dso_models[node_id].rho_ess)
            error_p_dso_tso = shared_ess_vars['dso']['current'][node_id]['p'] - shared_ess_vars['tso']['current'][node_id]['p']
            error_q_dso_tso = shared_ess_vars['dso']['current'][node_id]['q'] - shared_ess_vars['tso']['current'][node_id]['q']
            dual_vars['dso']['current'][node_id]['p'] += rho_ess_dso * error_p_dso_tso
            dual_vars['dso']['current'][node_id]['q'] += rho_ess_dso * error_q_dso_tso
            if params.previous_iter['ess']['dso']:
                error_p_dso_prev = shared_ess_vars['dso']['current'][node_id]['p'] - shared_ess_vars['dso']['prev'][node_id]['p']
                error_q_dso_prev = shared_ess_vars['dso']['current'][node_id]['q'] - shared_ess_vars['dso']['prev'][node_id]['q']
                dual_vars['dso']['prev'][node_id]['p'] += rho_ess_dso * error_p_dso_prev
                dual_vars['dso']['prev'][node_id]['q'] += rho_ess_dso * error_q_dso_prev


def _get_primal_value(operational_planning, tso_model, dso_models):

    transmission_network = operational_planning.transmission_network
    distribution_networks = operational_planning.distribution_networks

    primal_value = 0.0
    primal_value += transmission_network.get_primal_value(tso_model)
    for node_id in distribution_networks:
        primal_value += distribution_networks[node_id].get_primal_value(dso_models[node_id])

    return primal_value


def check_admm_convergence(operational_planning, consensus_vars, params, consider_shared_ess, debug_flag=False):
    if check_consensus_convergence(operational_planning, consensus_vars, params, consider_shared_ess, debug_flag=debug_flag):
        if check_stationary_convergence(operational_planning, consensus_vars, params, consider_shared_ess):
            print(f'[INFO]\t\t - Converged!')
            return True
    return False


def check_consensus_convergence(operational_planning, consensus_vars, params, consider_shared_ess, debug_flag=False):

    sum_rel_abs_error_vmag, sum_rel_abs_error_pf, sum_rel_abs_error_ess = 0.00, 0.00, 0.00
    num_elems_vmag, num_elems_pf, num_elems_ess = 0, 0, 0
    for node_id in operational_planning.active_distribution_network_nodes:

        s_base = operational_planning.transmission_network.baseMVA

        interface_v_base = operational_planning.transmission_network.get_node_base_kv(node_id)
        interface_transf_rating = operational_planning.distribution_networks[node_id].get_interface_branch_rating()
        if consider_shared_ess:
            shared_ess_idx = operational_planning.transmission_network.get_shared_energy_storage_idx(node_id)
            shared_ess_rating = abs(operational_planning.transmission_network.network.shared_energy_storages[shared_ess_idx].s) * s_base
            if isclose(shared_ess_rating, 0.00, abs_tol=SMALL_TOLERANCE):
                shared_ess_rating = 1.00

        sum_rel_abs_error_vmag += abs(sqrt(consensus_vars['v_sqr']['tso']['current'][node_id]) - sqrt(consensus_vars['v_sqr']['dso']['current'][node_id])) / interface_v_base
        num_elems_vmag += 2

        sum_rel_abs_error_pf += abs(consensus_vars['pf']['tso']['current'][node_id]['p'] - consensus_vars['pf']['dso']['current'][node_id]['p']) / interface_transf_rating
        sum_rel_abs_error_pf += abs(consensus_vars['pf']['tso']['current'][node_id]['q'] - consensus_vars['pf']['dso']['current'][node_id]['q']) / interface_transf_rating
        num_elems_pf += 4

        if consider_shared_ess:
            sum_rel_abs_error_ess += abs(consensus_vars['ess']['tso']['current'][node_id]['p'] - consensus_vars['ess']['dso']['current'][node_id]['p']) / shared_ess_rating
            sum_rel_abs_error_ess += abs(consensus_vars['ess']['tso']['current'][node_id]['q'] - consensus_vars['ess']['dso']['current'][node_id]['q']) / shared_ess_rating
            num_elems_ess += 4

    convergence = True
    if error_within_limits(sum_rel_abs_error_vmag, num_elems_vmag, params.tol['consensus']['v']):
        if error_within_limits(sum_rel_abs_error_pf, num_elems_pf, params.tol['consensus']['pf']):
            if consider_shared_ess:
                if error_within_limits(sum_rel_abs_error_ess, num_elems_ess, params.tol['consensus']['ess']):
                    print('[INFO]\t\t - Consensus constraints ok!')
                else:
                    convergence = False
                    print('[INFO]\t\t - Convergence shared ESS consensus constraints failed. {:.3f} > {:.3f}'.format(sum_rel_abs_error_ess, params.tol['consensus']['ess'] * num_elems_ess))
                    if debug_flag:
                        print_debug_info(operational_planning, consensus_vars, print_ess=True)
        else:
            convergence = False
            print('[INFO]\t\t - Convergence interface PF consensus constraints failed. {:.3f} > {:.3f}'.format(sum_rel_abs_error_pf, params.tol['consensus']['pf'] * num_elems_pf))
            if debug_flag:
                print_debug_info(operational_planning, consensus_vars, print_pf=True)
    else:
        convergence = False
        print('[INFO]\t\t - Convergence interface Vmag consensus constraints failed. {:.3f} > {:.3f}'.format(sum_rel_abs_error_vmag, params.tol['consensus']['v'] * num_elems_vmag))
        if debug_flag:
            print_debug_info(operational_planning, consensus_vars, print_vmag=True)

    return convergence


def check_stationary_convergence(operational_planning, consensus_vars, params, consider_shared_ess):

    rho_tso_v = params.rho['v'][operational_planning.transmission_network.name]
    rho_tso_pf = params.rho['pf'][operational_planning.transmission_network.name]
    rho_tso_ess = params.rho['ess'][operational_planning.transmission_network.name]

    sum_rel_abs_error_vmag, sum_rel_abs_error_pf, sum_rel_abs_error_ess = 0.00, 0.00, 0.00
    num_elems_vmag, num_elems_pf, num_elems_ess = 0, 0, 0
    for node_id in operational_planning.distribution_networks:

        rho_dso_v = params.rho['v'][operational_planning.distribution_networks[node_id].name]
        rho_dso_pf = params.rho['pf'][operational_planning.distribution_networks[node_id].name]
        rho_dso_ess = params.rho['ess'][operational_planning.distribution_networks[node_id].name]

        s_base = operational_planning.transmission_network.baseMVA
        interface_v_base = operational_planning.transmission_network.get_node_base_kv(node_id)
        interface_transf_rating = operational_planning.distribution_networks[node_id].get_interface_branch_rating()
        if consider_shared_ess:
            shared_ess_idx = operational_planning.transmission_network.get_shared_energy_storage_idx(node_id)
            shared_ess_rating = abs(operational_planning.transmission_network.shared_energy_storages[shared_ess_idx].s) * s_base
            if isclose(shared_ess_rating, 0.00, abs_tol=SMALL_TOLERANCE):
                shared_ess_rating = 1.00

        sum_rel_abs_error_vmag += rho_tso_v * abs(sqrt(consensus_vars['v_sqr']['tso']['current'][node_id]) - sqrt(consensus_vars['v_sqr']['tso']['prev'][node_id])) / interface_v_base
        sum_rel_abs_error_vmag += rho_dso_v * abs(sqrt(consensus_vars['v_sqr']['dso']['current'][node_id]) - sqrt(consensus_vars['v_sqr']['dso']['prev'][node_id])) / interface_v_base
        num_elems_vmag += 2

        sum_rel_abs_error_pf += rho_tso_pf * abs(consensus_vars['pf']['tso']['current'][node_id]['p'] - consensus_vars['pf']['tso']['prev'][node_id]['p']) / interface_transf_rating
        sum_rel_abs_error_pf += rho_tso_pf * abs(consensus_vars['pf']['tso']['current'][node_id]['q'] - consensus_vars['pf']['tso']['prev'][node_id]['q']) / interface_transf_rating
        sum_rel_abs_error_pf += rho_dso_pf * abs(consensus_vars['pf']['dso']['current'][node_id]['p'] - consensus_vars['pf']['dso']['prev'][node_id]['p']) / interface_transf_rating
        sum_rel_abs_error_pf += rho_dso_pf * abs(consensus_vars['pf']['dso']['current'][node_id]['q'] - consensus_vars['pf']['dso']['prev'][node_id]['q']) / interface_transf_rating
        num_elems_pf += 4

        if consider_shared_ess:
            sum_rel_abs_error_ess += rho_tso_ess * abs(consensus_vars['ess']['tso']['current'][node_id]['p'] - consensus_vars['ess']['tso']['prev'][node_id]['p']) / shared_ess_rating
            sum_rel_abs_error_ess += rho_tso_ess * abs(consensus_vars['ess']['tso']['current'][node_id]['q'] - consensus_vars['ess']['tso']['prev'][node_id]['q']) / shared_ess_rating
            sum_rel_abs_error_ess += rho_dso_ess * abs(consensus_vars['ess']['dso']['current'][node_id]['p'] - consensus_vars['ess']['dso']['prev'][node_id]['p']) / shared_ess_rating
            sum_rel_abs_error_ess += rho_dso_ess * abs(consensus_vars['ess']['dso']['current'][node_id]['q'] - consensus_vars['ess']['dso']['prev'][node_id]['q']) / shared_ess_rating
            num_elems_ess += 4

    convergence = True
    if error_within_limits(sum_rel_abs_error_vmag, num_elems_vmag, params.tol['stationarity']['v']):
        if error_within_limits(sum_rel_abs_error_pf, num_elems_pf, params.tol['stationarity']['pf']):
            if consider_shared_ess:
                if error_within_limits(sum_rel_abs_error_ess, num_elems_ess, params.tol['stationarity']['ess']):
                    print('[INFO]\t\t - Stationary constraints ok!')
                else:
                    convergence = False
                    print('[INFO]\t\t - Convergence shared ESS stationary constraints failed. {:.3f} > {:.3f}'.format(sum_rel_abs_error_ess, params.tol['stationarity']['ess'] * num_elems_ess))
        else:
            convergence = False
            print('[INFO]\t\t - Convergence interface PF stationary constraints failed. {:.3f} > {:.3f}'.format(sum_rel_abs_error_pf, params.tol['stationarity']['pf'] * num_elems_pf))
    else:
        convergence = False
        print('[INFO]\t\t - Convergence interface Vmag stationary constraints failed. {:.3f} > {:.3f}'.format(sum_rel_abs_error_vmag, params.tol['stationarity']['v'] * num_elems_vmag))

    return convergence


def error_within_limits(sum_abs_error, num_elems, tol):
    if sum_abs_error > tol * num_elems:
        if not isclose(sum_abs_error, tol * num_elems, rel_tol=ADMM_CONVERGENCE_REL_TOL, abs_tol=ADMM_CONVERGENCE_ABS_TOL):
            return False
    return True


# ======================================================================================================================
#  RESULTS PROCESSING functions
# ======================================================================================================================
def _process_operational_planning_results(operational_planning, tso_model, dso_models, optimization_results, t):

    transmission_network = operational_planning.transmission_network
    distribution_networks = operational_planning.distribution_networks

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()
    processed_results['interface'] = dict()
    processed_results['summary_detail'] = dict()

    processed_results['tso'] = transmission_network.process_results(tso_model, t, optimization_results['tso'])
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results(dso_model, t, optimization_results['dso'][node_id])
    processed_results['interface'] = _process_results_interface(operational_planning, tso_model, dso_models)
    processed_results['summary_detail'] = _process_results_summary_detail(operational_planning, tso_model, dso_models, t)

    return processed_results


def _process_results_interface(operational_planning, tso_model, dso_models):

    transmission_network = operational_planning.transmission_network
    distribution_networks = operational_planning.distribution_networks

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()

    processed_results['tso'] = transmission_network.process_results_interface(tso_model)
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results_interface(dso_model)

    return processed_results


def _process_results_summary_detail(operational_planning, tso_model, dso_models, t):

    transmission_network = operational_planning.transmission_network
    distribution_networks = operational_planning.distribution_networks

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()

    processed_results['tso'] = transmission_network.process_results_summary_detail(tso_model, t)
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results_summary_detail(dso_model, t)

    return processed_results


# ======================================================================================================================
#  RESULTS OPERATIONAL PLANNING - write functions
# ======================================================================================================================
def _write_operational_planning_results_to_excel(operational_planning, results, t, primal_evolution=list(), consider_shared_ess=False, filename='operation_planning'):

    wb = Workbook()

    _write_operational_planning_main_info_to_excel(operational_planning, wb, results, t)
    _write_operational_planning_main_info_to_excel_detailed(operational_planning, wb, results['summary_detail'], t)

    # Primal evolution
    if primal_evolution:
        _write_objective_function_evolution_to_excel(wb, primal_evolution)

    # Interface Power Flow
    _write_interface_results_to_excel(operational_planning, wb, t, results['interface'])

    # Shared Energy Storages results
    if consider_shared_ess:
        _write_shared_energy_storages_results_to_excel(operational_planning, wb, t, results)

    #  TSO and DSOs' results
    _write_network_voltage_results_to_excel(operational_planning, wb, t, results)
    _write_network_consumption_results_to_excel(operational_planning, wb, t, results)
    _write_network_generation_results_to_excel(operational_planning, wb, t, results)
    _write_network_branch_results_to_excel(operational_planning, wb, t, results, 'losses')
    _write_network_branch_results_to_excel(operational_planning, wb, t, results, 'ratio')
    _write_network_branch_loading_results_to_excel(operational_planning, wb, t, results)
    _write_network_branch_power_flow_results_to_excel(operational_planning, wb, t, results)
    _write_network_energy_storages_results_to_excel(operational_planning, wb, t, results)
    _write_relaxation_slacks_results_to_excel(operational_planning, wb, t, results)

    # Save results
    results_filename = os.path.join(operational_planning.results_dir, f'{filename}_operational_planning_results.xlsx')
    try:
        wb.save(results_filename)
        print('[INFO] Operational Planning Results written to {}.'.format(results_filename))
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(operational_planning.results_dir, f"{filename.replace('.xlsx', '')}_{current_time}.xlsx")
        print(f"[WARNING] Results saved to file {backup_filename}.xlsx")
        wb.save(backup_filename)


def _write_operational_planning_main_info_to_excel(operational_planning, workbook, results, t):

    sheet = workbook.worksheets[0]
    sheet.title = 'Main Info'

    # Write Header
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'SO'
    sheet.cell(row=line_idx, column=2).value = 'Node ID'
    sheet.cell(row=line_idx, column=3).value = 'Value'
    sheet.cell(row=line_idx, column=4).value = t

    # TSO
    line_idx = _write_operational_planning_main_info_per_operator(operational_planning.transmission_network, sheet, 'TSO', line_idx, results['tso'])

    # DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        line_idx = _write_operational_planning_main_info_per_operator(distribution_network, sheet, 'DSO', line_idx, dso_results, tn_node_id=tn_node_id)


def _write_operational_planning_main_info_per_operator(network, sheet, operator_type, line_idx, results, tn_node_id='-'):

    decimal_style = '0.00'

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id

    # - Objective
    obj_string = 'Objective'
    if network.params.obj_type == OBJ_MIN_COST:
        obj_string += ' (cost), [€]'
    elif network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
        obj_string += ' (congestion management)'
    sheet.cell(row=line_idx, column=3).value = obj_string
    sheet.cell(row=line_idx, column=4).value = results['obj']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    # Total Load
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Load, [MWh]'
    sheet.cell(row=line_idx, column=4).value = results['total_load']['p']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Load, [MVArh]'
    sheet.cell(row=line_idx, column=4).value = results['total_load']['q']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    # Flexibility used
    if network.params.fl_reg:
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = operator_type
        sheet.cell(row=line_idx, column=2).value = tn_node_id
        sheet.cell(row=line_idx, column=3).value = 'Flexibility used, [MWh]'
        sheet.cell(row=line_idx, column=4).value = results['flex_used']['p']
        sheet.cell(row=line_idx, column=4).number_format = decimal_style

        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = operator_type
        sheet.cell(row=line_idx, column=2).value = tn_node_id
        sheet.cell(row=line_idx, column=3).value = 'Flexibility used, [MVArh]'
        sheet.cell(row=line_idx, column=4).value = results['flex_used']['q']
        sheet.cell(row=line_idx, column=4).number_format = decimal_style

    # Total Load curtailed
    if network.params.l_curt:
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = operator_type
        sheet.cell(row=line_idx, column=2).value = tn_node_id
        sheet.cell(row=line_idx, column=3).value = 'Load curtailed, [MWh]'
        sheet.cell(row=line_idx, column=4).value = results['load_curt']['p']
        sheet.cell(row=line_idx, column=4).number_format = decimal_style

        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = operator_type
        sheet.cell(row=line_idx, column=2).value = tn_node_id
        sheet.cell(row=line_idx, column=3).value = 'Load curtailed, [MVArh]'
        sheet.cell(row=line_idx, column=4).value = results['load_curt']['q']
        sheet.cell(row=line_idx, column=4).number_format = decimal_style

    # Total Generation
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Generation, [MWh]'
    sheet.cell(row=line_idx, column=4).value = results['total_gen']['p']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Generation, [MVArh]'
    sheet.cell(row=line_idx, column=4).value = results['total_gen']['q']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    # Total Conventional Generation
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Conventional Generation, [MWh]'
    sheet.cell(row=line_idx, column=4).value = results['total_conventional_gen']['p']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Conventional Generation, [MVArh]'
    sheet.cell(row=line_idx, column=4).value = results['total_conventional_gen']['q']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    # Total Renewable Generation
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Renewable generation, [MWh]'
    sheet.cell(row=line_idx, column=4).value = results['total_renewable_gen']['p']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Renewable generation, [MVArh]'
    sheet.cell(row=line_idx, column=4).value = results['total_renewable_gen']['q']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Renewable generation, [MVAh]'
    sheet.cell(row=line_idx, column=4).value = results['total_renewable_gen']['s']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    # Renewable Generation Curtailed
    if network.params.rg_curt:
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = operator_type
        sheet.cell(row=line_idx, column=2).value = tn_node_id
        sheet.cell(row=line_idx, column=3).value = 'Renewable generation curtailed, [MVAh]'
        sheet.cell(row=line_idx, column=4).value = results['gen_curt']['s']
        sheet.cell(row=line_idx, column=4).number_format = decimal_style

    # Losses
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Losses, [MWh]'
    sheet.cell(row=line_idx, column=4).value = results['losses']
    sheet.cell(row=line_idx, column=4).number_format = decimal_style

    # Number of operation (generation and consumption) scenarios
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = operator_type
    sheet.cell(row=line_idx, column=2).value = tn_node_id
    sheet.cell(row=line_idx, column=3).value = 'Number of operation scenarios'
    sheet.cell(row=line_idx, column=4).value = len(network.prob_operation_scenarios)

    return line_idx


def _write_operational_planning_main_info_to_excel_detailed(operational_planning, workbook, results, t):

    sheet = workbook.create_sheet('Main Info, Detailed')

    # Write Header -- Year
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Operator'
    sheet.cell(row=line_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=line_idx, column=3).value = 'Operation Scenario'
    sheet.cell(row=line_idx, column=4).value = 'Probability, [%]'
    sheet.cell(row=line_idx, column=5).value = 'OF Value'
    sheet.cell(row=line_idx, column=6).value = 'Load, [MWh]'
    sheet.cell(row=line_idx, column=7).value = 'Load, [MVArh]'
    sheet.cell(row=line_idx, column=8).value = 'Flexibility used, [MWh]'
    sheet.cell(row=line_idx, column=9).value = 'Flexibility used, [MVArh]'
    sheet.cell(row=line_idx, column=10).value = 'Flexibility Cost, [€]'
    sheet.cell(row=line_idx, column=11).value = 'Generation, [MWh]'
    sheet.cell(row=line_idx, column=12).value = 'Generation, [MVArh]'
    sheet.cell(row=line_idx, column=13).value = 'Conventional Generation, [MWh]'
    sheet.cell(row=line_idx, column=14).value = 'Conventional Generation, [MVArh]'
    sheet.cell(row=line_idx, column=15).value = 'Conventional Generation Cost, [€]'
    sheet.cell(row=line_idx, column=16).value = 'Renewable Generation, [MWh]'
    sheet.cell(row=line_idx, column=17).value = 'Renewable Generation, [MVArh]'
    sheet.cell(row=line_idx, column=18).value = 'Renewable Generation, [MVAh]'
    sheet.cell(row=line_idx, column=19).value = 'Renewable Generation Curtailed, [MVAh]'
    sheet.cell(row=line_idx, column=20).value = 'Losses, [MWh]'

    # TSO
    line_idx += 1
    line_idx = _write_operational_planning_main_info_per_operator_detailed(sheet, 'TSO', line_idx, results['tso'])

    # DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        line_idx = _write_operational_planning_main_info_per_operator_detailed(sheet, 'DSO', line_idx, dso_results, tn_node_id=tn_node_id)


def _write_operational_planning_main_info_per_operator_detailed(sheet, operator_type, line_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    percent_style = '0.00%'

    sheet.cell(row=line_idx, column=1).value = operator_type
    for s_o in results['scenarios']:

        sheet.cell(row=line_idx, column=1).value = operator_type
        sheet.cell(row=line_idx, column=2).value = tn_node_id
        sheet.cell(row=line_idx, column=3).value = s_o

        # Probability, [%]
        sheet.cell(row=line_idx, column=4).value = results['scenarios'][s_o]['probability']
        sheet.cell(row=line_idx, column=4).number_format = percent_style

        # OF
        sheet.cell(row=line_idx, column=5).value = results['scenarios'][s_o]['obj']
        sheet.cell(row=line_idx, column=5).number_format = decimal_style

        # Load
        sheet.cell(row=line_idx, column=6).value = results['scenarios'][s_o]['load']['p']
        sheet.cell(row=line_idx, column=6).number_format = decimal_style
        sheet.cell(row=line_idx, column=7).value = results['scenarios'][s_o]['load']['q']
        sheet.cell(row=line_idx, column=7).number_format = decimal_style

        # Flexibility
        sheet.cell(row=line_idx, column=8).value = results['scenarios'][s_o]['flexibility']['p']
        sheet.cell(row=line_idx, column=8).number_format = decimal_style
        sheet.cell(row=line_idx, column=9).value = results['scenarios'][s_o]['flexibility']['q']
        sheet.cell(row=line_idx, column=9).number_format = decimal_style

        # Flexibility Cost, [€]
        sheet.cell(row=line_idx, column=10).value = results['scenarios'][s_o]['cost_flexibility']
        sheet.cell(row=line_idx, column=10).number_format = decimal_style

        # Generation
        sheet.cell(row=line_idx, column=11).value = results['scenarios'][s_o]['generation']['p']
        sheet.cell(row=line_idx, column=11).number_format = decimal_style
        sheet.cell(row=line_idx, column=12).value = results['scenarios'][s_o]['generation']['q']
        sheet.cell(row=line_idx, column=12).number_format = decimal_style

        # Conventional Generation
        sheet.cell(row=line_idx, column=13).value = results['scenarios'][s_o]['generation_conventional']['p']
        sheet.cell(row=line_idx, column=13).number_format = decimal_style
        sheet.cell(row=line_idx, column=14).value = results['scenarios'][s_o]['generation_conventional']['q']
        sheet.cell(row=line_idx, column=14).number_format = decimal_style

        # Conventional Generation Cost
        sheet.cell(row=line_idx, column=15).value = results['scenarios'][s_o]['generation_conventional_cost']
        sheet.cell(row=line_idx, column=15).number_format = decimal_style

        # Renewable Generation
        sheet.cell(row=line_idx, column=16).value = results['scenarios'][s_o]['generation_renewable']['p']
        sheet.cell(row=line_idx, column=16).number_format = decimal_style
        sheet.cell(row=line_idx, column=17).value = results['scenarios'][s_o]['generation_renewable']['q']
        sheet.cell(row=line_idx, column=17).number_format = decimal_style
        sheet.cell(row=line_idx, column=18).value = results['scenarios'][s_o]['generation_renewable']['s']
        sheet.cell(row=line_idx, column=18).number_format = decimal_style
        sheet.cell(row=line_idx, column=19).value = results['scenarios'][s_o]['generation_renewable_curtailed']['s']
        sheet.cell(row=line_idx, column=19).number_format = decimal_style

        # Losses
        sheet.cell(row=line_idx, column=20).value = results['scenarios'][s_o]['losses']
        sheet.cell(row=line_idx, column=20).number_format = decimal_style

        line_idx += 1

    return line_idx


def _write_objective_function_evolution_to_excel(workbook, primal_evolution):

    sheet = workbook.create_sheet('Primal Evolution')

    decimal_style = '0.000000'
    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Iteration'
    sheet.cell(row=row_idx, column=2).value = 'OF value'
    row_idx = row_idx + 1
    for i in range(len(primal_evolution)):
        sheet.cell(row=row_idx, column=1).value = i
        sheet.cell(row=row_idx, column=2).value = primal_evolution[i]
        sheet.cell(row=row_idx, column=2).number_format = decimal_style
        sheet.cell(row=row_idx, column=2).value = primal_evolution[i]
        sheet.cell(row=row_idx, column=2).number_format = decimal_style
        row_idx = row_idx + 1


def _write_interface_results_to_excel(operational_planning, workbook, t, results):

    sheet = workbook.create_sheet('Interface PF')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Operator'
    sheet.cell(row=row_idx, column=3).value = 'Quantity'
    sheet.cell(row=row_idx, column=4).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=5).value = t
    row_idx = row_idx + 1

    # TSO's results
    for node_id in results['tso']:

        expected_vmag = 0.00
        expected_p = 0.00
        expected_q = 0.00

        for s_o in results['tso'][node_id]:

            omega_s = operational_planning.transmission_network.prob_operation_scenarios[s_o]
            interface_vmag = results['tso'][node_id][s_o]['v']
            interface_p = results['tso'][node_id][s_o]['p']
            interface_q = results['tso'][node_id][s_o]['q']

            # Voltage magnitude
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'TSO'
            sheet.cell(row=row_idx, column=3).value = 'Vmag, [p.u.]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = interface_vmag
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            expected_vmag += interface_vmag * omega_s
            row_idx += 1

            # Active Power
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'TSO'
            sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = interface_p
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            expected_p += interface_p * omega_s
            row_idx += 1

            # Reactive Power
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'TSO'
            sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = interface_q
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            expected_q += interface_q * omega_s
            row_idx += 1

        # Expected Active Power
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'TSO'
        sheet.cell(row=row_idx, column=3).value = 'Vmag, [p.u.]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_vmag
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx += 1

        # Expected Active Power
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'TSO'
        sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_p
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx += 1

        # Expected Reactive Power
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'TSO'
        sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_q
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx += 1

    # DSOs' results
    for node_id in results['dso']:

        expected_vmag = 0.00
        expected_p = 0.00
        expected_q = 0.00

        for s_o in results['dso'][node_id]:

            omega_s = operational_planning.distribution_networks[node_id].prob_operation_scenarios[s_o]
            interface_vmag = results['dso'][node_id][s_o]['v']
            interface_p = results['dso'][node_id][s_o]['p']
            interface_q = results['dso'][node_id][s_o]['q']

            # Voltage magnitude
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'DSO'
            sheet.cell(row=row_idx, column=3).value = 'Vmag, [p.u.]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = interface_vmag
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            expected_vmag += interface_vmag * omega_s
            row_idx += 1

            # Active Power
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'DSO'
            sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = interface_p
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            expected_p += interface_p * omega_s
            row_idx += 1

            # Reactive Power
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'DSO'
            sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = interface_q
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            expected_q += interface_q * omega_s
            row_idx += 1

        # Expected Voltage magnitude
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'DSO'
        sheet.cell(row=row_idx, column=3).value = 'Vmag, [p.u.]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_vmag
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx += 1

        # Expected Active Power
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'DSO'
        sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_p
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx += 1

        # Expected Reactive Power
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'DSO'
        sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_q
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx += 1


def _write_shared_energy_storages_results_to_excel(operational_planning, workbook, t, results):

    sheet = workbook.create_sheet('Shared ESS')

    row_idx = 1
    decimal_style = '0.00'
    percent_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Operator'
    sheet.cell(row=row_idx, column=3).value = 'Quantity'
    sheet.cell(row=row_idx, column=4).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=5).value = t

    # TSO's results
    expected_p = dict()
    expected_q = dict()
    expected_s = dict()
    expected_soc = dict()
    expected_soc_percent = dict()
    for node_id in operational_planning.active_distribution_network_nodes:
        expected_p[node_id] = 0.00
        expected_q[node_id] = 0.00
        expected_s[node_id] = 0.00
        expected_soc[node_id] = 0.00
        expected_soc_percent[node_id] = 0.00

    for s_o in results['tso']['results']['scenarios']:

        omega_s = operational_planning.transmission_network.prob_operation_scenarios[s_o]

        for node_id in operational_planning.active_distribution_network_nodes:

            ess_p = results['tso']['results']['scenarios'][s_o]['shared_energy_storages']['p'][node_id]
            ess_q = results['tso']['results']['scenarios'][s_o]['shared_energy_storages']['q'][node_id]
            ess_s = results['tso']['results']['scenarios'][s_o]['shared_energy_storages']['s'][node_id]
            ess_soc = results['tso']['results']['scenarios'][s_o]['shared_energy_storages']['soc'][node_id]
            ess_soc_percent = results['tso']['results']['scenarios'][s_o]['shared_energy_storages']['soc_percent'][node_id]

            # Active power
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'TSO'
            sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_p
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if ess_p != 'N/A':
                expected_p[node_id] += ess_p * omega_s
            else:
                expected_p[node_id] = ess_p

            # Reactive power
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'TSO'
            sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_q
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if ess_q != 'N/A':
                expected_q[node_id] += ess_q * omega_s
            else:
                expected_q[node_id] = ess_q

            # Apparent power
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'TSO'
            sheet.cell(row=row_idx, column=3).value = 'S, [MVA]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_s
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if ess_s != 'N/A':
                expected_s[node_id] += ess_s * omega_s
            else:
                expected_s[node_id] = ess_s

            # State-of-Charge, [MVAh]
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'TSO'
            sheet.cell(row=row_idx, column=3).value = 'SoC, [MVAh]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_soc
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if ess_soc != 'N/A':
                expected_soc[node_id] += ess_soc * omega_s
            else:
                expected_soc[node_id] = ess_soc

            # State-of-Charge, [%]
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'TSO'
            sheet.cell(row=row_idx, column=3).value = 'SoC, [%]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_soc_percent
            sheet.cell(row=row_idx, column=5).number_format = percent_style
            if ess_soc_percent != 'N/A':
                expected_soc_percent[node_id] += ess_soc_percent * omega_s
            else:
                expected_soc_percent[node_id] = ess_soc_percent

    for node_id in operational_planning.active_distribution_network_nodes:

        # Active Power, [MW]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'TSO'
        sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_p[node_id]
        sheet.cell(row=row_idx, column=5).number_format = decimal_style

        # Reactive Power, [MVAr]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'TSO'
        sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_q[node_id]
        sheet.cell(row=row_idx, column=5).number_format = decimal_style

        # Apparent Power, [MVA]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'TSO'
        sheet.cell(row=row_idx, column=3).value = 'S, [MVA]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_s[node_id]
        sheet.cell(row=row_idx, column=5).number_format = decimal_style

        # State-of-Charge, [MVAh]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'TSO'
        sheet.cell(row=row_idx, column=3).value = 'SoC, [MVAh]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_soc[node_id]
        sheet.cell(row=row_idx, column=5).number_format = decimal_style

        # State-of-Charge, [%]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'TSO'
        sheet.cell(row=row_idx, column=3).value = 'SoC, [%]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_soc_percent[node_id]
        sheet.cell(row=row_idx, column=5).number_format = percent_style

    # DSO's results
    for node_id in results['dso']:

        distribution_network = operational_planning.distribution_networks[node_id]
        ref_node_id = distribution_network.get_reference_node_id()

        expected_p = 0.00
        expected_q = 0.00
        expected_s = 0.00
        expected_soc = 0.00
        expected_soc_percent = 0.00

        for s_o in results['dso'][node_id]['results']['scenarios']:

            omega_s = distribution_network.prob_operation_scenarios[s_o]
            ess_p = results['dso'][node_id]['results']['scenarios'][s_o]['shared_energy_storages']['p'][ref_node_id]
            ess_q = results['dso'][node_id]['results']['scenarios'][s_o]['shared_energy_storages']['q'][ref_node_id]
            ess_s = results['dso'][node_id]['results']['scenarios'][s_o]['shared_energy_storages']['s'][ref_node_id]
            ess_soc = results['dso'][node_id]['results']['scenarios'][s_o]['shared_energy_storages']['soc'][ref_node_id]
            ess_soc_percent = results['dso'][node_id]['results']['scenarios'][s_o]['shared_energy_storages']['soc_percent'][ref_node_id]

            # Active power
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'DSO'
            sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_p
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if ess_p != 'N/A':
                expected_p += ess_p * omega_s
            else:
                expected_p = ess_p

            # Reactive power
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'DSO'
            sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_q
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if ess_q != 'N/A':
                expected_q += ess_q * omega_s
            else:
                expected_q = ess_q

            # Apparent power
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'DSO'
            sheet.cell(row=row_idx, column=3).value = 'S, [MVA]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_s
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if ess_s != 'N/A':
                expected_s += ess_s * omega_s
            else:
                expected_s = ess_s

            # State-of-Charge, [MVAh]
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'DSO'
            sheet.cell(row=row_idx, column=3).value = 'SoC, [MVAh]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_soc
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if ess_soc != 'N/A':
                expected_soc += ess_soc * omega_s
            else:
                expected_soc = ess_soc

            # State-of-Charge, [%]
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'DSO'
            sheet.cell(row=row_idx, column=3).value = 'SoC, [%]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = ess_soc_percent
            sheet.cell(row=row_idx, column=5).number_format = percent_style
            if ess_soc_percent != 'N/A':
                expected_soc_percent += ess_soc_percent * omega_s
            else:
                expected_soc_percent = ess_soc_percent

        # Expected values

        # Active Power, [MW]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'DSO'
        sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_p
        sheet.cell(row=row_idx, column=5).number_format = decimal_style

        # Reactive Power, [MW]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'DSO'
        sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_q
        sheet.cell(row=row_idx, column=5).number_format = decimal_style

        # Apparent Power, [MW]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'DSO'
        sheet.cell(row=row_idx, column=3).value = 'S, [MVA]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_s
        sheet.cell(row=row_idx, column=5).number_format = decimal_style

        # State-of-Charge, [MVAh]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'DSO'
        sheet.cell(row=row_idx, column=3).value = 'SoC, [MVAh]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_soc
        sheet.cell(row=row_idx, column=5).number_format = decimal_style

        # State-of-Charge, [%]
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'DSO'
        sheet.cell(row=row_idx, column=3).value = 'SoC, [%]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_soc_percent
        sheet.cell(row=row_idx, column=5).number_format = percent_style


def _write_network_voltage_results_to_excel(operational_planning, workbook, t, results):

    sheet = workbook.create_sheet('Voltage')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Network Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=6).value = t
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = operational_planning.transmission_network
    row_idx = _write_network_voltage_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        row_idx = _write_network_voltage_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_voltage_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'

    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    ref_node_id = network.get_reference_node_id()
    expected_vmag = dict()
    expected_vang = dict()
    for node in network.nodes:
        expected_vmag[node.bus_i] = 0.00
        expected_vang[node.bus_i] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for node_id in results['scenarios'][s_o]['voltage']['vmag']:

            v_min, v_max = network.get_node_voltage_limits(node_id)
            v_mag = results['scenarios'][s_o]['voltage']['vmag'][node_id]
            v_ang = results['scenarios'][s_o]['voltage']['vang'][node_id]

            # Voltage magnitude
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = node_id
            sheet.cell(row=row_idx, column=4).value = 'Vmag, [p.u.]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = v_mag
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            if node_id != ref_node_id and (v_mag > v_max + SMALL_TOLERANCE or v_mag < v_min - SMALL_TOLERANCE):
                sheet.cell(row=row_idx, column=6).fill = violation_fill
            expected_vmag[node_id] += v_mag * omega_s
            row_idx = row_idx + 1

            # Voltage angle
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = node_id
            sheet.cell(row=row_idx, column=4).value = 'Vang, [º]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = v_ang
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            expected_vang[node_id] += v_ang * omega_s
            row_idx = row_idx + 1

    for node in network.nodes:

        node_id = node.bus_i
        v_min, v_max = network.get_node_voltage_limits(node_id)

        # Expected voltage magnitude
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = node_id
        sheet.cell(row=row_idx, column=4).value = 'Vmag, [p.u.]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_vmag[node_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        if node_id != ref_node_id and (expected_vmag[node_id] > v_max + SMALL_TOLERANCE or expected_vmag[node_id] < v_min - SMALL_TOLERANCE):
            sheet.cell(row=row_idx, column=6).fill = violation_fill
        row_idx = row_idx + 1

        # Expected voltage angle
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = node_id
        sheet.cell(row=row_idx, column=4).value = 'Vang, [º]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_vang[node_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        row_idx = row_idx + 1

    return row_idx


def _write_network_consumption_results_to_excel(operational_planning, workbook, t, results):

    sheet = workbook.create_sheet('Consumption')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Load ID'
    sheet.cell(row=row_idx, column=4).value = 'Node ID'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=7).value = t
    row_idx = row_idx + 1

    # Write results -- TSO
    tso_results = results['tso']
    transmission_network = operational_planning.transmission_network
    row_idx = _write_network_consumption_results_per_operator(transmission_network, sheet, 'TSO', row_idx, tso_results)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        row_idx = _write_network_consumption_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_consumption_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    expected_pc = dict()
    expected_flex_p = dict()
    expected_pc_curt = dict()
    expected_pnet = dict()
    expected_qc = dict()
    expected_flex_q = dict()
    expected_qc_curt = dict()
    expected_qnet = dict()
    for load in network.loads:
        expected_pc[load.load_id] = 0.00
        expected_flex_p[load.load_id] = 0.00
        expected_pc_curt[load.load_id] = 0.00
        expected_pnet[load.load_id] = 0.00
        expected_qc[load.load_id] = 0.00
        expected_flex_q[load.load_id] = 0.00
        expected_qc_curt[load.load_id] = 0.00
        expected_qnet[load.load_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for load in network.loads:

            pc = results['scenarios'][s_o]['consumption']['pc'][load.load_id]
            qc = results['scenarios'][s_o]['consumption']['qc'][load.load_id]

            # - Active Power
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = load.load_id
            sheet.cell(row=row_idx, column=4).value = load.bus
            sheet.cell(row=row_idx, column=5).value = 'Pc, [MW]'
            sheet.cell(row=row_idx, column=6).value = s_o
            sheet.cell(row=row_idx, column=7).value = pc
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            expected_pc[load.load_id] += pc * omega_s
            row_idx = row_idx + 1

            if network.params.fl_reg:

                pc_flex = results['scenarios'][s_o]['consumption']['pc_flex'][load.load_id]

                # - Flexibility, active
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = load.load_id
                sheet.cell(row=row_idx, column=4).value = load.bus
                sheet.cell(row=row_idx, column=5).value = 'Pc_flex, [MW]'
                sheet.cell(row=row_idx, column=6).value = s_o
                sheet.cell(row=row_idx, column=7).value = pc_flex
                sheet.cell(row=row_idx, column=7).number_format = decimal_style
                expected_flex_p[load.load_id] += pc_flex * omega_s
                row_idx = row_idx + 1

            if network.params.l_curt:

                pc_curt = results['scenarios'][s_o]['consumption']['pc_curt'][load.load_id]

                # - Active power curtailment
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = load.load_id
                sheet.cell(row=row_idx, column=4).value = load.bus
                sheet.cell(row=row_idx, column=5).value = 'Pc_curt, [MW]'
                sheet.cell(row=row_idx, column=6).value = s_o
                sheet.cell(row=row_idx, column=7).value = pc_curt
                sheet.cell(row=row_idx, column=7).number_format = decimal_style
                if pc_curt >= SMALL_TOLERANCE:
                    sheet.cell(row=row_idx, column=7).fill = violation_fill
                expected_pc_curt[load.load_id] += pc_curt * omega_s
                row_idx = row_idx + 1

            if network.params.fl_reg or network.params.l_curt:

                p_net = results['scenarios'][s_o]['consumption']['pc_net'][load.load_id]

                # - Active power net consumption
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = load.load_id
                sheet.cell(row=row_idx, column=4).value = load.bus
                sheet.cell(row=row_idx, column=5).value = 'Pc_net, [MW]'
                sheet.cell(row=row_idx, column=6).value = s_o
                sheet.cell(row=row_idx, column=7).value = p_net
                sheet.cell(row=row_idx, column=7).number_format = decimal_style
                expected_pnet[load.load_id] += p_net * omega_s
                row_idx = row_idx + 1

            # - Reactive power
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = load.load_id
            sheet.cell(row=row_idx, column=4).value = load.bus
            sheet.cell(row=row_idx, column=5).value = 'Qc, [MVAr]'
            sheet.cell(row=row_idx, column=6).value = s_o
            sheet.cell(row=row_idx, column=7).value = qc
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            expected_qc[load.load_id] += qc * omega_s
            row_idx = row_idx + 1

            if network.params.fl_reg:

                qc_flex = results['scenarios'][s_o]['consumption']['qc_flex'][load.load_id]

                # - Flexibility, active
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = load.load_id
                sheet.cell(row=row_idx, column=4).value = load.bus
                sheet.cell(row=row_idx, column=5).value = 'Qc_flex, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = s_o
                sheet.cell(row=row_idx, column=7).value = qc_flex
                sheet.cell(row=row_idx, column=7).number_format = decimal_style
                expected_flex_q[load.load_id] += qc_flex * omega_s
                row_idx = row_idx + 1

            if network.params.l_curt:

                qc_curt = results['scenarios'][s_o]['consumption']['qc_curt'][load.load_id]

                # - Active power curtailment
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = load.load_id
                sheet.cell(row=row_idx, column=4).value = load.bus
                sheet.cell(row=row_idx, column=5).value = 'Qc_curt, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = s_o
                sheet.cell(row=row_idx, column=7).value = qc_curt
                sheet.cell(row=row_idx, column=7).number_format = decimal_style
                if qc_curt >= SMALL_TOLERANCE:
                    sheet.cell(row=row_idx, column=7).fill = violation_fill
                expected_qc_curt[load.load_id] += qc_curt * omega_s
                row_idx = row_idx + 1

            if network.params.fl_reg or network.params.l_curt:

                q_net = results['scenarios'][s_o]['consumption']['qc_net'][load.load_id]

                # - Active power net consumption
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = load.load_id
                sheet.cell(row=row_idx, column=4).value = load.bus
                sheet.cell(row=row_idx, column=5).value = 'Qc_net, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = s_o
                sheet.cell(row=row_idx, column=7).value = q_net
                sheet.cell(row=row_idx, column=7).number_format = decimal_style
                expected_qnet[load.load_id] += q_net * omega_s
                row_idx = row_idx + 1

    for load in network.loads:

        # - Active Power
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = load.load_id
        sheet.cell(row=row_idx, column=4).value = load.bus
        sheet.cell(row=row_idx, column=5).value = 'Pc, [MW]'
        sheet.cell(row=row_idx, column=6).value = 'Expected'
        sheet.cell(row=row_idx, column=7).value = expected_pc[load.load_id]
        sheet.cell(row=row_idx, column=7).number_format = decimal_style
        row_idx = row_idx + 1

        if network.params.fl_reg:

            # - Flexibility, active
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = load.load_id
            sheet.cell(row=row_idx, column=4).value = load.bus
            sheet.cell(row=row_idx, column=5).value = 'Pc_flex, [MW]'
            sheet.cell(row=row_idx, column=6).value = 'Expected'
            sheet.cell(row=row_idx, column=7).value = expected_flex_p[load.load_id]
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            row_idx = row_idx + 1

        if network.params.l_curt:

            # - Load curtailment (active power)
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = load.load_id
            sheet.cell(row=row_idx, column=4).value = load.bus
            sheet.cell(row=row_idx, column=5).value = 'Pc_curt, [MW]'
            sheet.cell(row=row_idx, column=6).value = 'Expected'
            sheet.cell(row=row_idx, column=7).value = expected_pc_curt[load.load_id]
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            if expected_pc_curt[load.load_id] >= SMALL_TOLERANCE:
                sheet.cell(row=row_idx, column=7).fill = violation_fill
            row_idx = row_idx + 1

        if network.params.fl_reg or network.params.l_curt:

            # - Active power net consumption
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = load.load_id
            sheet.cell(row=row_idx, column=4).value = load.bus
            sheet.cell(row=row_idx, column=5).value = 'Pc_net, [MW]'
            sheet.cell(row=row_idx, column=6).value = 'Expected'
            sheet.cell(row=row_idx, column=7).value = expected_pnet[load.load_id]
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            row_idx = row_idx + 1

        # - Reactive power
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = load.load_id
        sheet.cell(row=row_idx, column=4).value = load.bus
        sheet.cell(row=row_idx, column=5).value = 'Qc, [MVAr]'
        sheet.cell(row=row_idx, column=6).value = 'Expected'
        sheet.cell(row=row_idx, column=7).value = expected_qc[load.load_id]
        sheet.cell(row=row_idx, column=7).number_format = decimal_style
        row_idx = row_idx + 1

        if network.params.fl_reg:

            # - Flexibility, reactive
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = load.load_id
            sheet.cell(row=row_idx, column=4).value = load.bus
            sheet.cell(row=row_idx, column=5).value = 'Qc_flex, [MVAr]'
            sheet.cell(row=row_idx, column=6).value = 'Expected'
            sheet.cell(row=row_idx, column=7).value = expected_flex_q[load.load_id]
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            row_idx = row_idx + 1

        if network.params.l_curt:

            # - Load curtailment (reactive power)
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = load.load_id
            sheet.cell(row=row_idx, column=4).value = load.bus
            sheet.cell(row=row_idx, column=5).value = 'Qc_curt, [MVAr]'
            sheet.cell(row=row_idx, column=6).value = 'Expected'
            sheet.cell(row=row_idx, column=7).value = expected_qc_curt[load.load_id]
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            if expected_qc_curt[load.load_id] >= SMALL_TOLERANCE:
                sheet.cell(row=row_idx, column=7).fill = violation_fill
            row_idx = row_idx + 1

        if network.params.fl_reg or network.params.l_curt:

            # - Reactive power net consumption
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = load.load_id
            sheet.cell(row=row_idx, column=4).value = load.bus
            sheet.cell(row=row_idx, column=5).value = 'Qc_net, [MVAr]'
            sheet.cell(row=row_idx, column=6).value = 'Expected'
            sheet.cell(row=row_idx, column=7).value = expected_qnet[load.load_id]
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            row_idx = row_idx + 1

    return row_idx


def _write_network_generation_results_to_excel(operational_planning, workbook, t, results):

    sheet = workbook.create_sheet('Generation')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Generator ID'
    sheet.cell(row=row_idx, column=4).value = 'Node ID'
    sheet.cell(row=row_idx, column=5).value = 'Type'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=8).value = t
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = operational_planning.transmission_network
    row_idx = _write_network_generation_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        row_idx = _write_network_generation_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_generation_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    expected_pg = dict()
    expected_pg_net = dict()
    expected_qg = dict()
    expected_qg_net = dict()
    expected_sg = dict()
    expected_sg_curt = dict()
    expected_sg_net = dict()
    for generator in network.generators:
        expected_pg[generator.gen_id] = 0.00
        expected_qg[generator.gen_id] = 0.00
        expected_sg[generator.gen_id] = 0.00
        expected_pg_net[generator.gen_id] = 0.00
        expected_qg_net[generator.gen_id] = 0.00
        expected_sg_curt[generator.gen_id] = 0.00
        expected_sg_net[generator.gen_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for generator in network.generators:

            node_id = generator.bus
            gen_id = generator.gen_id
            gen_type = network.get_gen_type(gen_id)
            pg = results['scenarios'][s_o]['generation']['pg'][gen_id]
            qg = results['scenarios'][s_o]['generation']['qg'][gen_id]

            # Active Power
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = gen_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = gen_type
            sheet.cell(row=row_idx, column=6).value = 'Pg, [MW]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = pg
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            expected_pg[gen_id] += pg * omega_s
            row_idx = row_idx + 1

            # Active Power net
            if generator.is_curtaillable() and network.params.rg_curt:

                pg_net = results['scenarios'][s_o]['generation']['pg_net'][gen_id]

                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = gen_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = gen_type
                sheet.cell(row=row_idx, column=6).value = 'Pg_net, [MW]'
                sheet.cell(row=row_idx, column=7).value = s_o
                sheet.cell(row=row_idx, column=8).value = pg_net
                sheet.cell(row=row_idx, column=8).number_format = decimal_style
                expected_pg_net[gen_id] += pg_net * omega_s
                row_idx = row_idx + 1

            # Reactive Power
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = gen_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = gen_type
            sheet.cell(row=row_idx, column=6).value = 'Qg, [MVAr]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = qg
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            expected_qg[gen_id] += qg * omega_s
            row_idx = row_idx + 1

            # Reactive Power net
            if generator.is_curtaillable() and network.params.rg_curt:

                qg_net = results['scenarios'][s_o]['generation']['qg_net'][gen_id]

                # Reactive Power net
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = gen_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = gen_type
                sheet.cell(row=row_idx, column=6).value = 'Qg_net, [MVAr]'
                sheet.cell(row=row_idx, column=7).value = s_o
                sheet.cell(row=row_idx, column=8).value = qg_net
                sheet.cell(row=row_idx, column=8).number_format = decimal_style
                expected_qg_net[gen_id] += qg_net * omega_s
                row_idx = row_idx + 1

            # Apparent Power
            if generator.is_curtaillable() and network.params.rg_curt:

                sg = results['scenarios'][s_o]['generation']['sg'][gen_id]
                sg_curt = results['scenarios'][s_o]['generation']['sg_curt'][gen_id]
                sg_net = results['scenarios'][s_o]['generation']['sg_net'][gen_id]

                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = gen_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = gen_type
                sheet.cell(row=row_idx, column=6).value = 'Sg, [MVA]'
                sheet.cell(row=row_idx, column=7).value = s_o
                sheet.cell(row=row_idx, column=8).value = sg
                sheet.cell(row=row_idx, column=8).number_format = decimal_style
                expected_sg[gen_id] += sg * omega_s
                row_idx = row_idx + 1

                # Apparent Power curtailment
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = gen_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = gen_type
                sheet.cell(row=row_idx, column=6).value = 'Sg_curt, [MVA]'
                sheet.cell(row=row_idx, column=7).value = s_o
                sheet.cell(row=row_idx, column=8).value = sg_curt
                sheet.cell(row=row_idx, column=8).number_format = decimal_style
                if not isclose(sg_curt, 0.00, abs_tol=VIOLATION_TOLERANCE):
                    sheet.cell(row=row_idx, column=8).fill = violation_fill
                expected_sg_curt[gen_id] += sg_curt * omega_s
                row_idx = row_idx + 1

                # Apparent Power Net
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = gen_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = gen_type
                sheet.cell(row=row_idx, column=6).value = 'Sg_net, [MVA]'
                sheet.cell(row=row_idx, column=7).value = s_o
                sheet.cell(row=row_idx, column=8).value = sg_net
                sheet.cell(row=row_idx, column=8).number_format = decimal_style
                expected_sg_net[gen_id] += sg_net * omega_s
                row_idx = row_idx + 1

    for generator in network.generators:

        node_id = generator.bus
        gen_id = generator.gen_id
        gen_type = network.get_gen_type(gen_id)

        # Active Power
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = gen_id
        sheet.cell(row=row_idx, column=4).value = node_id
        sheet.cell(row=row_idx, column=5).value = gen_type
        sheet.cell(row=row_idx, column=6).value = 'Pg, [MW]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = '-'
        sheet.cell(row=row_idx, column=8).value = expected_pg[gen_id]
        sheet.cell(row=row_idx, column=8).number_format = decimal_style
        row_idx = row_idx + 1

        # Active Power Net
        if generator.is_curtaillable() and network.params.rg_curt:

            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = gen_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = gen_type
            sheet.cell(row=row_idx, column=6).value = 'Pg_net, [MW]'
            sheet.cell(row=row_idx, column=7).value = 'Expected'
            sheet.cell(row=row_idx, column=8).value = expected_pg_net[gen_id]
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            row_idx = row_idx + 1

        # Reactive Power
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = gen_id
        sheet.cell(row=row_idx, column=4).value = node_id
        sheet.cell(row=row_idx, column=5).value = gen_type
        sheet.cell(row=row_idx, column=6).value = 'Qg, [MVAr]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = expected_qg[gen_id]
        sheet.cell(row=row_idx, column=8).number_format = decimal_style
        row_idx = row_idx + 1

        # Reactive Power net
        if generator.is_curtaillable() and network.params.rg_curt:

            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = gen_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = gen_type
            sheet.cell(row=row_idx, column=6).value = 'Qg_net, [MVAr]'
            sheet.cell(row=row_idx, column=7).value = 'Expected'
            sheet.cell(row=row_idx, column=8).value = expected_qg_net[gen_id]
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            row_idx = row_idx + 1

        # Apparent Power
        if generator.is_curtaillable() and network.params.rg_curt:

            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = gen_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = gen_type
            sheet.cell(row=row_idx, column=6).value = 'Sg, [MVA]'
            sheet.cell(row=row_idx, column=7).value = 'Expected'
            sheet.cell(row=row_idx, column=8).value = expected_sg[gen_id]
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            row_idx = row_idx + 1

            # Apparent Power curtailment
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = gen_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = gen_type
            sheet.cell(row=row_idx, column=6).value = 'Sg_curt, [MVA]'
            sheet.cell(row=row_idx, column=7).value = 'Expected'
            sheet.cell(row=row_idx, column=8).value = expected_sg_curt[gen_id]
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            if not isclose(expected_sg_curt[gen_id], 0.00, abs_tol=VIOLATION_TOLERANCE):
                sheet.cell(row=row_idx, column=8).fill = violation_fill
            row_idx = row_idx + 1

            # Apparent Power Net
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = gen_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = gen_type
            sheet.cell(row=row_idx, column=6).value = 'Sg_net, [MVA]'
            sheet.cell(row=row_idx, column=7).value = 'Expected'
            sheet.cell(row=row_idx, column=8).value = expected_sg_net[gen_id]
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            row_idx = row_idx + 1

    return row_idx


def _write_network_branch_results_to_excel(operational_planning, workbook, t, results, result_type):

    sheet_name = str()
    if result_type == 'losses':
        sheet_name = 'Branch Losses'
    elif result_type == 'ratio':
        sheet_name = 'Transformer Ratio'

    sheet = workbook.create_sheet(sheet_name)

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Branch ID'
    sheet.cell(row=row_idx, column=4).value = 'From Node ID'
    sheet.cell(row=row_idx, column=5).value = 'To Node ID'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=8).value = t
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = operational_planning.transmission_network
    row_idx = _write_network_branch_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso'], result_type)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        row_idx = _write_network_branch_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, result_type, tn_node_id=tn_node_id)


def _write_network_branch_results_per_operator(network, sheet, operator_type, row_idx, results, result_type, tn_node_id='-'):

    decimal_style = '0.00'

    aux_string = str()
    if result_type == 'losses':
        aux_string = 'P, [MW]'
    elif result_type == 'ratio':
        aux_string = 'Ratio'

    expected_values = dict()
    for branch in network.branches:
        expected_values[branch.branch_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for branch in network.branches:

            branch_id = branch.branch_id

            if not(result_type == 'ratio' and not branch.is_transformer):

                value = results['scenarios'][s_o]['branches'][result_type][branch_id]

                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = branch.tbus
                sheet.cell(row=row_idx, column=6).value = aux_string
                sheet.cell(row=row_idx, column=7).value = s_o
                sheet.cell(row=row_idx, column=8).value = value
                sheet.cell(row=row_idx, column=8).number_format = decimal_style
                expected_values[branch_id] += value * omega_s
                row_idx = row_idx + 1

    for branch in network.branches:

        branch_id = branch.branch_id

        if not (result_type == 'ratio' and not branch.is_transformer):

            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.fbus
            sheet.cell(row=row_idx, column=5).value = branch.tbus
            sheet.cell(row=row_idx, column=6).value = aux_string
            sheet.cell(row=row_idx, column=7).value = 'Expected'
            sheet.cell(row=row_idx, column=8).value = expected_values[branch_id]
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            row_idx = row_idx + 1

    return row_idx


def _write_network_branch_loading_results_to_excel(operational_planning, workbook, t, results):

    sheet = workbook.create_sheet('Branch Loading')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Branch ID'
    sheet.cell(row=row_idx, column=4).value = 'From Node ID'
    sheet.cell(row=row_idx, column=5).value = 'To Node ID'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=8).value = t
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = operational_planning.transmission_network
    row_idx = _write_network_branch_loading_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        row_idx = _write_network_branch_loading_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_branch_loading_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    perc_style = '0.00%'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    expected_values = {'flow_ij': {}}
    for branch in network.branches:
        expected_values['flow_ij'][branch.branch_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for branch in network.branches:

            value = results['scenarios'][s_o]['branches']['branch_flow']['flow_ij_perc'][branch.branch_id]

            # flow ij, [%]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.fbus
            sheet.cell(row=row_idx, column=5).value = branch.tbus
            sheet.cell(row=row_idx, column=6).value = 'Flow_ij, [%]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = value
            sheet.cell(row=row_idx, column=8).number_format = perc_style
            if value > 1.00 + VIOLATION_TOLERANCE:
                sheet.cell(row=row_idx, column=8).fill = violation_fill
            expected_values['flow_ij'][branch.branch_id] += value * omega_s
            row_idx = row_idx + 1

    for branch in network.branches:

        value = expected_values['flow_ij'][branch.branch_id]

        # flow ij, [%]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = branch.branch_id
        sheet.cell(row=row_idx, column=4).value = branch.fbus
        sheet.cell(row=row_idx, column=5).value = branch.tbus
        sheet.cell(row=row_idx, column=6).value = 'Flow_ij, [%]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = value
        sheet.cell(row=row_idx, column=8).number_format = perc_style
        if value > 1.00 + VIOLATION_TOLERANCE:
            sheet.cell(row=row_idx, column=8).fill = violation_fill
        row_idx = row_idx + 1

    return row_idx


def _write_network_branch_power_flow_results_to_excel(operational_planning, workbook, t, results):

    sheet = workbook.create_sheet('Power Flows')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Branch ID'
    sheet.cell(row=row_idx, column=4).value = 'From Node ID'
    sheet.cell(row=row_idx, column=5).value = 'To Node ID'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=8).value = t
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = operational_planning.transmission_network
    row_idx = _write_network_power_flow_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        row_idx = _write_network_power_flow_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_power_flow_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    perc_style = '0.00%'

    expected_values = {'pij': {}, 'pji': {}, 'qij': {}, 'qji': {}, 'sij': {}, 'sji': {}}
    for branch in network.branches:
        expected_values['pij'][branch.branch_id] = 0.00
        expected_values['pji'][branch.branch_id] = 0.00
        expected_values['qij'][branch.branch_id] = 0.00
        expected_values['qji'][branch.branch_id] = 0.00
        expected_values['sij'][branch.branch_id] = 0.00
        expected_values['sji'][branch.branch_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for branch in network.branches:

            branch_id = branch.branch_id
            rating = branch.rate
            if rating == 0.0:
                rating = BRANCH_UNKNOWN_RATING

            pij = results['scenarios'][s_o]['branches']['power_flow']['pij'][branch_id]
            pji = results['scenarios'][s_o]['branches']['power_flow']['pji'][branch_id]
            qij = results['scenarios'][s_o]['branches']['power_flow']['qij'][branch_id]
            qji = results['scenarios'][s_o]['branches']['power_flow']['qji'][branch_id]
            sij = results['scenarios'][s_o]['branches']['power_flow']['sij'][branch_id]
            sji = results['scenarios'][s_o]['branches']['power_flow']['sji'][branch_id]

            # Pij, [MW]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.fbus
            sheet.cell(row=row_idx, column=5).value = branch.tbus
            sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = pij
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            expected_values['pij'][branch_id] += pij * omega_s
            row_idx = row_idx + 1

            # Pji, [MW]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.tbus
            sheet.cell(row=row_idx, column=5).value = branch.fbus
            sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = pji
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            expected_values['pji'][branch_id] += pji * omega_s
            row_idx = row_idx + 1

            # Qij, [MVAr]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.fbus
            sheet.cell(row=row_idx, column=5).value = branch.tbus
            sheet.cell(row=row_idx, column=6).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = qij
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            expected_values['qij'][branch_id] += qij * omega_s
            row_idx = row_idx + 1

            # Qji, [MW]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.tbus
            sheet.cell(row=row_idx, column=5).value = branch.fbus
            sheet.cell(row=row_idx, column=6).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = qji
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            expected_values['qji'][branch_id] += qji * omega_s
            row_idx = row_idx + 1

            # Sij, [MVA]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.fbus
            sheet.cell(row=row_idx, column=5).value = branch.tbus
            sheet.cell(row=row_idx, column=6).value = 'S, [MVA]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = sij
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            expected_values['sij'][branch_id] += sij * omega_s
            row_idx = row_idx + 1

            # Sij, [%]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.fbus
            sheet.cell(row=row_idx, column=5).value = branch.tbus
            sheet.cell(row=row_idx, column=6).value = 'S, [%]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = sij / rating
            sheet.cell(row=row_idx, column=8).number_format = perc_style
            row_idx = row_idx + 1

            # Sji, [MW]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.tbus
            sheet.cell(row=row_idx, column=5).value = branch.fbus
            sheet.cell(row=row_idx, column=6).value = 'S, [MVA]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = sji
            sheet.cell(row=row_idx, column=8).number_format = decimal_style
            expected_values['sji'][branch_id] += sji * omega_s
            row_idx = row_idx + 1

            # Sji, [%]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = branch.branch_id
            sheet.cell(row=row_idx, column=4).value = branch.tbus
            sheet.cell(row=row_idx, column=5).value = branch.fbus
            sheet.cell(row=row_idx, column=6).value = 'S, [%]'
            sheet.cell(row=row_idx, column=7).value = s_o
            sheet.cell(row=row_idx, column=8).value = sji / rating
            sheet.cell(row=row_idx, column=8).number_format = perc_style
            row_idx = row_idx + 1

    for branch in network.branches:

        branch_id = branch.branch_id
        rating = branch.rate
        if rating == 0.0:
            rating = BRANCH_UNKNOWN_RATING

        # Pij, [MW]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = branch.branch_id
        sheet.cell(row=row_idx, column=4).value = branch.fbus
        sheet.cell(row=row_idx, column=5).value = branch.tbus
        sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = expected_values['pij'][branch_id]
        sheet.cell(row=row_idx, column=8).number_format = decimal_style
        row_idx = row_idx + 1

        # Pji, [MW]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = branch.branch_id
        sheet.cell(row=row_idx, column=4).value = branch.tbus
        sheet.cell(row=row_idx, column=5).value = branch.fbus
        sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = expected_values['pji'][branch_id]
        sheet.cell(row=row_idx, column=8).number_format = decimal_style
        row_idx = row_idx + 1

        # Qij, [MVAr]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = branch.branch_id
        sheet.cell(row=row_idx, column=4).value = branch.fbus
        sheet.cell(row=row_idx, column=5).value = branch.tbus
        sheet.cell(row=row_idx, column=6).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = expected_values['qij'][branch_id]
        sheet.cell(row=row_idx, column=8).number_format = decimal_style
        row_idx = row_idx + 1

        # Qji, [MVAr]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = branch.branch_id
        sheet.cell(row=row_idx, column=4).value = branch.tbus
        sheet.cell(row=row_idx, column=5).value = branch.fbus
        sheet.cell(row=row_idx, column=6).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = expected_values['qji'][branch_id]
        sheet.cell(row=row_idx, column=8).number_format = decimal_style
        row_idx = row_idx + 1

        # Sij, [MVA]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = branch.branch_id
        sheet.cell(row=row_idx, column=4).value = branch.fbus
        sheet.cell(row=row_idx, column=5).value = branch.tbus
        sheet.cell(row=row_idx, column=6).value = 'S, [MVA]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = expected_values['sij'][branch_id]
        sheet.cell(row=row_idx, column=8).number_format = decimal_style
        row_idx = row_idx + 1

        # Sij, [%]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = branch.branch_id
        sheet.cell(row=row_idx, column=4).value = branch.fbus
        sheet.cell(row=row_idx, column=5).value = branch.tbus
        sheet.cell(row=row_idx, column=6).value = 'S, [%]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = abs(expected_values['sij'][branch_id]) / rating
        sheet.cell(row=row_idx, column=8).number_format = perc_style
        row_idx = row_idx + 1

        # Sji, [MVA]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = branch.branch_id
        sheet.cell(row=row_idx, column=4).value = branch.tbus
        sheet.cell(row=row_idx, column=5).value = branch.fbus
        sheet.cell(row=row_idx, column=6).value = 'S, [MVA]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = expected_values['sji'][branch_id]
        sheet.cell(row=row_idx, column=8).number_format = decimal_style
        row_idx = row_idx + 1

        # Sji, [%]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = branch.branch_id
        sheet.cell(row=row_idx, column=4).value = branch.tbus
        sheet.cell(row=row_idx, column=5).value = branch.fbus
        sheet.cell(row=row_idx, column=6).value = 'S, [%]'
        sheet.cell(row=row_idx, column=7).value = 'Expected'
        sheet.cell(row=row_idx, column=8).value = abs(expected_values['sji'][branch_id]) / rating
        sheet.cell(row=row_idx, column=8).number_format = perc_style
        row_idx = row_idx + 1

    return row_idx


def _write_network_energy_storages_results_to_excel(operational_planning, workbook, t, results):

    sheet = workbook.create_sheet('Energy Storage')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'ESS ID'
    sheet.cell(row=row_idx, column=4).value = 'Node ID'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=7).value = t
    row_idx = row_idx + 1

    # Write results -- TSO
    tso_results = results['tso']
    transmission_network = operational_planning.transmission_network
    if operational_planning.transmission_network.params.es_reg:
        row_idx = _write_network_energy_storages_results_per_operator(transmission_network, sheet, 'TSO', row_idx, tso_results)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        if operational_planning.distribution_networks[tn_node_id].params.es_reg:
            row_idx = _write_network_energy_storages_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_energy_storages_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    percent_style = '0.00%'

    expected_p = dict()
    expected_q = dict()
    expected_s = dict()
    expected_soc = dict()
    expected_soc_percent = dict()
    for energy_storage in network.energy_storages:
        es_id = energy_storage.es_id
        expected_p[es_id] = 0.00
        expected_q[es_id] = 0.00
        expected_s[es_id] = 0.00
        expected_soc[es_id] = 0.00
        expected_soc_percent[es_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for energy_storage in network.energy_storages:

            es_id = energy_storage.es_id
            node_id = energy_storage.bus
            ess_p = results['scenarios'][s_o]['energy_storages']['p'][es_id]
            ess_q = results['scenarios'][s_o]['energy_storages']['q'][es_id]
            ess_s = results['scenarios'][s_o]['energy_storages']['s'][es_id]
            ess_soc = results['scenarios'][s_o]['energy_storages']['soc'][es_id]
            ess_soc_percent = results['scenarios'][s_o]['energy_storages']['soc_percent'][es_id]

            # - Active Power
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = es_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=6).value = s_o
            sheet.cell(row=row_idx, column=7).value = ess_p
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            expected_p[es_id] += ess_p * omega_s
            row_idx = row_idx + 1

            # - Reactive Power
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = es_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=6).value = s_o
            sheet.cell(row=row_idx, column=7).value = ess_q
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            expected_q[es_id] += ess_q * omega_s
            row_idx = row_idx + 1

            # - Apparent Power
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = es_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
            sheet.cell(row=row_idx, column=6).value = s_o
            sheet.cell(row=row_idx, column=7).value = ess_s
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            expected_s[es_id] += ess_s  * omega_s
            row_idx = row_idx + 1

            # State-of-Charge, [MWh]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = es_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = 'SoC, [MWh]'
            sheet.cell(row=row_idx, column=6).value = s_o
            sheet.cell(row=row_idx, column=7).value = ess_soc
            sheet.cell(row=row_idx, column=7).number_format = decimal_style
            if ess_soc != 'N/A':
                expected_soc[es_id] += ess_soc * omega_s
            else:
                expected_soc[es_id] = ess_soc
            row_idx = row_idx + 1

            # State-of-Charge, [%]
            sheet.cell(row=row_idx, column=1).value = operator_type
            sheet.cell(row=row_idx, column=2).value = tn_node_id
            sheet.cell(row=row_idx, column=3).value = es_id
            sheet.cell(row=row_idx, column=4).value = node_id
            sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
            sheet.cell(row=row_idx, column=6).value = s_o
            sheet.cell(row=row_idx, column=7).value = ess_soc_percent
            sheet.cell(row=row_idx, column=7).number_format = percent_style
            if ess_soc_percent != 'N/A':
                expected_soc_percent[es_id] += ess_soc_percent * omega_s
            else:
                expected_soc_percent[es_id] = ess_soc_percent
            row_idx = row_idx + 1

    for energy_storage in network.energy_storages:

        es_id = energy_storage.es_id
        node_id = energy_storage.bus

        # - Active Power
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = es_id
        sheet.cell(row=row_idx, column=4).value = node_id
        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=6).value = 'Expected'
        sheet.cell(row=row_idx, column=7).value = expected_p[es_id]
        sheet.cell(row=row_idx, column=7).number_format = decimal_style
        row_idx = row_idx + 1

        # - Reactive Power
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = es_id
        sheet.cell(row=row_idx, column=4).value = node_id
        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=6).value = 'Expected'
        sheet.cell(row=row_idx, column=7).value = expected_q[es_id]
        sheet.cell(row=row_idx, column=7).number_format = decimal_style
        row_idx = row_idx + 1

        # - Apparent Power
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = es_id
        sheet.cell(row=row_idx, column=4).value = node_id
        sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
        sheet.cell(row=row_idx, column=6).value = 'Expected'
        sheet.cell(row=row_idx, column=7).value = expected_s[es_id]
        sheet.cell(row=row_idx, column=7).number_format = decimal_style
        row_idx = row_idx + 1

        # State-of-Charge, [MWh]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = es_id
        sheet.cell(row=row_idx, column=4).value = node_id
        sheet.cell(row=row_idx, column=5).value = 'SoC, [MWh]'
        sheet.cell(row=row_idx, column=6).value = 'Expected'
        sheet.cell(row=row_idx, column=7).value = expected_soc[es_id]
        sheet.cell(row=row_idx, column=7).number_format = decimal_style
        row_idx = row_idx + 1

        # State-of-Charge, [%]
        sheet.cell(row=row_idx, column=1).value = operator_type
        sheet.cell(row=row_idx, column=2).value = tn_node_id
        sheet.cell(row=row_idx, column=3).value = es_id
        sheet.cell(row=row_idx, column=4).value = node_id
        sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
        sheet.cell(row=row_idx, column=6).value = 'Expected'
        sheet.cell(row=row_idx, column=7).value = expected_soc_percent[es_id]
        sheet.cell(row=row_idx, column=7).number_format = percent_style
        row_idx = row_idx + 1

    return row_idx


def _write_relaxation_slacks_results_to_excel(operational_planning, workbook, t, results):
    _write_relaxation_slacks_results_network_operators_to_excel(operational_planning, workbook, t, results)


def _write_relaxation_slacks_results_network_operators_to_excel(operational_planning, workbook, t, results):

    sheet = workbook.create_sheet('Relaxation Slacks TSO, DSOs')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Resource ID'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=6).value = t
    row_idx = row_idx + 1

    # Write results -- TSO
    tso_results = results['tso']
    transmission_network = operational_planning.transmission_network
    tn_params = operational_planning.transmission_network.params
    if tn_params.slacks:
        row_idx = _write_relaxation_slacks_results_per_operator(transmission_network, sheet, 'TSO', row_idx, tso_results)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = operational_planning.distribution_networks[tn_node_id]
        if distribution_network.params.slacks:
            row_idx = _write_relaxation_slacks_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_relaxation_slacks_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    params = network.params

    for s_o in results['scenarios']:

        # Voltage slacks
        if params.slacks.grid_operation.voltage:

            for node in network.nodes:

                node_id = node.bus_i
                slack_e = results['scenarios'][s_o]['relaxation_slacks']['voltage']['e'][node_id]
                slack_f = results['scenarios'][s_o]['relaxation_slacks']['voltage']['f'][node_id]

                # - e
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = 'Voltage, e'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = slack_e
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

                # - f
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = 'Voltage, f'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = slack_f
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

        # Branch flow slacks
        if params.slacks.grid_operation.branch_flow:

            for branch in network.branches:

                branch_id = branch.branch_id
                iij_sqr = results['scenarios'][s_o]['relaxation_slacks']['branch_flow']['flow_ij_sqr'][branch_id]

                # - flow_ij
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch_id
                sheet.cell(row=row_idx, column=4).value = 'Flow_ij_sqr'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = iij_sqr
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

        # Node balance
        if params.slacks.node_balance:

            for node in network.nodes:

                node_id = node.bus_i
                slack_node_balance_p = results['scenarios'][s_o]['relaxation_slacks']['node_balance']['p'][node_id]
                slack_node_balance_q = results['scenarios'][s_o]['relaxation_slacks']['node_balance']['q'][node_id]

                # - p
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = 'Node balance, p'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = slack_node_balance_p
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

                # - q
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = 'Node balance, q'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = slack_node_balance_q
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

        # Shared ESS
        for shared_energy_storage in network.shared_energy_storages:

            node_id = shared_energy_storage.bus

            # - Complementarity
            if params.slacks.shared_ess.complementarity:

                comp = results['scenarios'][s_o]['relaxation_slacks']['shared_energy_storages']['comp'][node_id]

                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = 'Shared Energy Storage, comp'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = comp
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

            # - Charging
            if params.slacks.shared_ess.charging:

                sch_up = results['scenarios'][s_o]['relaxation_slacks']['shared_energy_storages']['sch_up'][node_id]
                sch_down = results['scenarios'][s_o]['relaxation_slacks']['shared_energy_storages']['sch_down'][node_id]
                sdch_up = results['scenarios'][s_o]['relaxation_slacks']['shared_energy_storages']['sdch_up'][node_id]
                sdch_down = results['scenarios'][s_o]['relaxation_slacks']['shared_energy_storages']['sdch_down'][node_id]


                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = 'Shared Energy Storage, sch_up'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = sch_up
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = 'Shared Energy Storage, sch_down'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = sch_down
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = 'Shared Energy Storage, sdch_up'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = sdch_up
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = 'Shared Energy Storage, sdch_down'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = sdch_down
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                row_idx = row_idx + 1

        # ESS
        if params.es_reg:

            for energy_storage in network.energy_storages:

                es_id = energy_storage.es_id

                # - Complementarity
                if params.slacks.ess.complementarity:

                    comp = results['scenarios'][s_o]['relaxation_slacks']['energy_storages']['comp'][es_id]

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = es_id
                    sheet.cell(row=row_idx, column=4).value = 'Energy Storage, comp'
                    sheet.cell(row=row_idx, column=5).value = s_o
                    sheet.cell(row=row_idx, column=6).value = comp
                    sheet.cell(row=row_idx, column=6).number_format = decimal_style
                    row_idx = row_idx + 1

                # - Charging
                if params.slacks.ess.charging:

                    sch_up = results['scenarios'][s_o]['relaxation_slacks']['energy_storages']['sch_up'][es_id]
                    sch_down = results['scenarios'][s_o]['relaxation_slacks']['energy_storages']['sch_down'][es_id]
                    sdch_up = results['scenarios'][s_o]['relaxation_slacks']['energy_storages']['sch_up'][es_id]
                    sdch_down = results['scenarios'][s_o]['relaxation_slacks']['energy_storages']['sdch_down'][es_id]

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = es_id
                    sheet.cell(row=row_idx, column=4).value = 'Energy Storage, sch_up'
                    sheet.cell(row=row_idx, column=5).value = s_o
                    sheet.cell(row=row_idx, column=6).value = sch_up
                    sheet.cell(row=row_idx, column=6).number_format = decimal_style
                    row_idx = row_idx + 1

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = es_id
                    sheet.cell(row=row_idx, column=4).value = 'Energy Storage, sch_down'
                    sheet.cell(row=row_idx, column=5).value = s_o
                    sheet.cell(row=row_idx, column=6).value = sch_down
                    sheet.cell(row=row_idx, column=6).number_format = decimal_style
                    row_idx = row_idx + 1

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = es_id
                    sheet.cell(row=row_idx, column=4).value = 'Energy Storage, sdch_up'
                    sheet.cell(row=row_idx, column=5).value = s_o
                    sheet.cell(row=row_idx, column=6).value = sdch_up
                    sheet.cell(row=row_idx, column=6).number_format = decimal_style
                    row_idx = row_idx + 1

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = es_id
                    sheet.cell(row=row_idx, column=4).value = 'Energy Storage, sdch_down'
                    sheet.cell(row=row_idx, column=5).value = s_o
                    sheet.cell(row=row_idx, column=6).value = sdch_down
                    sheet.cell(row=row_idx, column=6).number_format = decimal_style
                    row_idx = row_idx + 1

    return row_idx




# ======================================================================================================================
#   Aux functions
# ======================================================================================================================
def _check_interface_nodes_base_voltage_consistency(operational_planning):
    for node_id in operational_planning.distribution_networks:
        tn_node_base_kv = operational_planning.transmission_network.get_node_base_kv(node_id)
        dn_ref_node_id = operational_planning.distribution_networks[node_id].get_reference_node_id()
        dn_node_base_kv = operational_planning.distribution_networks[node_id].get_node_base_kv(dn_ref_node_id)
        if not isclose(tn_node_base_kv, dn_node_base_kv, rel_tol=5e-2):
            print(f'[ERROR] Inconsistent TN-DN base voltage at node {node_id}! Check network(s). Exiting')
            exit(ERROR_SPECIFICATION_FILE)


def _add_shared_energy_storage_to_transmission_network(operational_planning):
    s_base = operational_planning.transmission_network.baseMVA
    for node_id in operational_planning.active_distribution_network_nodes:
        for shared_ess in operational_planning.shared_ess_data.shared_energy_storages:
            if shared_ess.bus == node_id:
                shared_energy_storage = EnergyStorage()
                shared_energy_storage.bus = node_id
                shared_energy_storage.s = shared_ess.s / s_base
                shared_energy_storage.e = shared_ess.e / s_base
                shared_energy_storage.e_init = shared_ess.e_init / s_base
                shared_energy_storage.e_min = shared_ess.e_min / s_base
                shared_energy_storage.e_max = shared_ess.e_max / s_base
                shared_energy_storage.eff_ch = shared_ess.eff_ch
                shared_energy_storage.eff_dch = shared_ess.eff_dch
                shared_energy_storage.max_pf = shared_ess.max_pf
                shared_energy_storage.min_pf = shared_ess.min_pf
                operational_planning.transmission_network.shared_energy_storages.append(shared_energy_storage)


def _add_shared_energy_storage_to_distribution_network(operational_planning):
    for node_id in operational_planning.distribution_networks:
        s_base = operational_planning.distribution_networks[node_id].baseMVA
        for shared_ess in operational_planning.shared_ess_data.shared_energy_storages:
            if shared_ess.bus == node_id:
                shared_energy_storage = EnergyStorage()
                shared_energy_storage.bus = operational_planning.distribution_networks[node_id].get_reference_node_id()
                shared_energy_storage.s = shared_energy_storage.s / s_base
                shared_energy_storage.e = shared_energy_storage.e / s_base
                shared_energy_storage.e_init = shared_ess.e_init / s_base
                shared_energy_storage.e_min = shared_ess.e_min / s_base
                shared_energy_storage.e_max = shared_ess.e_max / s_base
                shared_energy_storage.eff_ch = shared_ess.eff_ch
                shared_energy_storage.eff_dch = shared_ess.eff_dch
                shared_energy_storage.max_pf = shared_ess.max_pf
                shared_energy_storage.min_pf = shared_ess.min_pf
                operational_planning.distribution_networks[node_id].shared_energy_storages.append(shared_energy_storage)


def print_debug_info(operational_planning, consensus_vars, print_vmag=False, print_pf=False, print_ess=False):
    for node_id in operational_planning.active_distribution_network_nodes:
        if print_vmag:
            print(f"\tNode {node_id}, PF, TSO,  V  {[sqrt(vmag) for vmag in consensus_vars['v_sqr']['tso']['current'][node_id]]}")
            print(f"\tNode {node_id}, PF, DSO,  V  {[sqrt(vmag) for vmag in consensus_vars['v_sqr']['dso']['current'][node_id]]}")
        if print_pf:
            print(f"\tNode {node_id}, PF, TSO,  P {consensus_vars['pf']['tso']['current'][node_id]['p']}")
            print(f"\tNode {node_id}, PF, DSO,  P {consensus_vars['pf']['dso']['current'][node_id]['p']}")
            print(f"\tNode {node_id}, PF, TSO,  Q {consensus_vars['pf']['tso']['current'][node_id]['q']}")
            print(f"\tNode {node_id}, PF, DSO,  Q {consensus_vars['pf']['dso']['current'][node_id]['q']}")
        if print_ess:
            print(f"\tNode {node_id}, ESS, TSO,  P {consensus_vars['ess']['tso']['current'][node_id]['p']}")
            print(f"\tNode {node_id}, ESS, DSO,  P {consensus_vars['ess']['dso']['current'][node_id]['p']}")
            print(f"\tNode {node_id}, ESS, TSO,  Q {consensus_vars['ess']['tso']['current'][node_id]['q']}")
            print(f"\tNode {node_id}, ESS, DSO,  Q {consensus_vars['ess']['dso']['current'][node_id]['q']}")
