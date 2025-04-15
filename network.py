import os
import pandas as pd
import pyomo.opt as po
import pyomo.environ as pe
from math import acos, sqrt, tan, atan2, pi, isclose
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from node import Node
from load import Load
from branch import Branch
from generator import Generator
from energy_storage import EnergyStorage
from network_parameters import NetworkParameters
from helper_functions import *


# ======================================================================================================================
#   Class NETWORK
# ======================================================================================================================
class Network:

    def __init__(self):
        self.name = str()
        self.data_dir = str()
        self.results_dir = str()
        self.diagrams_dir = str()
        self.num_instants = 24
        self.data_loaded = False
        self.is_transmission = False
        self.baseMVA = 100.0
        self.nodes = list()
        self.loads = list()
        self.branches = list()
        self.generators = list()
        self.energy_storages = list()
        self.shared_energy_storages = list()
        self.prob_operation_scenarios = list()          # Probability of operation (generation and consumption) scenarios
        self.cost_energy_p = list()
        self.cost_flex = list()
        self.active_distribution_network_nodes = list()
        self.params = NetworkParameters()

    def run_opf(self):
        model = self.build_model()
        results = self.optimize(model)
        processed_results = self.process_results(model, results)
        self.write_optimization_results_to_excel(processed_results)

    def determine_pq_map(self, num_steps=12):
        model = _build_pq_map_model(self)

    def build_model(self):
        _pre_process_network(self)
        return _build_model(self)

    def optimize(self, model, from_warm_start=False):
        return _optimize(self, model, from_warm_start=from_warm_start)

    def read_network_data(self, operational_data_filename):
        _read_network_data(self, operational_data_filename)

    def read_network_from_json_file(self, network_filename):
        filename = os.path.join(self.data_dir, self.name, network_filename)
        _read_network_from_json_file(self, filename)
        self.perform_network_check()

    def read_network_operational_data_from_file(self, operational_data_filename):
        filename = os.path.join(self.data_dir, self.name, operational_data_filename)
        data = _read_network_operational_data_from_file(self, filename)
        _update_network_with_excel_data(self, data)

    def read_network_parameters(self, params_filename):
        filename = os.path.join(self.data_dir, self.name, params_filename)
        self.params.read_parameters_from_file(filename)

    def get_reference_node_id(self):
        for node in self.nodes:
            if node.type == BUS_REF:
                return node.bus_i
        print(f'[ERROR] Network {self.name}. No REF NODE found! Check network.')
        exit(ERROR_NETWORK_FILE)

    def get_node_idx(self, node_id):
        for i in range(len(self.nodes)):
            if self.nodes[i].bus_i == node_id:
                return i
        print(f'[ERROR] Network {self.name}. Bus ID {node_id} not found! Check network model.')
        exit(ERROR_NETWORK_FILE)

    def get_node_type(self, node_id):
        for node in self.nodes:
            if node.bus_i == node_id:
                return node.type
        print(f'[ERROR] Network {self.name}. Node {node_id} not found! Check network.')
        exit(ERROR_NETWORK_FILE)

    def get_node_base_kv(self, node_id):
        for node in self.nodes:
            if node.bus_i == node_id:
                return node.base_kv
        print(f'[ERROR] Network {self.name}. Node {node_id} not found! Check network.')
        exit(ERROR_NETWORK_FILE)

    def get_node_voltage_limits(self, node_id):
        for node in self.nodes:
            if node.bus_i == node_id:
                return node.v_min, node.v_max
        print(f'[ERROR] Network {self.name}. Node {node_id} not found! Check network.')
        exit(ERROR_NETWORK_FILE)

    def node_exists(self, node_id):
        for node in self.nodes:
            if node.bus_i == node_id:
                return True
        return False

    def get_branch_idx(self, branch):
        for b in range(len(self.branches)):
            if self.branches[b].branch_id == branch.branch_id:
                return b
        print(f'[ERROR] Network {self.name}. No Branch connecting bus {branch.fbus} and bus {branch.tbus} found! Check network.')
        exit(ERROR_NETWORK_FILE)

    def get_num_renewable_gens(self):
        num_renewable_gens = 0
        for generator in self.generators:
            if generator.gen_type in GEN_CURTAILLABLE_TYPES:
                num_renewable_gens += 1
        return num_renewable_gens

    def get_reference_gen_idx(self):
        ref_node_id = self.get_reference_node_id()
        for i in range(len(self.generators)):
            gen = self.generators[i]
            if gen.bus == ref_node_id:
                return i
        print(f'[ERROR] Network {self.name}. No REF NODE found! Check network.')
        exit(ERROR_NETWORK_FILE)

    def get_gen_idx(self, node_id):
        for g in range(len(self.generators)):
            gen = self.generators[g]
            if gen.bus == node_id:
                return g
        print(f'[ERROR] Network {self.name}. No Generator in bus {node_id} found! Check network.')
        exit(ERROR_NETWORK_FILE)

    def get_gen_type(self, gen_id):
        description = 'Unknown'
        for gen in self.generators:
            if gen.gen_id == gen_id:
                if gen.gen_type == GEN_REFERENCE:
                    description = 'Reference (TN)'
                elif gen.gen_type == GEN_CONV:
                    description = 'Conventional'
                elif gen.gen_type == GEN_RES_CONTROLLABLE:
                    description = 'RES (Generic, Controllable)'
                elif gen.gen_type == GEN_RES_SOLAR:
                    description = 'RES (Solar)'
                elif gen.gen_type == GEN_RES_WIND:
                    description = 'RES (Wind)'
                elif gen.gen_type == GEN_RES_OTHER:
                    description = 'RES (Generic, Non-controllable)'
        return description

    def process_results(self, model, results=dict()):
        return _process_results(self, model, results=results)

    def compute_series_admittance(self):
        for branch in self.branches:
            branch.g = branch.r / (branch.r ** 2 + branch.x ** 2)
            branch.b = -branch.x / (branch.r ** 2 + branch.x ** 2)

    def perform_network_check(self):
        _perform_network_check(self)

    def write_optimization_results_to_excel(self, results, filename=str()):
        if not filename:
            filename = self.name
        _write_optimization_results_to_excel(self, self.results_dir, results, filename=filename)


# ======================================================================================================================
#   NETWORK optimization functions
# ======================================================================================================================
def _build_model(network, n=0):

    network.compute_series_admittance()

    model = pe.ConcreteModel()
    model.name = network.name
    params = network.params

    # ------------------------------------------------------------------------------------------------------------------
    # Sets
    model.scenarios_operation = range(len(network.prob_operation_scenarios))
    model.nodes = range(len(network.nodes))
    model.loads = range(len(network.loads))
    model.generators = range(len(network.generators))
    model.branches = range(len(network.branches))
    model.energy_storages = range(len(network.energy_storages))
    model.shared_energy_storages = range(len(network.shared_energy_storages))

    # ------------------------------------------------------------------------------------------------------------------
    # Decision variables
    # - Voltage
    model.e = pe.Var(model.nodes, model.scenarios_operation, domain=pe.Reals, initialize=1.0)
    model.f = pe.Var(model.nodes, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
    model.e_actual = pe.Var(model.nodes, model.scenarios_operation, domain=pe.Reals, initialize=1.0)
    model.f_actual = pe.Var(model.nodes, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
    if params.slacks.grid_operation.voltage:
        model.slack_e = pe.Var(model.nodes, model.scenarios_operation, domain=pe.Reals, initialize=0.00)
        model.slack_f = pe.Var(model.nodes, model.scenarios_operation, domain=pe.Reals, initialize=0.00)
    for i in model.nodes:
        node = network.nodes[i]
        e_lb, e_ub = -node.v_max, node.v_max
        f_lb, f_ub = -node.v_max, node.v_max
        for s_o in model.scenarios_operation:
            if params.slacks.grid_operation.voltage:
                model.slack_e[i, s_o].setub(VMAG_VIOLATION_ALLOWED)
                model.slack_e[i, s_o].setlb(-VMAG_VIOLATION_ALLOWED)
                model.slack_f[i, s_o].setub(VMAG_VIOLATION_ALLOWED)
                model.slack_f[i, s_o].setlb(-VMAG_VIOLATION_ALLOWED)
            if node.type == BUS_REF:
                if network.is_transmission:
                    model.e[i, s_o].setub(e_ub)
                    model.e[i, s_o].setlb(e_lb)
                else:
                    ref_gen_idx = network.get_gen_idx(node.bus_i)
                    vg = network.generators[ref_gen_idx].vg
                    model.e[i, s_o].setub(vg + EQUALITY_TOLERANCE)
                    model.e[i, s_o].setlb(vg - EQUALITY_TOLERANCE)
                    if params.slacks.grid_operation.voltage:
                        model.slack_e[i, s_o].setub(EQUALITY_TOLERANCE)
                        model.slack_e[i, s_o].setlb(-EQUALITY_TOLERANCE)
                model.f[i, s_o].setub(EQUALITY_TOLERANCE)
                model.f[i, s_o].setlb(-EQUALITY_TOLERANCE)
                if params.slacks.grid_operation.voltage:
                    model.slack_f[i, s_o].setub(EQUALITY_TOLERANCE)
                    model.slack_f[i, s_o].setlb(-EQUALITY_TOLERANCE)
            else:
                model.e[i, s_o].setub(e_ub)
                model.e[i, s_o].setlb(e_lb)
                model.f[i, s_o].setub(f_ub)
                model.f[i, s_o].setlb(f_lb)
    if params.slacks.node_balance:
        model.slack_node_balance_p = pe.Var(model.nodes, model.scenarios_operation, domain=pe.Reals, initialize=0.00)
        model.slack_node_balance_q = pe.Var(model.nodes, model.scenarios_operation, domain=pe.Reals, initialize=0.00)

    # - Generation
    model.pg = pe.Var(model.generators, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
    model.qg = pe.Var(model.generators, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
    for g in model.generators:
        generator = network.generators[g]
        pg_ub, pg_lb = generator.pmax, generator.pmin
        qg_ub, qg_lb = generator.qmax, generator.qmin
        for s_o in model.scenarios_operation:
            if generator.status[n]:
                model.pg[g, s_o] = max(pg_lb, 0.00)
                model.qg[g, s_o] = max(qg_lb, 0.00)
                model.pg[g, s_o].setub(pg_ub)
                model.pg[g, s_o].setlb(pg_lb)
                model.qg[g, s_o].setub(qg_ub)
                model.qg[g, s_o].setlb(qg_lb)
            else:
                model.pg[g, s_o].setub(EQUALITY_TOLERANCE)
                model.pg[g, s_o].setlb(-EQUALITY_TOLERANCE)
                model.qg[g, s_o].setub(EQUALITY_TOLERANCE)
                model.qg[g, s_o].setlb(-EQUALITY_TOLERANCE)
    if params.rg_curt:
        model.sg_abs = pe.Var(model.generators, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.sg_sqr = pe.Var(model.generators, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.sg_curt = pe.Var(model.generators, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        for g in model.generators:
            generator = network.generators[g]
            for s_o in model.scenarios_operation:
                if generator.is_curtaillable():
                    # - Renewable Generation
                    init_sg = 0.0
                    if generator.status[n]:
                        init_sg = sqrt(generator.pg[s_o][n] ** 2 + generator.qg[s_o][n] ** 2)
                    model.sg_abs[g, s_o].setub(init_sg)
                    model.sg_sqr[g, s_o].setub(init_sg ** 2)
                    model.sg_curt[g, s_o].setub(init_sg)
                else:
                    model.sg_abs[g, s_o].setub(EQUALITY_TOLERANCE)
                    model.sg_sqr[g, s_o].setub(EQUALITY_TOLERANCE)
                    model.sg_curt[g, s_o].setub(EQUALITY_TOLERANCE)

    # - Branch power flows (squared) -- used in branch limits
    model.flow_ij_sqr = pe.Var(model.branches, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
    if params.slacks.grid_operation.branch_flow:
        model.slack_flow_ij_sqr = pe.Var(model.branches, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
    for b in model.branches:
        for s_o in model.scenarios_operation:
            if network.branches[b].status:
                if params.slacks.grid_operation.branch_flow:
                    rating = network.branches[b].rate / network.baseMVA
                    model.slack_flow_ij_sqr[b, s_o].setub(SIJ_VIOLATION_ALLOWED * rating)
            else:
                model.flow_ij_sqr[b, s_o].setub(EQUALITY_TOLERANCE)
                if params.slacks.grid_operation.branch_flow:
                    model.slack_flow_ij_sqr[b, s_o].setub(EQUALITY_TOLERANCE)

    # - Loads
    model.pc = pe.Var(model.loads, model.scenarios_operation, domain=pe.Reals)
    model.qc = pe.Var(model.loads, model.scenarios_operation, domain=pe.Reals)
    for c in model.loads:
        load = network.loads[c]
        for s_o in model.scenarios_operation:
            model.pc[c, s_o].setub(load.pd[s_o][n] + EQUALITY_TOLERANCE)
            model.pc[c, s_o].setlb(load.pd[s_o][n] - EQUALITY_TOLERANCE)
            model.qc[c, s_o].setub(load.qd[s_o][n] + EQUALITY_TOLERANCE)
            model.qc[c, s_o].setlb(load.qd[s_o][n] - EQUALITY_TOLERANCE)
    if params.fl_reg:
        model.flex_p_up = pe.Var(model.loads, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.flex_p_down = pe.Var(model.loads, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        for c in model.loads:
            load = network.loads[c]
            for s_o in model.scenarios_operation:
                if load.fl_reg:
                    flex_up = load.flexibility.upward[n]
                    flex_down = load.flexibility.downward[n]
                    model.flex_p_up[c, s_o].setub(abs(flex_up))
                    model.flex_p_down[c, s_o].setub(abs(flex_down))
                else:
                    model.flex_p_up[c, s_o].setub(EQUALITY_TOLERANCE)
                    model.flex_p_down[c, s_o].setub(EQUALITY_TOLERANCE)
    if params.l_curt:
        model.pc_curt_down = pe.Var(model.loads, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.pc_curt_up = pe.Var(model.loads, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.qc_curt_down = pe.Var(model.loads, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.qc_curt_up = pe.Var(model.loads, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        for c in model.loads:
            load = network.loads[c]
            for s_o in model.scenarios_operation:
                if load.pd[s_o][n] >= 0.00:
                    model.pc_curt_down[c, s_o].setub(abs(load.pd[s_o][n]))
                    model.pc_curt_up[c, s_o].setub(EQUALITY_TOLERANCE)
                else:
                    model.pc_curt_up[c, s_o].setub(abs(load.pd[s_o][n]))
                    model.pc_curt_down[c, s_o].setub(EQUALITY_TOLERANCE)

                if load.qd[s_o][n] >= 0.00:
                    model.qc_curt_down[c, s_o].setub(abs(load.qd[s_o][n]))
                    model.qc_curt_up[c, s_o].setub(EQUALITY_TOLERANCE)
                else:
                    model.qc_curt_up[c, s_o].setub(abs(load.qd[s_o][n]))
                    model.qc_curt_down[c, s_o].setub(EQUALITY_TOLERANCE)

    # - Transformers
    model.r = pe.Var(model.branches, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=1.0)
    for i in model.branches:
        branch = network.branches[i]
        for s_o in model.scenarios_operation:
            if branch.is_transformer:
                # - Transformer
                if params.transf_reg and branch.vmag_reg:
                    model.r[i, s_o].setub(TRANSFORMER_MAXIMUM_RATIO)
                    model.r[i, s_o].setlb(TRANSFORMER_MINIMUM_RATIO)
                else:
                    model.r[i, s_o].setub(branch.ratio + EQUALITY_TOLERANCE)
                    model.r[i, s_o].setlb(branch.ratio - EQUALITY_TOLERANCE)
            else:
                model.r[i, s_o].setub(1.00 + EQUALITY_TOLERANCE)
                model.r[i, s_o].setlb(1.00 - EQUALITY_TOLERANCE)

    # - Energy Storage devices
    if params.es_reg:
        model.es_soc = pe.Var(model.energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_sch = pe.Var(model.energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.es_pch = pe.Var(model.energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.es_qch = pe.Var(model.energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
        model.es_sdch = pe.Var(model.energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.es_pdch = pe.Var(model.energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        model.es_qdch = pe.Var(model.energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
        for e in model.energy_storages:
            energy_storage = network.energy_storages[e]
            for s_o in model.scenarios_operation:
                model.es_soc[e, s_o] = energy_storage.e_init
                model.es_soc[e, s_o].setlb(energy_storage.e_min)
                model.es_soc[e, s_o].setub(energy_storage.e_max)
                model.es_sch[e, s_o].setub(energy_storage.s)
                model.es_pch[e, s_o].setub(energy_storage.s)
                model.es_qch[e, s_o].setub(energy_storage.s)
                model.es_qch[e, s_o].setlb(-energy_storage.s)
                model.es_sdch[e, s_o].setub(energy_storage.s)
                model.es_pdch[e, s_o].setub(energy_storage.s)
                model.es_qdch[e, s_o].setub(energy_storage.s)
                model.es_qdch[e, s_o].setlb(-energy_storage.s)
        if params.slacks.ess.complementarity:
            model.slack_es_comp = pe.Var(model.energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
        if params.slacks.ess.charging:
            model.slack_es_ch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
            model.slack_es_dch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)

    # - Shared Energy Storage devices
    model.shared_es_soc = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.00)
    model.shared_es_sch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
    model.shared_es_pch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
    model.shared_es_qch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
    model.shared_es_sdch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
    model.shared_es_pdch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
    model.shared_es_qdch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
    model.shared_es_pnet = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
    model.shared_es_qnet = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
    for e in model.shared_energy_storages:
        shared_energy_storage = network.shared_energy_storages[e]
        for s_o in model.scenarios_operation:
            model.shared_es_soc[e, s_o] = shared_energy_storage.e_init
            model.shared_es_soc[e, s_o].setlb(shared_energy_storage.e_min)
            model.shared_es_soc[e, s_o].setub(shared_energy_storage.e_max)
            model.shared_es_sch[e, s_o].setub(shared_energy_storage.s)
            model.shared_es_pch[e, s_o].setub(shared_energy_storage.s)
            model.shared_es_qch[e, s_o].setub(shared_energy_storage.s)
            model.shared_es_qch[e, s_o].setlb(-shared_energy_storage.s)
            model.shared_es_sdch[e, s_o].setub(shared_energy_storage.s)
            model.shared_es_pdch[e, s_o].setub(shared_energy_storage.s)
            model.shared_es_qdch[e, s_o].setub(shared_energy_storage.s)
            model.shared_es_qdch[e, s_o].setlb(-shared_energy_storage.s)
    if params.slacks.shared_ess.complementarity:
        model.slack_shared_es_comp = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.0)
    if params.slacks.shared_ess.charging:
        model.slack_shared_es_ch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)
        model.slack_shared_es_dch = pe.Var(model.shared_energy_storages, model.scenarios_operation, domain=pe.Reals, initialize=0.0)

    # ------------------------------------------------------------------------------------------------------------------
    # Constraints
    # - Voltage
    model.voltage_cons = pe.ConstraintList()
    for i in model.nodes:
        node = network.nodes[i]
        for s_o in model.scenarios_operation:
            # e_actual and f_actual definition
            e_actual = model.e[i, s_o]
            f_actual = model.f[i, s_o]
            if params.slacks.grid_operation.voltage:
                e_actual += model.slack_e[i, s_o]
                f_actual += model.slack_f[i, s_o]

            model.voltage_cons.add(model.e_actual[i, s_o] == e_actual)
            model.voltage_cons.add(model.f_actual[i, s_o] == f_actual)

            # voltage magnitude constraints
            if node.type == BUS_PV:
                if params.enforce_vg:
                    # - Enforce voltage controlled bus
                    gen_idx = network.get_gen_idx(node.bus_i)
                    vg = network.generators[gen_idx].vg
                    e = model.e[i, s_o]
                    f = model.f[i, s_o]
                    model.voltage_cons.add(e ** 2 + f ** 2 <= vg[n] ** 2 + EQUALITY_TOLERANCE)
                    model.voltage_cons.add(e ** 2 + f ** 2 >= vg[n] ** 2 - EQUALITY_TOLERANCE)
                else:
                    # - Voltage at the bus is not controlled
                    e = model.e[i, s_o]
                    f = model.f[i, s_o]
                    model.voltage_cons.add(e ** 2 + f ** 2 >= node.v_min**2)
                    model.voltage_cons.add(e ** 2 + f ** 2 <= node.v_max**2)
            else:
                e = model.e[i, s_o]
                f = model.f[i, s_o]
                model.voltage_cons.add(e ** 2 + f ** 2 >= node.v_min**2)
                model.voltage_cons.add(e ** 2 + f ** 2 <= node.v_max**2)

    model.generation_apparent_power = pe.ConstraintList()
    model.generation_power_factor = pe.ConstraintList()
    if params.rg_curt:
        for g in model.generators:
            generator = network.generators[g]
            for s_o in model.scenarios_operation:
                if generator.is_curtaillable():
                    init_sg = 0.0
                    if generator.status[n]:
                        init_sg = sqrt(generator.pg[s_o][n] ** 2 + generator.qg[s_o][n] ** 2)
                    model.generation_apparent_power.add(model.sg_sqr[g, s_o] <= model.pg[g, s_o] ** 2 + model.qg[g, s_o] ** 2 + EQUALITY_TOLERANCE)
                    model.generation_apparent_power.add(model.sg_sqr[g, s_o] >= model.pg[g, s_o] ** 2 + model.qg[g, s_o] ** 2 - EQUALITY_TOLERANCE)
                    model.generation_apparent_power.add(model.sg_abs[g, s_o] ** 2 <= model.sg_sqr[g, s_o] + EQUALITY_TOLERANCE)
                    model.generation_apparent_power.add(model.sg_abs[g, s_o] ** 2 >= model.sg_sqr[g, s_o] - EQUALITY_TOLERANCE)
                    model.generation_apparent_power.add(model.sg_abs[g, s_o] <= init_sg - model.sg_curt[g, s_o] + EQUALITY_TOLERANCE)
                    model.generation_apparent_power.add(model.sg_abs[g, s_o] >= init_sg - model.sg_curt[g, s_o] - EQUALITY_TOLERANCE)
                    if generator.power_factor_control:
                        # Power factor control, variable phi
                        max_phi = acos(generator.max_pf)
                        min_phi = acos(generator.min_pf)
                        model.generation_power_factor.add(model.qg[g, s_o] <= tan(max_phi) * model.pg[g, s_o])
                        model.generation_power_factor.add(model.qg[g, s_o] >= tan(min_phi) * model.pg[g, s_o])
                    else:
                        # No power factor control, maintain given phi
                        phi = atan2(generator.qg[s_o][n], generator.pg[s_o][n])
                        model.generation_power_factor.add(model.qg[g, s_o] <= tan(phi) * model.pg[g, s_o])
                        model.generation_power_factor.add(model.qg[g, s_o] >= tan(phi) * model.pg[g, s_o])

    # - Energy Storage constraints
    if params.es_reg:

        model.energy_storage_balance = pe.ConstraintList()
        model.energy_storage_operation = pe.ConstraintList()
        model.energy_storage_ch_dch_exclusion = pe.ConstraintList()

        for e in model.energy_storages:

            energy_storage = network.energy_storages[e]
            soc_init = energy_storage.e_init
            eff_charge = energy_storage.eff_ch
            eff_discharge = energy_storage.eff_dch
            max_phi = acos(energy_storage.max_pf)
            min_phi = acos(energy_storage.min_pf)

            for s_o in model.scenarios_operation:

                sch = model.es_sch[e, s_o]
                pch = model.es_pch[e, s_o]
                qch = model.es_qch[e, s_o]
                sdch = model.es_sdch[e, s_o]
                pdch = model.es_pdch[e, s_o]
                qdch = model.es_qdch[e, s_o]

                # ESS operation
                model.energy_storage_operation.add(qch <= tan(max_phi) * pch)
                model.energy_storage_operation.add(qch >= tan(min_phi) * pch)
                model.energy_storage_operation.add(qdch <= tan(max_phi) * pdch)
                model.energy_storage_operation.add(qdch >= tan(min_phi) * pdch)

                if params.slacks.ess.charging:
                    model.energy_storage_operation.add(sch ** 2 == pch ** 2 + qch ** 2 + model.slack_es_ch[e, s_o])
                    model.energy_storage_operation.add(sdch ** 2 == pdch ** 2 + qdch ** 2 + model.slack_es_dch[e, s_o])
                else:
                    model.energy_storage_operation.add(sch ** 2 <= pch ** 2 + qch ** 2 + EQUALITY_TOLERANCE)
                    model.energy_storage_operation.add(sch ** 2 >= pch ** 2 + qch ** 2 - EQUALITY_TOLERANCE)
                    model.energy_storage_operation.add(sdch ** 2 <= pdch ** 2 + qdch ** 2 + EQUALITY_TOLERANCE)
                    model.energy_storage_operation.add(sdch ** 2 >= pdch ** 2 + qdch ** 2 - EQUALITY_TOLERANCE)

                # Charging/discharging complementarity constraints
                if params.slacks.ess.complementarity:
                    model.energy_storage_ch_dch_exclusion.add(sch * sdch == model.slack_es_comp[e, s_o])
                else:
                    model.energy_storage_ch_dch_exclusion.add(sch * sdch <= EQUALITY_TOLERANCE)

                # State-of-Charge
                soc_prev = soc_init
                model.energy_storage_balance.add(model.es_soc[e, s_o] <= soc_prev + (sch * eff_charge - sdch / eff_discharge) + EQUALITY_TOLERANCE)
                model.energy_storage_balance.add(model.es_soc[e, s_o] >= soc_prev + (sch * eff_charge - sdch / eff_discharge) - EQUALITY_TOLERANCE)

    # - Shared Energy Storage constraints
    model.shared_energy_storage_balance = pe.ConstraintList()
    model.shared_energy_storage_operation = pe.ConstraintList()
    model.shared_energy_storage_ch_dch_exclusion = pe.ConstraintList()
    for e in model.shared_energy_storages:

        shared_energy_storage = network.shared_energy_storages[e]
        soc_init = shared_energy_storage.e_init
        eff_charge = shared_energy_storage.eff_ch
        eff_discharge = shared_energy_storage.eff_dch
        max_phi = acos(shared_energy_storage.max_pf)
        min_phi = acos(shared_energy_storage.min_pf)

        for s_o in model.scenarios_operation:

            sch = model.shared_es_sch[e, s_o]
            pch = model.shared_es_pch[e, s_o]
            qch = model.shared_es_qch[e, s_o]
            sdch = model.shared_es_sdch[e, s_o]
            pdch = model.shared_es_pdch[e, s_o]
            qdch = model.shared_es_qdch[e, s_o]

            # ESS operation
            model.shared_energy_storage_operation.add(qch <= tan(max_phi) * pch)
            model.shared_energy_storage_operation.add(qch >= tan(min_phi) * pch)
            model.shared_energy_storage_operation.add(qdch <= tan(max_phi) * pdch)
            model.shared_energy_storage_operation.add(qdch >= tan(min_phi) * pdch)

            # Pnet and Qnet definition
            model.shared_energy_storage_operation.add(model.shared_es_pnet[e, s_o] <= pch - pdch + EQUALITY_TOLERANCE)
            model.shared_energy_storage_operation.add(model.shared_es_pnet[e, s_o] >= pch - pdch - EQUALITY_TOLERANCE)
            model.shared_energy_storage_operation.add(model.shared_es_qnet[e, s_o] <= qch - qdch + EQUALITY_TOLERANCE)
            model.shared_energy_storage_operation.add(model.shared_es_qnet[e, s_o] >= qch - qdch - EQUALITY_TOLERANCE)

            if params.slacks.shared_ess.charging:
                model.shared_energy_storage_operation.add(sch ** 2 == pch ** 2 + qch ** 2 + model.slack_shared_es_ch[e, s_o])
                model.shared_energy_storage_operation.add(sdch ** 2 == pdch ** 2 + qdch ** 2 + model.slack_shared_es_dch[e, s_o])
            else:
                model.shared_energy_storage_operation.add(sch ** 2 <= pch ** 2 + qch ** 2 + EQUALITY_TOLERANCE)
                model.shared_energy_storage_operation.add(sch ** 2 >= pch ** 2 + qch ** 2 - EQUALITY_TOLERANCE)
                model.shared_energy_storage_operation.add(sdch ** 2 <= pdch ** 2 + qdch ** 2 + EQUALITY_TOLERANCE)
                model.shared_energy_storage_operation.add(sdch ** 2 >= pdch ** 2 + qdch ** 2 - EQUALITY_TOLERANCE)

            # Charging/discharging complementarity constraints
            if params.slacks.shared_ess.complementarity:
                model.shared_energy_storage_ch_dch_exclusion.add(sch * sdch == model.slack_shared_es_comp[e, s_o])
            else:
                model.shared_energy_storage_ch_dch_exclusion.add(sch * sdch <= EQUALITY_TOLERANCE)

            # State-of-Charge
            soc_prev = soc_init
            model.shared_energy_storage_balance.add(model.shared_es_soc[e, s_o] <= soc_prev + (sch * eff_charge - sdch / eff_discharge) + EQUALITY_TOLERANCE)
            model.shared_energy_storage_balance.add(model.shared_es_soc[e, s_o] >= soc_prev + (sch * eff_charge - sdch / eff_discharge) - EQUALITY_TOLERANCE)

    # - Node Balance constraints
    model.node_balance_cons_p = pe.ConstraintList()
    model.node_balance_cons_q = pe.ConstraintList()
    for s_o in model.scenarios_operation:
        for i in range(len(network.nodes)):

            node = network.nodes[i]

            Pd = 0.00
            Qd = 0.00
            for c in model.loads:
                if network.loads[c].bus == node.bus_i:
                    Pd += model.pc[c, s_o]
                    Qd += model.qc[c, s_o]
                    if params.fl_reg and network.loads[c].fl_reg:
                        Pd += (model.flex_p_up[c, s_o] - model.flex_p_down[c, s_o])
                    if params.l_curt:
                        Pd -= (model.pc_curt_down[c, s_o] - model.pc_curt_up[c, s_o])
                        Qd -= (model.qc_curt_down[c, s_o] - model.qc_curt_up[c, s_o])
            if params.es_reg:
                for e in model.energy_storages:
                    if network.energy_storages[e].bus == node.bus_i:
                        Pd += (model.es_pch[e, s_o] - model.es_pdch[e, s_o])
                        Qd += (model.es_qch[e, s_o] - model.es_qdch[e, s_o])
            for e in model.shared_energy_storages:
                if network.shared_energy_storages[e].bus == node.bus_i:
                    Pd += (model.shared_es_pch[e, s_o] - model.shared_es_pdch[e, s_o])
                    Qd += (model.shared_es_qch[e, s_o] - model.shared_es_qdch[e, s_o])

            Pg = 0.0
            Qg = 0.0
            for g in model.generators:
                generator = network.generators[g]
                if generator.bus == node.bus_i:
                    Pg += model.pg[g, s_o]
                    Qg += model.qg[g, s_o]

            ei = model.e_actual[i, s_o]
            fi = model.f_actual[i, s_o]

            Pi = node.gs * (ei ** 2 + fi ** 2)
            Qi = -node.bs * (ei ** 2 + fi ** 2)
            for b in range(len(network.branches)):
                branch = network.branches[b]
                if branch.fbus == node.bus_i or branch.tbus == node.bus_i:

                    rij = model.r[b, s_o]
                    if not branch.is_transformer:
                        rij = 1.00

                    if branch.fbus == node.bus_i:
                        fnode_idx = network.get_node_idx(branch.fbus)
                        tnode_idx = network.get_node_idx(branch.tbus)

                        ei = model.e_actual[fnode_idx, s_o]
                        fi = model.f_actual[fnode_idx, s_o]
                        ej = model.e_actual[tnode_idx, s_o]
                        fj = model.f_actual[tnode_idx, s_o]

                        Pi += branch.g * (ei ** 2 + fi ** 2) * rij ** 2
                        Pi -= rij * (branch.g * (ei * ej + fi * fj) + branch.b * (fi * ej - ei * fj))
                        Qi -= (branch.b + branch.b_sh * 0.5) * (ei ** 2 + fi ** 2) * rij ** 2
                        Qi += rij * (branch.b * (ei * ej + fi * fj) - branch.g * (fi * ej - ei * fj))
                    else:
                        fnode_idx = network.get_node_idx(branch.tbus)
                        tnode_idx = network.get_node_idx(branch.fbus)

                        ei = model.e_actual[fnode_idx, s_o]
                        fi = model.f_actual[fnode_idx, s_o]
                        ej = model.e_actual[tnode_idx, s_o]
                        fj = model.f_actual[tnode_idx, s_o]

                        Pi += branch.g * (ei ** 2 + fi ** 2)
                        Pi -= rij * (branch.g * (ei * ej + fi * fj) + branch.b * (fi * ej - ei * fj))
                        Qi -= (branch.b + branch.b_sh * 0.5) * (ei ** 2 + fi ** 2)
                        Qi += rij * (branch.b * (ei * ej + fi * fj) - branch.g * (fi * ej - ei * fj))

            if params.slacks.node_balance:
                model.node_balance_cons_p.add(Pg == Pd + Pi + model.slack_node_balance_p[i, s_o])
                model.node_balance_cons_q.add(Qg == Qd + Qi + model.slack_node_balance_q[i, s_o])
            else:
                model.node_balance_cons_p.add(Pg <= Pd + Pi + EQUALITY_TOLERANCE)
                model.node_balance_cons_p.add(Pg >= Pd + Pi - EQUALITY_TOLERANCE)
                model.node_balance_cons_q.add(Qg <= Qd + Qi + EQUALITY_TOLERANCE)
                model.node_balance_cons_q.add(Qg >= Qd + Qi - EQUALITY_TOLERANCE)

    # - Branch Power Flow constraints (current)
    model.branch_power_flow_cons = pe.ConstraintList()
    model.branch_power_flow_lims = pe.ConstraintList()
    for s_o in model.scenarios_operation:
        for b in model.branches:

            branch = network.branches[b]
            rating = branch.rate / network.baseMVA
            if rating == 0.0:
                rating = BRANCH_UNKNOWN_RATING
            fnode_idx = network.get_node_idx(branch.fbus)
            tnode_idx = network.get_node_idx(branch.tbus)

            rij = model.r[b, s_o]
            if not branch.is_transformer:
                rij = 1.00
            ei = model.e_actual[fnode_idx, s_o]
            fi = model.f_actual[fnode_idx, s_o]
            ej = model.e_actual[tnode_idx, s_o]
            fj = model.f_actual[tnode_idx, s_o]

            flow_ij_sqr = 0.00

            if params.branch_limit_type == BRANCH_LIMIT_CURRENT:

                bij_sh = branch.b_sh * 0.50

                iij_sqr = (branch.g ** 2 + branch.b ** 2) * (((rij ** 2) * ei - rij * ej) ** 2 + ((rij ** 2) * fi - rij * fj) ** 2)
                iij_sqr += bij_sh ** 2 * (ei ** 2 + fi ** 2)
                iij_sqr += 2 * branch.g * bij_sh * (((rij ** 2) * fi - rij * fj) * ei - ((rij ** 2) * ei - rij * ej) * fi)
                iij_sqr += 2 * branch.b * bij_sh * (((rij ** 2) * ei - rij * ej) * ei + ((rij ** 2) * fi - rij * fj) * fi)
                flow_ij_sqr = iij_sqr

                # Previous (approximation?)
                # iji_sqr = (branch.g ** 2 + branch.b ** 2) * ((ej - rij * ei) ** 2 + (fj - rij * fi) ** 2)
                # iji_sqr += bij_sh ** 2 * (ej ** 2 + fj ** 2)
                # iji_sqr += 2 * branch.g * bij_sh * ((fj - rij * fi) * ej - (ej - rij * ei) * fj)
                # iji_sqr += 2 * branch.b * bij_sh * ((ej - rij * ei) * ej + (fj - rij * fi) * fj)

            elif params.branch_limit_type == BRANCH_LIMIT_APPARENT_POWER:

                pij = branch.g * (ei ** 2 + fi ** 2) * rij ** 2
                pij -= branch.g * (ei * ej + fi * fj) * rij
                pij -= branch.b * (fi * ej - ei * fj) * rij
                qij = - (branch.b + branch.b_sh * 0.50) * (ei ** 2 + fi ** 2) * rij ** 2
                qij += branch.b * (ei * ej + fi * fj) * rij
                qij -= branch.g * (fi * ej - ei * fj) * rij
                sij_sqr = pij ** 2 + qij ** 2
                flow_ij_sqr = sij_sqr

                # Without rij
                # pji = branch.g * (ej ** 2 + fj ** 2)
                # pji -= branch.g * (ej * ei + fj * fi) * rij
                # pji -= branch.b * (fj * ei - ej * fi) * rij
                # qji = - (branch.b + branch.b_sh * 0.50) * (ej ** 2 + fj ** 2)
                # qji += branch.b * (ej * ei + fj * fi) * rij
                # qji -= branch.g * (fj * ei - ej * fi) * rij
                # sji_sqr = pji ** 2 + qji ** 2

            elif params.branch_limit_type == BRANCH_LIMIT_MIXED:

                if branch.is_transformer:
                    pij = branch.g * (ei ** 2 + fi ** 2) * rij ** 2
                    pij -= branch.g * (ei * ej + fi * fj) * rij
                    pij -= branch.b * (fi * ej - ei * fj) * rij
                    qij = - (branch.b + branch.b_sh * 0.50) * (ei ** 2 + fi ** 2) * rij ** 2
                    qij += branch.b * (ei * ej + fi * fj) * rij
                    qij -= branch.g * (fi * ej - ei * fj) * rij
                    sij_sqr = pij ** 2 + qij ** 2
                    flow_ij_sqr = sij_sqr
                else:
                    bij_sh = branch.b_sh * 0.50
                    iij_sqr = (branch.g ** 2 + branch.b ** 2) * (((rij ** 2) * ei - rij * ej) ** 2 + ((rij ** 2) * fi - rij * fj) ** 2)
                    iij_sqr += bij_sh ** 2 * (ei ** 2 + fi ** 2)
                    iij_sqr += 2 * branch.g * bij_sh * (((rij ** 2) * fi - rij * fj) * ei - ((rij ** 2) * ei - rij * ej) * fi)
                    iij_sqr += 2 * branch.b * bij_sh * (((rij ** 2) * ei - rij * ej) * ei + ((rij ** 2) * fi - rij * fj) * fi)
                    flow_ij_sqr = iij_sqr

            # Flow_ij, definition
            model.branch_power_flow_cons.add(model.flow_ij_sqr[b, s_o] <= flow_ij_sqr + EQUALITY_TOLERANCE)
            model.branch_power_flow_cons.add(model.flow_ij_sqr[b, s_o] >= flow_ij_sqr - EQUALITY_TOLERANCE)

            # Branch flow limits
            if branch.status:
                if params.slacks.grid_operation.branch_flow:
                    model.branch_power_flow_lims.add(model.flow_ij_sqr[b, s_o] <= rating ** 2 + model.slack_flow_ij_sqr[b, s_o])
                else:
                    model.branch_power_flow_lims.add(model.flow_ij_sqr[b, s_o] <= rating ** 2)

    # ------------------------------------------------------------------------------------------------------------------
    # Costs (penalties)
    # Note: defined as variables (bus fixed) so that they can be changed later, if needed
    model.penalty_ess_usage = pe.Var(domain=pe.NonNegativeReals)
    model.penalty_ess_usage.fix(PENALTY_ESS_USAGE)
    if params.obj_type == OBJ_MIN_COST:
        model.cost_res_curtailment = pe.Var(domain=pe.NonNegativeReals)
        model.cost_load_curtailment = pe.Var(domain=pe.NonNegativeReals)
        model.cost_res_curtailment.fix(COST_GENERATION_CURTAILMENT)
        model.cost_load_curtailment.fix(COST_CONSUMPTION_CURTAILMENT)
    elif params.obj_type == OBJ_CONGESTION_MANAGEMENT:
        model.penalty_gen_curtailment = pe.Var(domain=pe.NonNegativeReals)
        model.penalty_load_curtailment = pe.Var(domain=pe.NonNegativeReals)
        model.penalty_flex_usage = pe.Var(domain=pe.NonNegativeReals)
        model.penalty_gen_curtailment.fix(PENALTY_GENERATION_CURTAILMENT)
        model.penalty_load_curtailment.fix(PENALTY_LOAD_CURTAILMENT)
        model.penalty_flex_usage.fix(PENALTY_FLEXIBILITY_USAGE)
    else:
        print(f'[ERROR] Unrecognized or invalid objective. Objective = {params.obj_type}. Exiting...')
        exit(ERROR_NETWORK_MODEL)

    # Objective Function
    obj = 0.0
    if params.obj_type == OBJ_MIN_COST:

        # Cost minimization
        c_p = network.cost_energy_p
        c_flex = network.cost_flex
        for s_o in model.scenarios_operation:

            obj_scenario = 0.0
            omega_oper = network.prob_operation_scenarios[s_o]

            # Generation
            for g in model.generators:
                if network.generators[g].is_controllable():
                    if (not network.is_transmission) and network.generators[g].gen_type == GEN_REFERENCE:
                        continue
                    pg = model.pg[g, s_o]
                    obj_scenario += c_p[n] * network.baseMVA * pg

            # Demand side flexibility
            if params.fl_reg:
                for c in model.loads:
                    flex_p_up = model.flex_p_up[c, s_o]
                    flex_p_down = model.flex_p_down[c, s_o]
                    obj_scenario += c_flex[n] * network.baseMVA * (flex_p_down + flex_p_up)

            # Load curtailment
            if params.l_curt:
                for c in model.loads:
                    pc_curt = (model.pc_curt_down[c, s_o] + model.pc_curt_up[c, s_o])
                    qc_curt = (model.qc_curt_down[c, s_o] + model.qc_curt_up[c, s_o])
                    obj_scenario += model.cost_load_curtailment * network.baseMVA * (pc_curt + qc_curt)

            # Generation curtailment
            if params.rg_curt:
                for g in model.generators:
                    if network.generators[g].is_curtaillable():
                        sg_curt = model.sg_curt[g, s_o]
                        obj_scenario += model.cost_res_curtailment * network.baseMVA * sg_curt

            # ESS utilization
            if params.es_reg:
                for e in model.energy_storages:
                    sch = model.es_sch[e, s_o]
                    sdch = model.es_sdch[e, s_o]
                    obj_scenario += model.penalty_ess_usage * network.baseMVA * (sch + sdch)

            # Shared ESS utilization
            for e in model.shared_energy_storages:
                sch = model.shared_es_sch[e, s_o]
                sdch = model.shared_es_sdch[e, s_o]
                obj_scenario += model.penalty_ess_usage * network.baseMVA * (sch + sdch)

            obj += obj_scenario * omega_oper
    elif params.obj_type == OBJ_CONGESTION_MANAGEMENT:

        # Congestion Management
        for s_o in model.scenarios_operation:

            omega_oper = network.prob_operation_scenarios[s_o]

            obj_scenario = 0.0

            # Load curtailment
            if params.l_curt:
                for c in model.loads:
                    pc_curt = (model.pc_curt_down[c, s_o] + model.pc_curt_up[c, s_o])
                    qc_curt = (model.qc_curt_down[c, s_o] + model.qc_curt_up[c, s_o])
                    obj_scenario += model.penalty_load_curtailment * network.baseMVA * (pc_curt + qc_curt)

            # Demand side flexibility
            if params.fl_reg:
                for c in model.loads:
                    flex_p_up = model.flex_p_up[c, s_o]
                    flex_p_down = model.flex_p_down[c, s_o]
                    obj_scenario += model.penalty_flex_usage * network.baseMVA * (flex_p_down + flex_p_up)

            # ESS utilization
            if params.es_reg:
                for e in model.energy_storages:
                    sch = model.es_sch[e, s_o]
                    sdch = model.es_sdch[e, s_o]
                    obj_scenario += model.penalty_ess_usage * network.baseMVA * (sch + sdch)

            # Shared ESS utilization
            for e in model.shared_energy_storages:
                sch = model.shared_es_sch[e, s_o]
                sdch = model.shared_es_sdch[e, s_o]
                obj_scenario += model.penalty_ess_usage * network.baseMVA * (sch + sdch)

            obj += obj_scenario * omega_oper

    # Slacks grid operation
    for s_o in model.scenarios_operation:

        omega_oper = network.prob_operation_scenarios[s_o]

        # Voltage slacks
        if params.slacks.grid_operation.voltage:
            for i in model.nodes:
                slack_e_sqr = model.slack_e[i, s_o] ** 2
                slack_f_sqr = model.slack_f[i, s_o] ** 2
                obj += PENALTY_VOLTAGE * network.baseMVA * omega_oper * (slack_e_sqr + slack_f_sqr)

        # Branch power flow slacks
        if params.slacks.grid_operation.branch_flow:
            for b in model.branches:
                slack_flow_ij_sqr = (model.slack_flow_ij_sqr[b, s_o])
                obj += PENALTY_CURRENT * network.baseMVA * omega_oper * slack_flow_ij_sqr

    # Operation slacks
    for s_o in model.scenarios_operation:

        omega_oper = network.prob_operation_scenarios[s_o]

        # Node balance
        if params.slacks.node_balance:
            for i in model.nodes:
                slack_p_sqr = model.slack_node_balance_p[i, s_o] ** 2
                slack_q_sqr = model.slack_node_balance_q[i, s_o] ** 2
                obj += PENALTY_NODE_BALANCE * network.baseMVA * omega_oper * (slack_p_sqr + slack_q_sqr)

        # ESS slacks
        if params.es_reg:
            if params.slacks.ess.complementarity:
                slack_comp = model.slack_es_comp[e, s_o]
                obj += PENALTY_ESS * network.baseMVA * omega_oper * slack_comp
            if params.slacks.ess.charging:
                slack_ch_sqr = model.slack_es_ch[e, s_o] ** 2
                slack_dch_sqr = model.slack_es_dch[e, s_o] ** 2
                obj += PENALTY_ESS * network.baseMVA * omega_oper * (slack_ch_sqr + slack_dch_sqr)

        # Shared ESS slacks
        for e in model.shared_energy_storages:
            if params.slacks.shared_ess.complementarity:
                slack_comp = model.slack_shared_es_comp[e, s_o]
                obj += PENALTY_SHARED_ESS * network.baseMVA * omega_oper * slack_comp
            if params.slacks.shared_ess.charging:
                slack_ch_sqr = model.slack_shared_es_ch[e, s_o] ** 2
                slack_dch_sqr = model.slack_shared_es_dch[e, s_o] ** 2
                obj += PENALTY_SHARED_ESS * network.baseMVA * omega_oper * (slack_ch_sqr + slack_dch_sqr)

    model.objective = pe.Objective(sense=pe.minimize, expr=obj)

    # Model suffixes (used for warm start)
    model.ipopt_zL_out = pe.Suffix(direction=pe.Suffix.IMPORT)  # Ipopt bound multipliers (obtained from solution)
    model.ipopt_zU_out = pe.Suffix(direction=pe.Suffix.IMPORT)
    model.ipopt_zL_in = pe.Suffix(direction=pe.Suffix.EXPORT)  # Ipopt bound multipliers (sent to solver)
    model.ipopt_zU_in = pe.Suffix(direction=pe.Suffix.EXPORT)
    model.dual = pe.Suffix(direction=pe.Suffix.IMPORT_EXPORT)  # Obtain dual solutions from previous solve and send to warm start

    return model


def _build_pq_map_model(network):

    model = network.build_model()
    ref_gen_idx = network.get_reference_gen_idx()

    # Add expected interface power flow variables
    expected_pf_p = 0.00
    expected_pf_q = 0.00
    model.expected_interface_pf_p = pe.Var(domain=pe.Reals, initialize=0.00)
    model.expected_interface_pf_q = pe.Var(domain=pe.Reals, initialize=0.00)
    model.interface_expected_values = pe.ConstraintList()
    for s_o in model.scenarios_operation:
        omega_oper = network.prob_operation_scenarios[s_o]
        expected_pf_p += omega_oper * model.pg[ref_gen_idx, s_o]
        expected_pf_q += omega_oper * model.qg[ref_gen_idx, s_o]
    model.interface_expected_values.add(model.expected_interface_pf_p <= expected_pf_p + EQUALITY_TOLERANCE)
    model.interface_expected_values.add(model.expected_interface_pf_p >= expected_pf_p - EQUALITY_TOLERANCE)
    model.interface_expected_values.add(model.expected_interface_pf_q <= expected_pf_q + EQUALITY_TOLERANCE)
    model.interface_expected_values.add(model.expected_interface_pf_q >= expected_pf_q - EQUALITY_TOLERANCE)

    # New objective function (PQ maps)
    obj = 0.00
    model.alpha = pe.Var(domain=pe.NonNegativeReals, initialize=0.00, bounds=(0.00, 1.00))
    for s_o in model.scenarios_operation:
        omega_oper = network.prob_operation_scenarios[s_o]
        obj += model.alpha * model.pg[ref_gen_idx, s_o] * omega_oper
        obj += (1 - model.alpha) * model.qg[ref_gen_idx, s_o] * omega_oper

    # Regularization -- Added to OF to minimize deviations from scenarios to expected values
    s_base = network.baseMVA
    ref_gen_idx = network.get_reference_gen_idx()
    model.penalty_regularization = pe.Var(domain=pe.NonNegativeReals)
    model.penalty_regularization.fix(PENALTY_REGULARIZATION)
    for s_o in model.scenarios_operation:
        obj += model.penalty_regularization * s_base * (model.pg[ref_gen_idx, s_o] - model.expected_interface_pf_p) ** 2
        obj += model.penalty_regularization * s_base * (model.qg[ref_gen_idx, s_o] - model.expected_interface_pf_q) ** 2

    model.objective.expr = obj

    return model


def _optimize(network, model, from_warm_start=False):

    params = network.params
    solver = po.SolverFactory(params.solver_params.solver, executable=params.solver_params.solver_path)

    if from_warm_start:
        model.ipopt_zL_in.update(model.ipopt_zL_out)
        model.ipopt_zU_in.update(model.ipopt_zU_out)
        solver.options['warm_start_init_point'] = 'yes'
        solver.options['warm_start_bound_push'] = 1e-9
        solver.options['warm_start_bound_frac'] = 1e-9
        solver.options['warm_start_slack_bound_frac'] = 1e-9
        solver.options['warm_start_slack_bound_push'] = 1e-9
        solver.options['warm_start_mult_bound_push'] = 1e-9

    if params.solver_params.verbose:
        solver.options['print_level'] = 6
        solver.options['output_file'] = 'optim_log.txt'

    if params.solver_params.solver == 'ipopt':
        solver.options['tol'] = params.solver_params.solver_tol
        solver.options['linear_solver'] = params.solver_params.linear_solver
        solver.options['mu_strategy'] = 'adaptive'

    result = solver.solve(model, tee=params.solver_params.verbose)

    if params.solver_params.verbose:
        import logging
        from pyomo.util.infeasible import log_infeasible_constraints

        # Create a logger object with DEBUG level
        logging_logger = logging.getLogger()
        logging_logger.setLevel(logging.DEBUG)

        # Create a console handler
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)

        # add the handler to the logger
        logging_logger.addHandler(ch)

        # Log the infeasible constraints of pyomo object
        print("Displaying Infeasible Constraints")
        log_infeasible_constraints(model, log_expression=True, log_variables=True, logger=logging_logger)

    return result


# ======================================================================================================================
#  NETWORK DATA read functions
# ======================================================================================================================
def _read_network_data(network, operational_data_filename):

    # Read info from file(s)
    network.read_network_from_json_file(f'{network.name}.json')
    network.read_network_operational_data_from_file(operational_data_filename)

    if network.params.print_to_screen:
        network.print_network_to_screen()
    if network.params.plot_diagram:
        network.plot_diagram()


def _read_network_from_json_file(network, filename):

    network_data = convert_json_to_dict(read_json_file(filename))

    # Network base
    network.baseMVA = float(network_data['baseMVA'])

    # Nodes
    for node_data in network_data['nodes']:
        node = Node()
        node.bus_i = int(node_data['bus_i'])
        node.type = int(node_data['type'])
        node.gs = float(node_data['Gs']) / network.baseMVA
        node.bs = float(node_data['Bs']) / network.baseMVA
        node.base_kv = float(node_data['baseKV'])
        node.v_max = float(node_data['Vmax'])
        node.v_min = float(node_data['Vmin'])
        network.nodes.append(node)

    # Generators
    for gen_data in network_data['generators']:
        generator = Generator()
        generator.gen_id = int(gen_data['gen_id'])
        generator.bus = int(gen_data['bus'])
        if not network.node_exists(generator.bus):
            print(f'[ERROR] Generator {generator.gen_id}. Node {generator.bus} does not exist! Exiting...')
            exit(ERROR_NETWORK_FILE)
        generator.pmax = float(gen_data['Pmax']) / network.baseMVA
        generator.pmin = float(gen_data['Pmin']) / network.baseMVA
        generator.qmax = float(gen_data['Qmax']) / network.baseMVA
        generator.qmin = float(gen_data['Qmin']) / network.baseMVA
        generator.vg = float(gen_data['Vg'])
        generator.status = bool(gen_data['status'])
        gen_type = gen_data['type']
        if gen_type == 'REF':
            generator.gen_type = GEN_REFERENCE
        elif gen_type == 'CONV':
            generator.gen_type = GEN_CONV
        elif gen_type == 'PV':
            generator.gen_type = GEN_RES_SOLAR
        elif gen_type == 'WIND':
            generator.gen_type = GEN_RES_WIND
        elif gen_type == 'RES_OTHER':
            generator.gen_type = GEN_RES_OTHER
        elif gen_type == 'RES_CONTROLLABLE':
            generator.gen_type = GEN_RES_CONTROLLABLE
        if 'pf_control' in gen_data:
            generator.power_factor_control = bool(gen_data['pf_control'])
            if 'pf_max' in gen_data:
                generator.max_pf = float(gen_data['pf_max'])
            if 'pf_min' in gen_data:
                generator.min_pf = float(gen_data['pf_min'])
        network.generators.append(generator)

    # Loads
    for load_data in network_data['loads']:
        load = Load()
        load.load_id = int(load_data['load_id'])
        load.bus = int(load_data['bus'])
        if not network.node_exists(load.bus):
            print(f'[ERROR] Load {load.load_id }. Node {load.bus} does not exist! Exiting...')
            exit(ERROR_NETWORK_FILE)
        load.status = bool(load_data['status'])
        load.fl_reg = bool(load_data['fl_reg'])
        network.loads.append(load)

    # Lines
    for line_data in network_data['lines']:
        branch = Branch()
        branch.branch_id = int(line_data['branch_id'])
        branch.fbus = int(line_data['fbus'])
        if not network.node_exists(branch.fbus):
            print(f'[ERROR] Line {branch.branch_id }. Node {branch.fbus} does not exist! Exiting...')
            exit(ERROR_NETWORK_FILE)
        branch.tbus = int(line_data['tbus'])
        if not network.node_exists(branch.tbus):
            print(f'[ERROR] Line {branch.branch_id }. Node {branch.tbus} does not exist! Exiting...')
            exit(ERROR_NETWORK_FILE)
        branch.r = float(line_data['r'])
        branch.x = float(line_data['x'])
        branch.b_sh = float(line_data['b'])
        branch.rate = float(line_data['rating'])
        branch.status = bool(line_data['status'])
        network.branches.append(branch)

    # Transformers
    if 'transformers' in network_data:
        for transf_data in network_data['transformers']:
            branch = Branch()
            branch.branch_id = int(transf_data['branch_id'])
            branch.fbus = int(transf_data['fbus'])
            if not network.node_exists(branch.fbus):
                print(f'[ERROR] Transformer {branch.branch_id}. Node {branch.fbus} does not exist! Exiting...')
                exit(ERROR_NETWORK_FILE)
            branch.tbus = int(transf_data['tbus'])
            if not network.node_exists(branch.tbus):
                print(f'[ERROR] Transformer {branch.branch_id}. Node {branch.tbus} does not exist! Exiting...')
                exit(ERROR_NETWORK_FILE)
            branch.r = float(transf_data['r'])
            branch.x = float(transf_data['x'])
            branch.b_sh = float(transf_data['b'])
            branch.rate = float(transf_data['rating'])
            branch.ratio = float(transf_data['ratio'])
            branch.status = bool(transf_data['status'])
            branch.is_transformer = True
            branch.vmag_reg = bool(transf_data['vmag_reg'])
            network.branches.append(branch)

    # Energy Storages
    if 'energy_storages' in network_data:
        for energy_storage_data in network_data['energy_storages']:
            energy_storage = EnergyStorage()
            energy_storage.es_id = int(energy_storage_data['es_id'])
            energy_storage.bus = int(energy_storage_data['bus'])
            if not network.node_exists(energy_storage.bus):
                print(f'[ERROR] Energy Storage {energy_storage.es_id}. Node {energy_storage.bus} does not exist! Exiting...')
                exit(ERROR_NETWORK_FILE)
            energy_storage.s = float(energy_storage_data['s']) / network.baseMVA
            energy_storage.e = float(energy_storage_data['e']) / network.baseMVA
            energy_storage.e_init = energy_storage.e * ENERGY_STORAGE_RELATIVE_INIT_SOC
            energy_storage.e_min = energy_storage.e * ENERGY_STORAGE_MIN_ENERGY_STORED
            energy_storage.e_max = energy_storage.e * ENERGY_STORAGE_MAX_ENERGY_STORED
            energy_storage.eff_ch = float(energy_storage_data['eff_ch'])
            energy_storage.eff_dch = float(energy_storage_data['eff_dch'])
            energy_storage.max_pf = float(energy_storage_data['max_pf'])
            energy_storage.min_pf = float(energy_storage_data['min_pf'])
            network.energy_storages.append(energy_storage)


# ======================================================================================================================
#   NETWORK OPERATIONAL DATA read functions
# ======================================================================================================================
def _read_network_operational_data_from_file(network, filename):

    data = {
        'consumption': {
            'pc': dict(), 'qc': dict()
        },
        'flexibility': {
            'upward': dict(),
            'downward': dict()
        },
        'generation': {
            'pg': dict(), 'qg': dict(), 'status': list()
        }
    }

    # Scenario information
    num_oper_scenarios, prob_oper_scenarios = _get_operational_scenario_info_from_excel_file(filename, 'Main')
    network.prob_operation_scenarios = prob_oper_scenarios

    # Consumption and Generation data -- by scenario
    for i in range(len(network.prob_operation_scenarios)):

        sheet_name_pc = f'Pc, S{i + 1}'
        sheet_name_qc = f'Qc, S{i + 1}'
        sheet_name_pg = f'Pg, S{i + 1}'
        sheet_name_qg = f'Qg, S{i + 1}'

        # Consumption per scenario (active, reactive power)
        pc_scenario = _get_consumption_flexibility_data_from_excel_file(filename, sheet_name_pc)
        qc_scenario = _get_consumption_flexibility_data_from_excel_file(filename, sheet_name_qc)
        if not pc_scenario:
            print(f'[ERROR] Network {network.name}. No active power consumption data provided for scenario {i + 1}. Exiting...')
            exit(ERROR_OPERATIONAL_DATA_FILE)
        if not qc_scenario:
            print(f'[ERROR] Network {network.name}. No reactive power consumption data provided for scenario {i + 1}. Exiting...')
            exit(ERROR_OPERATIONAL_DATA_FILE)
        data['consumption']['pc'][i] = pc_scenario
        data['consumption']['qc'][i] = qc_scenario

        # Generation per scenario (active, reactive power)
        num_renewable_gens = network.get_num_renewable_gens()
        if num_renewable_gens > 0:
            pg_scenario = _get_generation_data_from_excel_file(filename, sheet_name_pg)
            qg_scenario = _get_generation_data_from_excel_file(filename, sheet_name_qg)
            if not pg_scenario:
                print(f'[ERROR] Network {network.name}. No active power generation data provided for scenario {i + 1}. Exiting...')
                exit(ERROR_OPERATIONAL_DATA_FILE)
            if not qg_scenario:
                print(f'[ERROR] Network {network.name}. No reactive power generation data provided for scenario {i + 1}. Exiting...')
                exit(ERROR_OPERATIONAL_DATA_FILE)
            data['generation']['pg'][i] = pg_scenario
            data['generation']['qg'][i] = qg_scenario

    # Generators status. Note: common to all scenarios
    data['generation']['status'] = _get_generator_status_from_excel_file(filename, f'GenStatus')

    # Flexibility data
    flex_up_p = _get_consumption_flexibility_data_from_excel_file(filename, f'UpFlex')
    if not flex_up_p:
        for load in network.loads:
            flex_up_p[load.load_id] = [0.0 for _ in range(network.num_instants)]
    data['flexibility']['upward'] = flex_up_p

    flex_down_p = _get_consumption_flexibility_data_from_excel_file(filename, f'DownFlex')
    if not flex_down_p:
        for load in network.loads:
            flex_down_p[load.load_id] = [0.0 for _ in range(network.num_instants)]
    data['flexibility']['downward'] = flex_down_p

    return data


def _get_operational_scenario_info_from_excel_file(filename, sheet_name):

    num_scenarios = 0
    prob_scenarios = list()

    try:
        df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
        if is_int(df.iloc[0, 1]):
            num_scenarios = int(df.iloc[0, 1])
        else:
            print(f'[ERROR] File {filename}. Num scenarios should be an int!')
            exit(ERROR_OPERATIONAL_DATA_FILE)
        for i in range(num_scenarios):
            if is_float(df.iloc[0, i+2]):
                prob_scenarios.append(float(df.iloc[0, i+2]))
            else:
                print(f'[ERROR] File {filename}. Scenario probability should be a float!')
                exit(ERROR_OPERATIONAL_DATA_FILE)
    except:
        print(f'[ERROR] Workbook {filename}. Sheet {sheet_name} does not exist.')
        exit(ERROR_OPERATIONAL_DATA_FILE)

    if num_scenarios != len(prob_scenarios):
        print(f'[WARNING] Workbook {filename}. Data file. Number of scenarios different from the probability vector!')

    if round(sum(prob_scenarios), 2) != 1.00:
        print(f'[ERROR] Workbook {filename}. Probability of scenarios does not add up to 100%.')
        exit(ERROR_OPERATIONAL_DATA_FILE)

    return num_scenarios, prob_scenarios


def _get_consumption_flexibility_data_from_excel_file(filename, sheet_name):

    try:
        data = pd.read_excel(filename, sheet_name=sheet_name)
        num_rows, num_cols = data.shape
        processed_data = dict()
        for i in range(num_rows):
            node_id = data.iloc[i, 0]
            processed_data[node_id] = [0.0 for _ in range(num_cols - 1)]
        for node_id in processed_data:
            node_values = [0.0 for _ in range(num_cols - 1)]
            for i in range(0, num_rows):
                aux_node_id = data.iloc[i, 0]
                if aux_node_id == node_id:
                    for j in range(0, num_cols - 1):
                        node_values[j] += data.iloc[i, j + 1]
            processed_data[node_id] = node_values
    except:
        print(f'[WARNING] Workbook {filename}. Sheet {sheet_name} does not exist.')
        processed_data = {}

    return processed_data


def _get_generation_data_from_excel_file(filename, sheet_name):

    try:
        data = pd.read_excel(filename, sheet_name=sheet_name)
        num_rows, num_cols = data.shape
        processed_data = dict()
        for i in range(num_rows):
            gen_id = data.iloc[i, 0]
            processed_data[gen_id] = [0.0 for _ in range(num_cols - 1)]
        for gen_id in processed_data:
            processed_data_gen = [0.0 for _ in range(num_cols - 1)]
            for i in range(0, num_rows):
                aux_node_id = data.iloc[i, 0]
                if aux_node_id == gen_id:
                    for j in range(0, num_cols - 1):
                        processed_data_gen[j] += data.iloc[i, j + 1]
            processed_data[gen_id] = processed_data_gen
    except:
        print(f'[WARNING] Workbook {filename}. Sheet {sheet_name} does not exist.')
        processed_data = {}

    return processed_data


def _get_generator_status_from_excel_file(filename, sheet_name):

    try:
        data = pd.read_excel(filename, sheet_name=sheet_name)
        num_rows, num_cols = data.shape
        status_values = dict()
        for i in range(num_rows):
            gen_id = data.iloc[i, 0]
            status_values[gen_id] = list()
            for j in range(0, num_cols - 1):
                status_values[gen_id].append(bool(data.iloc[i, j + 1]))
    except:
        print(f'[WARNING] Workbook {filename}. Sheet {sheet_name} does not exist.')
        status_values = list()

    return status_values


def _update_network_with_excel_data(network, data):

    for load in network.loads:

        load_id = load.load_id
        load.pd = dict()         # Note: Changes Pd and Qd fields to dicts (per scenario)
        load.qd = dict()

        for s in range(len(network.prob_operation_scenarios)):
            pc = _get_consumption_from_data(data, load_id, network.num_instants, s, DATA_ACTIVE_POWER)
            qc = _get_consumption_from_data(data, load_id, network.num_instants, s, DATA_REACTIVE_POWER)
            load.pd[s] = [instant / network.baseMVA for instant in pc]
            load.qd[s] = [instant / network.baseMVA for instant in qc]
        flex_up_p = _get_flexibility_from_data(data, load_id, network.num_instants, DATA_UPWARD_FLEXIBILITY)
        flex_down_p = _get_flexibility_from_data(data, load_id, network.num_instants, DATA_DOWNWARD_FLEXIBILITY)
        load.flexibility.upward = [p / network.baseMVA for p in flex_up_p]
        load.flexibility.downward = [q / network.baseMVA for q in flex_down_p]

    for generator in network.generators:

        generator.pg = dict()  # Note: Changes Pg and Qg fields to dicts (per scenario)
        generator.qg = dict()

        # Active and Reactive power
        for s in range(len(network.prob_operation_scenarios)):
            if generator.gen_type in GEN_CURTAILLABLE_TYPES:
                pg = _get_generation_from_data(data, generator.gen_id, s, DATA_ACTIVE_POWER)
                qg = _get_generation_from_data(data, generator.gen_id, s, DATA_REACTIVE_POWER)
                generator.pg[s] = [instant / network.baseMVA for instant in pg]
                generator.qg[s] = [instant / network.baseMVA for instant in qg]
            else:
                generator.pg[s] = [0.00 for _ in range(network.num_instants)]
                generator.qg[s] = [0.00 for _ in range(network.num_instants)]

        # Status
        if generator.gen_id in data['generation']['status']:
            generator.status = data['generation']['status'][generator.gen_id]
        else:
            generator.status = [generator.status for _ in range(network.num_instants)]

    network.data_loaded = True


def _get_consumption_from_data(data, node_id, num_instants, idx_scenario, type):

    if type == DATA_ACTIVE_POWER:
        power_label = 'pc'
    else:
        power_label = 'qc'

    for node in data['consumption'][power_label][idx_scenario]:
        if node == node_id:
            return data['consumption'][power_label][idx_scenario][node_id]

    consumption = [0.0 for _ in range(num_instants)]

    return consumption


def _get_flexibility_from_data(data, node_id, num_instants, flex_type):

    flex_label = str()

    if flex_type == DATA_UPWARD_FLEXIBILITY:
        flex_label = 'upward'
    elif flex_type == DATA_DOWNWARD_FLEXIBILITY:
        flex_label = 'downward'
    elif flex_type == DATA_COST_FLEXIBILITY:
        flex_label = 'cost'
    else:
        print('[ERROR] Unrecognized flexibility type in get_flexibility_from_data. Exiting.')
        exit(1)

    for node in data['flexibility'][flex_label]:
        if node == node_id:
            return data['flexibility'][flex_label][node_id]

    flex = [0.0 for _ in range(num_instants)]   # Returns empty flexibility vector

    return flex


def _get_generation_from_data(data, gen_id, idx_scenario, type):

    if type == DATA_ACTIVE_POWER:
        power_label = 'pg'
    else:
        power_label = 'qg'

    return data['generation'][power_label][idx_scenario][gen_id]


# ======================================================================================================================
#   NETWORK RESULTS functions
# ======================================================================================================================
def _process_results(network, model, results=dict(), n=0):

    params = network.params
    s_base = network.baseMVA

    processed_results = dict()
    processed_results['obj'] = _compute_objective_function_value(network, model, params)
    processed_results['gen_cost'] = _compute_generation_cost(network, model)
    processed_results['total_load'] = _compute_total_load(network, model, params)
    processed_results['total_gen'] = _compute_total_generation(network, model, params)
    processed_results['total_conventional_gen'] = _compute_conventional_generation(network, model)
    processed_results['total_renewable_gen'] = _compute_renewable_generation(network, model, params)
    processed_results['losses'] = _compute_losses(network, model, params)
    processed_results['gen_curt'] = _compute_generation_curtailment(network, model, params)
    processed_results['load_curt'] = _compute_load_curtailment(network, model, params)
    processed_results['flex_used'] = _compute_flexibility_used(network, model, params)
    if results:
        processed_results['runtime'] = float(_get_info_from_results(results, 'Time:').strip()),

    processed_results['scenarios'] = dict()
    for s_o in model.scenarios_operation:

        processed_results['scenarios'][s_o] = {
            'voltage': {'vmag': {}, 'vang': {}},
            'consumption': {'pc': {}, 'qc': {}, 'pc_net': {}, 'qc_net': {}},
            'generation': {'pg': {}, 'qg': {}, 'sg': {}},
            'branches': {'power_flow': {'pij': {}, 'pji': {}, 'qij': {}, 'qji': {}, 'sij': {}, 'sji': {}},
                         'losses': {}, 'ratio': {}, 'branch_flow': {'flow_ij_perc': {}}},
            'energy_storages': {'p': {}, 'q': {}, 's': {}, 'soc': {}, 'soc_percent': {}},
            'shared_energy_storages': {'p': {}, 'q': {}, 's': {}, 'soc': {}, 'soc_percent': {}}
        }

        if params.transf_reg:
            processed_results['scenarios'][s_o]['branches']['ratio'] = dict()

        if params.fl_reg:
            processed_results['scenarios'][s_o]['consumption']['p_up'] = dict()
            processed_results['scenarios'][s_o]['consumption']['p_down'] = dict()

        if params.l_curt:
            processed_results['scenarios'][s_o]['consumption']['pc_curt'] = dict()
            processed_results['scenarios'][s_o]['consumption']['qc_curt'] = dict()

        if params.rg_curt:
            processed_results['scenarios'][s_o]['generation']['pg_net'] = dict()
            processed_results['scenarios'][s_o]['generation']['qg_net'] = dict()
            processed_results['scenarios'][s_o]['generation']['sg_net'] = dict()
            processed_results['scenarios'][s_o]['generation']['sg_curt'] = dict()

        if params.es_reg:
            processed_results['scenarios'][s_o]['energy_storages']['p'] = dict()
            processed_results['scenarios'][s_o]['energy_storages']['q'] = dict()
            processed_results['scenarios'][s_o]['energy_storages']['s'] = dict()
            processed_results['scenarios'][s_o]['energy_storages']['soc'] = dict()
            processed_results['scenarios'][s_o]['energy_storages']['soc_percent'] = dict()

        processed_results['scenarios'][s_o]['relaxation_slacks'] = dict()
        processed_results['scenarios'][s_o]['relaxation_slacks']['voltage'] = dict()
        if params.slacks.grid_operation.voltage:
            processed_results['scenarios'][s_o]['relaxation_slacks']['voltage']['e'] = dict()
            processed_results['scenarios'][s_o]['relaxation_slacks']['voltage']['f'] = dict()
        processed_results['scenarios'][s_o]['relaxation_slacks']['branch_flow'] = dict()
        if params.slacks.grid_operation.branch_flow:
            processed_results['scenarios'][s_o]['relaxation_slacks']['branch_flow']['flow_ij_sqr'] = dict()
        processed_results['scenarios'][s_o]['relaxation_slacks']['node_balance'] = dict()
        if params.slacks.node_balance:
            processed_results['scenarios'][s_o]['relaxation_slacks']['node_balance']['p'] = dict()
            processed_results['scenarios'][s_o]['relaxation_slacks']['node_balance']['q'] = dict()
        processed_results['scenarios'][s_o]['relaxation_slacks']['shared_energy_storages'] = dict()
        if params.slacks.shared_ess.complementarity:
            processed_results['scenarios'][s_o]['relaxation_slacks']['shared_energy_storages']['comp'] = dict()
        if params.fl_reg:
            processed_results['scenarios'][s_o]['relaxation_slacks']['flexibility'] = dict()
        if params.es_reg:
            processed_results['scenarios'][s_o]['relaxation_slacks']['energy_storages'] = dict()
            if params.slacks.ess.complementarity:
                processed_results['scenarios'][s_o]['relaxation_slacks']['energy_storages']['comp'] = dict()

        # Voltage
        for i in model.nodes:
            node_id = network.nodes[i].bus_i
            e = pe.value(model.e_actual[i, s_o])
            f = pe.value(model.f_actual[i, s_o])
            v_mag = sqrt(e**2 + f**2)
            v_ang = atan2(f, e) * (180.0 / pi)
            processed_results['scenarios'][s_o]['voltage']['vmag'][node_id] = v_mag
            processed_results['scenarios'][s_o]['voltage']['vang'][node_id] = v_ang

        # Consumption
        for c in model.loads:
            load_id = network.loads[c].load_id
            processed_results['scenarios'][s_o]['consumption']['pc_net'][load_id] = 0.00
            processed_results['scenarios'][s_o]['consumption']['qc_net'][load_id] = 0.00
            pc = pe.value(model.pc[c, s_o]) * network.baseMVA
            qc = pe.value(model.qc[c, s_o]) * network.baseMVA
            processed_results['scenarios'][s_o]['consumption']['pc'][load_id] = pc
            processed_results['scenarios'][s_o]['consumption']['qc'][load_id] = qc
            processed_results['scenarios'][s_o]['consumption']['pc_net'][load_id] += pc
            processed_results['scenarios'][s_o]['consumption']['qc_net'][load_id] += qc
            if params.fl_reg:
                pup = pe.value(model.flex_p_up[c, s_o]) * network.baseMVA
                pdown = pe.value(model.flex_p_down[c, s_o]) * network.baseMVA
                processed_results['scenarios'][s_o]['consumption']['p_up'][load_id] = pup
                processed_results['scenarios'][s_o]['consumption']['p_down'][load_id] = pdown
                processed_results['scenarios'][s_o]['consumption']['pc_net'][load_id] += pup - pdown
            if params.l_curt:
                pc_curt = pe.value(model.pc_curt_down[c, s_o] - model.pc_curt_up[c, s_o]) * network.baseMVA
                qc_curt = pe.value(model.qc_curt_down[c, s_o] - model.qc_curt_up[c, s_o]) * network.baseMVA
                processed_results['scenarios'][s_o]['consumption']['pc_curt'][load_id] = pc_curt
                processed_results['scenarios'][s_o]['consumption']['pc_net'][load_id] -= pc_curt
                processed_results['scenarios'][s_o]['consumption']['qc_curt'][load_id] = qc_curt
                processed_results['scenarios'][s_o]['consumption']['qc_net'][load_id] -= qc_curt

        # Generation
        for g in model.generators:
            generator = network.generators[g]
            gen_id = generator.gen_id
            if params.rg_curt:
                processed_results['scenarios'][s_o]['generation']['pg_net'][gen_id] = 0.00
                processed_results['scenarios'][s_o]['generation']['qg_net'][gen_id] = 0.00
                processed_results['scenarios'][s_o]['generation']['sg_net'][gen_id] = 0.00
                processed_results['scenarios'][s_o]['generation']['sg_curt'][gen_id] = 0.00
            if generator.is_curtaillable() and params.rg_curt:
                pg = generator.pg[s_o][n] * network.baseMVA
                qg = generator.qg[s_o][n] * network.baseMVA
                sg = sqrt(pg ** 2 + qg ** 2)
                pg_net = pe.value(model.pg[g, s_o]) * network.baseMVA
                qg_net = pe.value(model.qg[g, s_o]) * network.baseMVA
                sg_net = pe.value(model.sg_abs[g, s_o]) * network.baseMVA
                sg_curt = pe.value(model.sg_curt[g, s_o]) * network.baseMVA
                processed_results['scenarios'][s_o]['generation']['pg_net'][gen_id] = pg_net
                processed_results['scenarios'][s_o]['generation']['qg_net'][gen_id] = qg_net
                processed_results['scenarios'][s_o]['generation']['sg_net'][gen_id] = sg_net
                processed_results['scenarios'][s_o]['generation']['sg_curt'][gen_id] = sg_curt
            else:
                pg = pe.value(model.pg[g, s_o]) * network.baseMVA
                qg = pe.value(model.qg[g, s_o]) * network.baseMVA
                sg = sqrt(pg ** 2 + qg ** 2)
            processed_results['scenarios'][s_o]['generation']['pg'][gen_id] = pg
            processed_results['scenarios'][s_o]['generation']['qg'][gen_id] = qg
            processed_results['scenarios'][s_o]['generation']['sg'][gen_id] = sg

        # Branch current, transformers' ratio
        for k in model.branches:

            branch = network.branches[k]
            branch_id = branch.branch_id
            rating = branch.rate / network.baseMVA
            if rating == 0.0:
                rating = BRANCH_UNKNOWN_RATING

            # Power flows
            pij, qij = _get_branch_power_flow(network, params, branch, branch.fbus, branch.tbus, model, s_o)
            pji, qji = _get_branch_power_flow(network, params, branch, branch.tbus, branch.fbus, model, s_o)
            sij_sqr = pij**2 + qij**2
            sji_sqr = pji**2 + qji**2
            processed_results['scenarios'][s_o]['branches']['power_flow']['pij'][branch_id] = pij
            processed_results['scenarios'][s_o]['branches']['power_flow']['pji'][branch_id] = pji
            processed_results['scenarios'][s_o]['branches']['power_flow']['qij'][branch_id] = qij
            processed_results['scenarios'][s_o]['branches']['power_flow']['qji'][branch_id] = qji
            processed_results['scenarios'][s_o]['branches']['power_flow']['sij'][branch_id] = sqrt(sij_sqr)
            processed_results['scenarios'][s_o]['branches']['power_flow']['sji'][branch_id] = sqrt(sji_sqr)

            # Losses (active power)
            p_losses = _get_branch_power_losses(network, params, model, k, s_o)
            processed_results['scenarios'][s_o]['branches']['losses'][branch_id] = p_losses

            # Ratio
            if branch.is_transformer:
                r_ij = pe.value(model.r[k, s_o])
                processed_results['scenarios'][s_o]['branches']['ratio'][branch_id] = r_ij

            # Branch flow (limits)
            flow_ij_perc = sqrt(abs(pe.value(model.flow_ij_sqr[k, s_o]))) / rating
            processed_results['scenarios'][s_o]['branches']['branch_flow']['flow_ij_perc'][branch_id] = flow_ij_perc

        # Energy Storage devices
        if params.es_reg:
            for e in model.energy_storages:
                es_id = network.energy_storages[e].es_id
                capacity = network.energy_storages[e].e * network.baseMVA
                if isclose(capacity, 0.0, abs_tol=1e-6):
                    capacity = 1.00
                s_ess = pe.value(model.es_sch[e, s_o] - model.es_sdch[e, s_o]) * network.baseMVA
                p_ess = pe.value(model.es_pch[e, s_o] - model.es_pdch[e, s_o]) * network.baseMVA
                q_ess = pe.value(model.es_qch[e, s_o] - model.es_qdch[e, s_o]) * network.baseMVA
                soc_ess = pe.value(model.es_soc[e, s_o]) * network.baseMVA
                processed_results['scenarios'][s_o]['energy_storages']['p'][es_id] = p_ess
                processed_results['scenarios'][s_o]['energy_storages']['q'][es_id] = q_ess
                processed_results['scenarios'][s_o]['energy_storages']['s'][es_id] = s_ess
                processed_results['scenarios'][s_o]['energy_storages']['soc'][es_id] = soc_ess
                processed_results['scenarios'][s_o]['energy_storages']['soc_percent'][es_id] = soc_ess / capacity

        # Flexible loads
        if params.fl_reg:
            for i in model.loads:
                load_id = network.loads[i].load_id
                p_up = pe.value(model.flex_p_up[i, s_o]) * network.baseMVA
                p_down = pe.value(model.flex_p_down[i, s_o]) * network.baseMVA
                processed_results['scenarios'][s_o]['consumption']['p_up'][load_id] = p_up
                processed_results['scenarios'][s_o]['consumption']['p_down'][load_id] = p_down

        # Shared Energy Storages
        for e in model.shared_energy_storages:
            node_id = network.shared_energy_storages[e].bus
            capacity = network.shared_energy_storages[e].e * network.baseMVA
            if isclose(capacity, 0.0, abs_tol=1e-6):
                capacity = 1.00
            s_ess = pe.value(model.shared_es_sch[e, s_o] - model.shared_es_sdch[e, s_o]) * network.baseMVA
            p_ess = pe.value(model.shared_es_pch[e, s_o] - model.shared_es_pdch[e, s_o]) * network.baseMVA
            q_ess = pe.value(model.shared_es_qch[e, s_o] - model.shared_es_qdch[e, s_o]) * network.baseMVA
            soc_ess = pe.value(model.shared_es_soc[e, s_o]) * network.baseMVA
            processed_results['scenarios'][s_o]['shared_energy_storages']['p'][node_id] = p_ess
            processed_results['scenarios'][s_o]['shared_energy_storages']['q'][node_id] = q_ess
            processed_results['scenarios'][s_o]['shared_energy_storages']['s'][node_id] = s_ess
            processed_results['scenarios'][s_o]['shared_energy_storages']['soc'][node_id] = soc_ess
            processed_results['scenarios'][s_o]['shared_energy_storages']['soc_percent'][node_id] = soc_ess / capacity

        # Voltage slacks
        if params.slacks.grid_operation.voltage:
            for i in model.nodes:
                node_id = network.nodes[i].bus_i
                slack_e = pe.value(model.slack_e[i, s_o])
                slack_f = pe.value(model.slack_f[i, s_o])
                processed_results['scenarios'][s_o]['relaxation_slacks']['voltage']['e'][node_id] = slack_e
                processed_results['scenarios'][s_o]['relaxation_slacks']['voltage']['f'][node_id] = slack_f

        # Branch current slacks
        if params.slacks.grid_operation.branch_flow:
            for b in model.branches:
                branch_id = network.branches[b].branch_id
                slack_flow_ij_sqr = pe.value(model.slack_flow_ij_sqr[b, s_o])
                processed_results['scenarios'][s_o]['relaxation_slacks']['branch_flow']['flow_ij_sqr'][branch_id] = slack_flow_ij_sqr

        # Slacks
        # - Shared ESS
        for e in model.shared_energy_storages:
            node_id = network.shared_energy_storages[e].bus
            if params.slacks.shared_ess.complementarity:
                slack_comp = pe.value(model.slack_shared_es_comp[e, s_o]) * (s_base ** 2)
                processed_results['scenarios'][s_o]['relaxation_slacks']['shared_energy_storages']['comp'][node_id] = slack_comp

        # - Node balance
        if params.slacks.node_balance:
            for i in model.nodes:
                node_id = network.nodes[i].bus_i
                slack_p = pe.value(model.slack_node_balance_p[i, s_o]) * s_base
                slack_q = pe.value(model.slack_node_balance_q[i, s_o]) * s_base
                processed_results['scenarios'][s_o]['relaxation_slacks']['node_balance']['p'][node_id] = slack_p
                processed_results['scenarios'][s_o]['relaxation_slacks']['node_balance']['q'][node_id] = slack_q

        # - ESS slacks
        if params.es_reg:
            for e in model.energy_storages:
                es_id = network.energy_storages[e].es_id
                if params.slacks.ess.complementarity:
                    slack_comp = pe.value(model.slack_es_comp[e, s_o]) * (s_base ** 2)
                    processed_results['scenarios'][s_o]['relaxation_slacks']['energy_storages']['comp'][es_id] = slack_comp

    return processed_results


def _compute_objective_function_value(network, model, params, n=0):

    obj = 0.0

    if params.obj_type == OBJ_MIN_COST:

        c_p = network.cost_energy_p
        c_flex = network.cost_flex
        cost_res_curt = pe.value(model.cost_res_curtailment)
        cost_load_curt = pe.value(model.cost_load_curtailment)

        for s_o in model.scenarios_operation:

            obj_scenario = 0.0

            # Generation -- paid at market price
            for g in model.generators:
                if network.generators[g].is_controllable():
                    if (not network.is_transmission) and network.generators[g].gen_type == GEN_REFERENCE:
                        continue
                    pg = pe.value(model.pg[g, s_o])
                    obj_scenario += c_p[n] * network.baseMVA * pg

            # Demand side flexibility
            if params.fl_reg:
                for c in model.loads:
                    flex_up = pe.value(model.flex_p_up[c, s_o])
                    flex_down = pe.value(model.flex_p_down[c, s_o])
                    obj_scenario += c_flex[n] * network.baseMVA * (flex_down + flex_up)

            # Load curtailment
            if params.l_curt:
                for c in model.loads:
                    pc_curt = pe.value(model.pc_curt_down[c, s_o] + model.pc_curt_up[c, s_o])
                    qc_curt = pe.value(model.qc_curt_down[c, s_o] + model.qc_curt_up[c, s_o])
                    obj_scenario += cost_load_curt * network.baseMVA * (pc_curt + qc_curt)

            # Generation curtailment
            if params.rg_curt:
                for g in model.generators:
                    sg_curt = pe.value(model.sg_curt[g, s_o])
                    obj_scenario += cost_res_curt * network.baseMVA * sg_curt

            obj += obj_scenario * network.prob_operation_scenarios[s_o]

    elif params.obj_type == OBJ_CONGESTION_MANAGEMENT:

        pen_gen_curtailment = pe.value(model.penalty_gen_curtailment)
        pen_load_curtailment = pe.value(model.penalty_load_curtailment)
        pen_flex_usage = pe.value(model.penalty_flex_usage)

        for s_o in model.scenarios_operation:

            obj_scenario = 0.0

            # Generation curtailment
            if params.rg_curt:
                for g in model.generators:
                    sg_curt = pe.value(model.sg_curt[g, s_o])
                    obj_scenario += pen_gen_curtailment * network.baseMVA * sg_curt

            # Consumption curtailment
            if params.l_curt:
                for c in model.loads:
                    pc_curt = pe.value(model.pc_curt_down[c, s_o] + model.pc_curt_up[c, s_o])
                    qc_curt = pe.value(model.qc_curt_down[c, s_o] + model.qc_curt_up[c, s_o])
                    obj_scenario += pen_load_curtailment * network.baseMVA * (pc_curt + qc_curt)

            # Demand side flexibility
            if params.fl_reg:
                for c in model.loads:
                    flex_p_up = pe.value(model.flex_p_up[c, s_o])
                    flex_p_down = pe.value(model.flex_p_down[c, s_o])
                    obj_scenario += pen_flex_usage * network.baseMVA * (flex_p_down + flex_p_up)

            obj += obj_scenario * network.prob_operation_scenarios[s_o]

    return obj


def _compute_generation_cost(network, model, n=0):

    gen_cost = 0.0

    c_p = network.cost_energy_p

    for s_o in model.scenarios_operation:
        gen_cost_scenario = 0.0
        for g in model.generators:
            if network.generators[g].is_controllable():
                gen_cost_scenario += c_p[n] * network.baseMVA * pe.value(model.pg[g, s_o])
        gen_cost += gen_cost_scenario * network.prob_operation_scenarios[s_o]

    return gen_cost


def _compute_total_load(network, model, params):

    total_load = {'p': 0.00, 'q': 0.00}

    for s_o in model.scenarios_operation:
        total_load_scenario = {'p': 0.00, 'q': 0.00}
        for c in model.loads:
            total_load_scenario['p'] += network.baseMVA * pe.value(model.pc[c, s_o])
            total_load_scenario['q'] += network.baseMVA * pe.value(model.qc[c, s_o])
            if params.l_curt:
                total_load_scenario['p'] -= network.baseMVA * pe.value(model.pc_curt_down[c, s_o] - model.pc_curt_up[c, s_o])
                total_load_scenario['q'] -= network.baseMVA * pe.value(model.qc_curt_down[c, s_o] - model.qc_curt_up[c, s_o])

        total_load['p'] += total_load_scenario['p'] * network.prob_operation_scenarios[s_o]
        total_load['q'] += total_load_scenario['q'] * network.prob_operation_scenarios[s_o]

    return total_load


def _compute_total_generation(network, model, params):

    total_gen = {'p': 0.00, 'q': 0.00}

    for s_o in model.scenarios_operation:
        total_gen_scenario = {'p': 0.00, 'q': 0.00}
        for g in model.generators:
            total_gen_scenario['p'] += network.baseMVA * pe.value(model.pg[g, s_o])
            total_gen_scenario['q'] += network.baseMVA * pe.value(model.qg[g, s_o])
        total_gen['p'] += total_gen_scenario['p'] * network.prob_operation_scenarios[s_o]
        total_gen['q'] += total_gen_scenario['q'] * network.prob_operation_scenarios[s_o]

    return total_gen


def _compute_conventional_generation(network, model):

    total_gen = {'p': 0.00, 'q': 0.00}

    for s_o in model.scenarios_operation:
        total_gen_scenario = {'p': 0.00, 'q': 0.00}
        for g in model.generators:
            if network.generators[g].gen_type == GEN_CONV:
                total_gen_scenario['p'] += network.baseMVA * pe.value(model.pg[g, s_o])
                total_gen_scenario['q'] += network.baseMVA * pe.value(model.qg[g, s_o])
        total_gen['p'] += total_gen_scenario['p'] * network.prob_operation_scenarios[s_o]
        total_gen['q'] += total_gen_scenario['q'] * network.prob_operation_scenarios[s_o]

    return total_gen


def _compute_renewable_generation(network, model, params):

    total_renewable_gen = {'p': 0.00, 'q': 0.00, 's': 0.00}

    for s_o in model.scenarios_operation:
        total_renewable_gen_scenario = {'p': 0.00, 'q': 0.00, 's': 0.00}
        for g in model.generators:
            if network.generators[g].is_renewable():
                total_renewable_gen_scenario['p'] += network.baseMVA * pe.value(model.pg[g, s_o])
                total_renewable_gen_scenario['q'] += network.baseMVA * pe.value(model.qg[g, s_o])
                if params.rg_curt:
                    total_renewable_gen_scenario['s'] += network.baseMVA * pe.value(model.sg_abs[g, s_o])
        total_renewable_gen['p'] += total_renewable_gen_scenario['p'] * network.prob_operation_scenarios[s_o]
        total_renewable_gen['q'] += total_renewable_gen_scenario['q'] * network.prob_operation_scenarios[s_o]
        total_renewable_gen['s'] += total_renewable_gen_scenario['s'] * network.prob_operation_scenarios[s_o]

    return total_renewable_gen


def _compute_losses(network, model, params):
    power_losses = 0.0
    for s_o in model.scenarios_operation:
        power_losses_scenario = 0.0
        for k in model.branches:
            power_losses_scenario += _get_branch_power_losses(network, params, model, k, s_o)
        power_losses += power_losses_scenario * network.prob_operation_scenarios[s_o]
    return power_losses


def _compute_generation_curtailment(network, model, params):

    gen_curtailment = {'p': 0.00, 'q': 0.00, 's': 0.00}

    if params.rg_curt:
        for s_o in model.scenarios_operation:
            gen_curtailment_scenario = {'p': 0.00, 'q': 0.00, 's': 0.00}
            for g in model.generators:
                if network.generators[g].is_curtaillable():
                    gen_curtailment_scenario['s'] += pe.value(model.sg_curt[g, s_o]) * network.baseMVA
            gen_curtailment['p'] += gen_curtailment_scenario['p'] * network.prob_operation_scenarios[s_o]
            gen_curtailment['q'] += gen_curtailment_scenario['q'] * network.prob_operation_scenarios[s_o]
            gen_curtailment['s'] += gen_curtailment_scenario['s'] * network.prob_operation_scenarios[s_o]

    return gen_curtailment


def _compute_load_curtailment(network, model, params):

    load_curtailment = {'p': 0.00, 'q': 0.00}

    if params.l_curt:
        for s_m in model.scenarios_market:
            for s_o in model.scenarios_operation:
                load_curtailment_scenario = {'p': 0.00, 'q': 0.00}
                for c in model.loads:
                    for p in model.periods:
                        load_curtailment_scenario['p'] += pe.value(model.pc_curt_down[c, s_m, s_o, p] - model.pc_curt_up[c, s_m, s_o, p]) * network.baseMVA
                        load_curtailment_scenario['q'] += pe.value(model.qc_curt_down[c, s_m, s_o, p] - model.qc_curt_up[c, s_m, s_o, p]) * network.baseMVA

                load_curtailment['p'] += load_curtailment_scenario['p'] * (network.prob_market_scenarios[s_m] * network.prob_operation_scenarios[s_o])
                load_curtailment['q'] += load_curtailment_scenario['q'] * (network.prob_market_scenarios[s_m] * network.prob_operation_scenarios[s_o])

    return load_curtailment


def _compute_flexibility_used(network, model, params):

    flexibility_used = 0.0

    if params.fl_reg:
        for s_o in model.scenarios_operation:
            flexibility_used_scenario = 0.0
            for c in model.loads:
                flexibility_used_scenario += pe.value(model.flex_p_up[c, s_o]) * network.baseMVA
                flexibility_used_scenario += pe.value(model.flex_p_down[c, s_o]) * network.baseMVA

            flexibility_used += flexibility_used_scenario * network.prob_operation_scenarios[s_o]

    return flexibility_used


def _write_optimization_results_to_excel(network, data_dir, processed_results, filename):

    wb = Workbook()

    _write_main_info_to_excel(network, wb, processed_results)
    _write_shared_network_energy_storage_results_to_excel(network, wb, processed_results)
    _write_network_voltage_results_to_excel(network, wb, processed_results)
    _write_network_consumption_results_to_excel(network, wb, processed_results)
    _write_network_generation_results_to_excel(network, wb, processed_results)
    _write_network_branch_results_to_excel(network, wb, processed_results, 'losses')
    _write_network_branch_results_to_excel(network, wb, processed_results, 'ratio')
    _write_network_branch_loading_results_to_excel(network, wb, processed_results)
    _write_network_branch_power_flow_results_to_excel(network, wb, processed_results)
    if network.params.es_reg:
        _write_network_energy_storage_results_to_excel(network, wb, processed_results)
    _write_relaxation_slacks_scenarios_results_to_excel(network, wb, processed_results)

    results_filename = os.path.join(data_dir, f'{filename}_results.xlsx')
    try:
        wb.save(results_filename)
        print('[INFO] S-MPOPF Results written to {}.'.format(results_filename))
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(data_dir, f'{network.name}_results_{current_time}.xlsx')
        print('[INFO] S-MPOPF Results written to {}.'.format(backup_filename))
        wb.save(backup_filename)


def _write_main_info_to_excel(network, workbook, results):

    sheet = workbook.worksheets[0]
    sheet.title = 'Main Info'

    decimal_style = '0.00'
    line_idx = 1
    sheet.cell(row=line_idx, column=2).value = 'Value'

    # Objective function value
    line_idx += 1
    obj_string = 'Objective'
    if network.params.obj_type == OBJ_MIN_COST:
        obj_string += ' (cost), [€]'
    elif network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
        obj_string += ' (congestion management)'
    sheet.cell(row=line_idx, column=1).value = obj_string
    sheet.cell(row=line_idx, column=2).value = results['obj']
    sheet.cell(row=line_idx, column=2).number_format = decimal_style

    # Total Load
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Load, [MWh]'
    sheet.cell(row=line_idx, column=2).value = results['total_load']['p']
    sheet.cell(row=line_idx, column=2).number_format = decimal_style

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Load, [MVArh]'
    sheet.cell(row=line_idx, column=2).value = results['total_load']['q']
    sheet.cell(row=line_idx, column=2).number_format = decimal_style

    # Flexibility used
    if network.params.fl_reg:
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Flexibility used, [MWh]'
        sheet.cell(row=line_idx, column=2).value = results['flex_used']
        sheet.cell(row=line_idx, column=2).number_format = decimal_style

    # Total Load curtailed
    if network.params.l_curt:

        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Load curtailed, [MWh]'
        sheet.cell(row=line_idx, column=2).value = results['load_curt']['p']
        sheet.cell(row=line_idx, column=2).number_format = decimal_style

        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Load curtailed, [MVArh]'
        sheet.cell(row=line_idx, column=2).value = results['load_curt']['q']
        sheet.cell(row=line_idx, column=2).number_format = decimal_style

    # Total Generation
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Generation, [MWh]'
    sheet.cell(row=line_idx, column=2).value = results['total_gen']['p']
    sheet.cell(row=line_idx, column=2).number_format = decimal_style

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Generation, [MVArh]'
    sheet.cell(row=line_idx, column=2).value = results['total_gen']['q']
    sheet.cell(row=line_idx, column=2).number_format = decimal_style

    # Total Renewable Generation
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Renewable generation, [MWh]'
    sheet.cell(row=line_idx, column=2).value = results['total_renewable_gen']['p']
    sheet.cell(row=line_idx, column=2).number_format = decimal_style

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Renewable generation, [MVArh]'
    sheet.cell(row=line_idx, column=2).value = results['total_renewable_gen']['q']
    sheet.cell(row=line_idx, column=2).number_format = decimal_style

    # Renewable Generation Curtailed
    if network.params.rg_curt:

        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Renewable generation curtailed, [MWh]'
        sheet.cell(row=line_idx, column=2).value = results['gen_curt']['p']
        sheet.cell(row=line_idx, column=2).number_format = decimal_style

        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Renewable generation curtailed, [MVArh]'
        sheet.cell(row=line_idx, column=2).value = results['gen_curt']['q']
        sheet.cell(row=line_idx, column=2).number_format = decimal_style

    # Losses
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Losses, [MWh]'
    sheet.cell(row=line_idx, column=2).value = results['losses']
    sheet.cell(row=line_idx, column=2).number_format = decimal_style

    # Execution time
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Execution time, [s]'
    sheet.cell(row=line_idx, column=2).value = results['runtime'][0]
    sheet.cell(row=line_idx, column=2).number_format = decimal_style

    # Number of operation (generation and consumption) scenarios
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Number of operation scenarios'
    sheet.cell(row=line_idx, column=2).value = len(network.prob_operation_scenarios)


def _write_shared_network_energy_storage_results_to_excel(network, workbook, results, n=0):

    sheet = workbook.create_sheet('Shared Energy Storage')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Quantity'
    sheet.cell(row=row_idx, column=3).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=4).value = n
    row_idx = row_idx + 1

    expected_p = dict()
    expected_q = dict()
    expected_s = dict()
    expected_soc = dict()
    expected_soc_perc = dict()

    for energy_storage in network.shared_energy_storages:
        expected_p[energy_storage.bus] = 0.00
        expected_q[energy_storage.bus] = 0.00
        expected_s[energy_storage.bus] = 0.00
        expected_soc[energy_storage.bus] = 0.00
        expected_soc_perc[energy_storage.bus] = 0.0

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for node_id in results['scenarios'][s_o]['shared_energy_storages']['p']:

            pc = results['scenarios'][s_o]['shared_energy_storages']['p'][node_id]
            qc = results['scenarios'][s_o]['shared_energy_storages']['q'][node_id]
            sc = results['scenarios'][s_o]['shared_energy_storages']['s'][node_id]
            soc = results['scenarios'][s_o]['shared_energy_storages']['soc'][node_id]
            soc_perc = results['scenarios'][s_o]['shared_energy_storages']['soc_percent'][node_id]

            # - Active Power
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=3).value = s_o
            sheet.cell(row=row_idx, column=4).value = pc
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            if pc != 'N/A':
                expected_p[node_id] += pc * omega_s
            else:
                expected_p[node_id] = 'N/A'
            row_idx = row_idx + 1

            # - Reactive Power
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=3).value = s_o
            sheet.cell(row=row_idx, column=4).value = qc
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            if qc != 'N/A':
                expected_q[node_id] += qc * omega_s
            else:
                expected_q[node_id] = 'N/A'
            row_idx = row_idx + 1

            # - Apparent Power
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'S, [MVA]'
            sheet.cell(row=row_idx, column=3).value = s_o
            sheet.cell(row=row_idx, column=4).value = sc
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            if sc != 'N/A':
                expected_s[node_id] += sc * omega_s
            else:
                expected_s[node_id] = 'N/A'
            row_idx = row_idx + 1

            # - SoC, [MWh]
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'SoC, [MWh]'
            sheet.cell(row=row_idx, column=3).value = s_o
            sheet.cell(row=row_idx, column=4).value = soc
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            if soc != 'N/A':
                expected_soc[node_id] += soc * omega_s
            else:
                expected_soc[node_id] = 'N/A'
            row_idx = row_idx + 1

            # - SoC, [%]
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'SoC, [%]'
            sheet.cell(row=row_idx, column=3).value = s_o
            sheet.cell(row=row_idx, column=4).value = soc_perc
            sheet.cell(row=row_idx, column=4).number_format = perc_style
            if soc_perc != 'N/A':
                expected_soc_perc[node_id] += soc_perc * omega_s
            else:
                expected_soc_perc[node_id] = 'N/A'
            row_idx = row_idx + 1

    for energy_storage in network.shared_energy_storages:

        node_id = energy_storage.bus

        # - Active Power
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=3).value = 'Expected'
        sheet.cell(row=row_idx, column=4).value = expected_p[node_id]
        sheet.cell(row=row_idx, column=4).number_format = decimal_style
        row_idx = row_idx + 1

        # - Reactive Power
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=3).value = 'Expected'
        sheet.cell(row=row_idx, column=4).value = expected_q[node_id]
        sheet.cell(row=row_idx, column=4).number_format = decimal_style
        row_idx = row_idx + 1

        # - Apparent Power
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'S, [MVA]'
        sheet.cell(row=row_idx, column=3).value = 'Expected'
        sheet.cell(row=row_idx, column=4).value = expected_s[node_id]
        sheet.cell(row=row_idx, column=4).number_format = decimal_style
        row_idx = row_idx + 1

        # - SoC, [MWh]
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'SoC, [MWh]'
        sheet.cell(row=row_idx, column=3).value = 'Expected'
        sheet.cell(row=row_idx, column=4).value = expected_soc[node_id]
        sheet.cell(row=row_idx, column=4).number_format = decimal_style
        row_idx = row_idx + 1

        # - SoC, [%]
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'SoC, [%]'
        sheet.cell(row=row_idx, column=3).value = 'Expected'
        sheet.cell(row=row_idx, column=4).value = expected_soc_perc[node_id]
        sheet.cell(row=row_idx, column=4).number_format = perc_style
        row_idx = row_idx + 1


def _write_network_voltage_results_to_excel(network, workbook, results, n=0):

    sheet = workbook.create_sheet('Voltage')

    row_idx = 1
    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Quantity'
    sheet.cell(row=row_idx, column=3).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=4).value = n
    row_idx = row_idx + 1

    expected_vmag = dict()
    expected_vang = dict()
    for node in network.nodes:
        expected_vmag[node.bus_i] = 0.00
        expected_vang[node.bus_i] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for node_id in results['scenarios'][s_o]['voltage']['vmag']:

            v_min, v_max = network.get_node_voltage_limits(node_id)
            v_mag = results['scenarios'][s_o]['voltage']['vmag'][node_id]
            v_ang = results['scenarios'][s_o]['voltage']['vang'][node_id]

            # Voltage magnitude
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'Vmag, [p.u.]'
            sheet.cell(row=row_idx, column=3).value = s_o
            sheet.cell(row=row_idx, column=4).value = v_mag
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            if v_mag > v_max + VIOLATION_TOLERANCE or v_mag < v_min - VIOLATION_TOLERANCE:
                sheet.cell(row=row_idx, column=4).fill = violation_fill
            expected_vmag[node_id] += v_mag * omega_s
            row_idx = row_idx + 1

            # Voltage angle
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = 'Vang, [º]'
            sheet.cell(row=row_idx, column=3).value = s_o
            sheet.cell(row=row_idx, column=4).value = v_ang
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            expected_vang[node_id] += v_ang * omega_s
            row_idx = row_idx + 1

    for node in network.nodes:

        node_id = node.bus_i
        v_min, v_max = network.get_node_voltage_limits(node_id)

        # Expected voltage magnitude
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'Vmag, [p.u.]'
        sheet.cell(row=row_idx, column=3).value = '-'
        sheet.cell(row=row_idx, column=4).value = expected_vmag[node_id]
        sheet.cell(row=row_idx, column=4).number_format = decimal_style
        if expected_vmag[node_id] > v_max + VIOLATION_TOLERANCE or expected_vmag[node_id] < v_min - VIOLATION_TOLERANCE:
            sheet.cell(row=row_idx, column=4).fill = violation_fill
        row_idx = row_idx + 1

        # Expected voltage angle
        sheet.cell(row=row_idx, column=1).value = node_id
        sheet.cell(row=row_idx, column=2).value = 'Vang, [º]'
        sheet.cell(row=row_idx, column=3).value = 'Expected'
        sheet.cell(row=row_idx, column=4).value = expected_vang[node_id]
        sheet.cell(row=row_idx, column=4).number_format = decimal_style
        row_idx = row_idx + 1


def _write_network_consumption_results_to_excel(network, workbook, results, n=0):

    sheet = workbook.create_sheet('Consumption')

    row_idx = 1
    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Load ID'
    sheet.cell(row=row_idx, column=2).value = 'Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Quantity'
    sheet.cell(row=row_idx, column=4).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=5).value = n
    row_idx = row_idx + 1

    expected_pc = dict()
    expected_flex_up = dict()
    expected_flex_down = dict()
    expected_pc_curt = dict()
    expected_pnet = dict()
    expected_qc = dict()
    expected_qc_curt = dict()
    expected_qnet = dict()
    for load in network.loads:
        expected_pc[load.load_id] = 0.00
        expected_flex_up[load.load_id] = 0.00
        expected_flex_down[load.load_id] = 0.00
        expected_pc_curt[load.load_id] = 0.00
        expected_pnet[load.load_id] = 0.00
        expected_qc[load.load_id] = 0.00
        expected_qc_curt[load.load_id] = 0.00
        expected_qnet[load.load_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for load in network.loads:

            load_id = load.load_id
            node_id = load.bus

            pc = results['scenarios'][s_o]['consumption']['pc'][load_id]
            qc = results['scenarios'][s_o]['consumption']['qc'][load_id]

            # - Active Power
            sheet.cell(row=row_idx, column=1).value = load_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'Pc, [MW]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = pc
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            expected_pc[load_id] += pc * omega_s
            row_idx = row_idx + 1

            if network.params.fl_reg:

                flex_up = results['scenarios'][s_o]['consumption']['p_up'][load_id]
                flex_down = results['scenarios'][s_o]['consumption']['p_down'][load_id]

                # - Flexibility, up
                sheet.cell(row=row_idx, column=1).value = load_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = 'Flex Up, [MW]'
                sheet.cell(row=row_idx, column=4).value = s_o
                sheet.cell(row=row_idx, column=5).value = flex_up
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                expected_flex_up[load_id] += flex_up * omega_s
                row_idx = row_idx + 1

                # - Flexibility, down
                sheet.cell(row=row_idx, column=1).value = load_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = 'Flex Down, [MW]'
                sheet.cell(row=row_idx, column=4).value = s_o
                sheet.cell(row=row_idx, column=5).value = flex_down
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                expected_flex_down[load_id] += flex_down * omega_s
                row_idx = row_idx + 1

            if network.params.l_curt:

                pc_curt = results['scenarios'][s_o]['consumption']['pc_curt'][load_id]

                # - Active power curtailment
                sheet.cell(row=row_idx, column=1).value = load_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = 'Pc_curt, [MW]'
                sheet.cell(row=row_idx, column=4).value = s_o
                sheet.cell(row=row_idx, column=5).value = pc_curt
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                if not isclose(pc_curt, 0.00, abs_tol= VIOLATION_TOLERANCE):
                    sheet.cell(row=row_idx, column=5).fill = violation_fill
                expected_pc_curt[load_id] += pc_curt * omega_s
                row_idx = row_idx + 1

            if network.params.fl_reg or network.params.l_curt:

                p_net = results['scenarios'][s_o]['consumption']['pc_net'][load_id]

                # - Active power net consumption
                sheet.cell(row=row_idx, column=1).value = load_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = 'Pc_net, [MW]'
                sheet.cell(row=row_idx, column=4).value = s_o
                sheet.cell(row=row_idx, column=5).value = p_net
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                expected_pnet[load_id] += p_net * omega_s
                row_idx = row_idx + 1

            # - Reactive power
            sheet.cell(row=row_idx, column=1).value = load_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'Qc, [MVAr]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = qc
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            expected_qc[load_id] += qc * omega_s
            row_idx = row_idx + 1

            if network.params.l_curt:

                qc_curt = results['scenarios'][s_o]['consumption']['qc_curt'][load_id]
                q_net = results['scenarios'][s_o]['consumption']['qc_net'][load_id]

                # - Reactive power curtailment
                sheet.cell(row=row_idx, column=1).value = load_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = 'Qc_curt, [MW]'
                sheet.cell(row=row_idx, column=4).value = s_o
                sheet.cell(row=row_idx, column=5).value = qc_curt
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                if not isclose(qc_curt, 0.00, abs_tol=VIOLATION_TOLERANCE):
                    sheet.cell(row=row_idx, column=5).fill = violation_fill
                expected_qc_curt[load_id] += qc_curt * omega_s
                row_idx = row_idx + 1

                # - Reactive power net consumption
                sheet.cell(row=row_idx, column=1).value = load_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = 'Qc_net, [MW]'
                sheet.cell(row=row_idx, column=4).value = s_o
                sheet.cell(row=row_idx, column=5).value = q_net
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                expected_qnet[load_id] += q_net * omega_s
                row_idx = row_idx + 1

    for load in network.loads:

        load_id = load.load_id
        node_id = load.bus

        # - Active Power
        sheet.cell(row=row_idx, column=1).value = load_id
        sheet.cell(row=row_idx, column=2).value = node_id
        sheet.cell(row=row_idx, column=3).value = 'Pc, [MW]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_pc[load_id]
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx = row_idx + 1

        if network.params.fl_reg:

            # - Flexibility, up
            sheet.cell(row=row_idx, column=1).value = load_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'Flex Up, [MW]'
            sheet.cell(row=row_idx, column=4).value = 'Expected'
            sheet.cell(row=row_idx, column=5).value = expected_flex_up[load_id]
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            row_idx = row_idx + 1

            # - Flexibility, down
            sheet.cell(row=row_idx, column=1).value = load_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'Flex Down, [MW]'
            sheet.cell(row=row_idx, column=4).value = 'Expected'
            sheet.cell(row=row_idx, column=5).value = expected_flex_down[load_id]
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            row_idx = row_idx + 1

        if network.params.l_curt:

            # - Load curtailment (active power)
            sheet.cell(row=row_idx, column=1).value = load_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'Pc_curt, [MW]'
            sheet.cell(row=row_idx, column=4).value = 'Expected'
            sheet.cell(row=row_idx, column=5).value = expected_pc_curt[load_id]
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if not isclose(expected_pc_curt[load_id], 0.00, abs_tol=VIOLATION_TOLERANCE):
                sheet.cell(row=row_idx, column=p + 8).fill = violation_fill
            row_idx = row_idx + 1

        if network.params.fl_reg or network.params.l_curt:

            # - Active power net consumption
            sheet.cell(row=row_idx, column=1).value = load_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'Pc_net, [MW]'
            sheet.cell(row=row_idx, column=4).value = 'Expected'
            sheet.cell(row=row_idx, column=5).value = expected_pnet[load_id]
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            row_idx = row_idx + 1

        # - Reactive power
        sheet.cell(row=row_idx, column=1).value = load_id
        sheet.cell(row=row_idx, column=2).value = node_id
        sheet.cell(row=row_idx, column=3).value = 'Qc, [MVAr]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        for p in range(network.num_instants):
            sheet.cell(row=row_idx, column=5).value = expected_qc[load_id]
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx = row_idx + 1

        if network.params.l_curt:

            # - Load curtailment (reactive power)
            sheet.cell(row=row_idx, column=1).value = load_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'Qc_curt, [MW]'
            sheet.cell(row=row_idx, column=4).value = 'Expected'
            sheet.cell(row=row_idx, column=5).value = expected_qc_curt[load_id]
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if not isclose(expected_qc_curt[load_id], 0.00, abs_tol=VIOLATION_TOLERANCE):
                sheet.cell(row=row_idx, column=5).fill = violation_fill
            row_idx = row_idx + 1

            # - Reactive power net consumption
            sheet.cell(row=row_idx, column=1).value = load_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'Qc_net, [MW]'
            sheet.cell(row=row_idx, column=4).value = 'Expected'
            sheet.cell(row=row_idx, column=5).value = expected_qnet[load_id]
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            row_idx = row_idx + 1


def _write_network_generation_results_to_excel(network, workbook, results, n=0):

    sheet = workbook.create_sheet('Generation')

    row_idx = 1
    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Generator ID'
    sheet.cell(row=row_idx, column=2).value = 'Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Type'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=6).value = n
    row_idx = row_idx + 1

    expected_pg = dict()
    expected_pg_net = dict()
    expected_qg = dict()
    expected_qg_net = dict()
    expected_sg = dict()
    expected_sg_curt = dict()
    expected_sg_net = dict()
    for generator in network.generators:
        expected_pg[generator.gen_id] = 0.00
        expected_pg_net[generator.gen_id] = 0.00
        expected_qg[generator.gen_id] = 0.00
        expected_qg_net[generator.gen_id] = 0.00
        expected_sg[generator.gen_id] = 0.00
        expected_sg_curt[generator.gen_id] = 0.00
        expected_sg_net[generator.gen_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for generator in network.generators:

            gen_id = generator.gen_id
            node_id = generator.bus
            gen_type = network.get_gen_type(gen_id)

            pg = results['scenarios'][s_o]['generation']['pg'][gen_id]
            qg = results['scenarios'][s_o]['generation']['qg'][gen_id]

            # Active Power
            sheet.cell(row=row_idx, column=1).value = gen_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = gen_type
            sheet.cell(row=row_idx, column=4).value = 'Pg, [MW]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = pg
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            expected_pg[gen_id] += pg * omega_s
            row_idx = row_idx + 1

            # Active Power net
            if generator.is_curtaillable() and network.params.rg_curt:

                pg_net = results['scenarios'][s_o]['generation']['pg_net'][gen_id]

                sheet.cell(row=row_idx, column=1).value = gen_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = 'Pg_net, [MW]'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = pg_net
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                expected_pg_net[gen_id] += pg_net * omega_s
                row_idx = row_idx + 1

            # Reactive Power
            sheet.cell(row=row_idx, column=1).value = gen_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = gen_type
            sheet.cell(row=row_idx, column=4).value = 'Qg, [MVAr]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = qg
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            expected_qg[gen_id] += qg * omega_s
            row_idx = row_idx + 1

            # Reactive Power net
            if generator.is_curtaillable() and network.params.rg_curt:

                qg_net = results['scenarios'][s_o]['generation']['qg_net'][gen_id]

                sheet.cell(row=row_idx, column=1).value = gen_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = 'Qg_net, [MVAr]'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = qg_net
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                expected_qg_net[gen_id] += qg_net * omega_s
                row_idx = row_idx + 1

            # Apparent Power
            if generator.is_curtaillable() and network.params.rg_curt:

                sg = results['scenarios'][s_o]['generation']['sg'][gen_id]
                sg_curt = results['scenarios'][s_o]['generation']['sg_curt'][gen_id]
                sg_net = results['scenarios'][s_o]['generation']['sg_net'][gen_id]

                sheet.cell(row=row_idx, column=1).value = gen_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = 'Sg, [MVA]'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = sg
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                expected_sg[gen_id] += sg * omega_s
                row_idx = row_idx + 1

                # Apparent Power net
                sheet.cell(row=row_idx, column=1).value = gen_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = 'Qg_net, [MVAr]'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = sg_curt
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                expected_sg_curt[gen_id] += sg_curt * omega_s
                row_idx = row_idx + 1

                # Apparent Power net
                sheet.cell(row=row_idx, column=1).value = gen_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = 'Qg_net, [MVAr]'
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = sg_net
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                expected_sg_net[gen_id] += sg_net * omega_s
                row_idx = row_idx + 1

    for generator in network.generators:

        node_id = generator.bus
        gen_id = generator.gen_id
        gen_type = network.get_gen_type(gen_id)

        # Active Power
        sheet.cell(row=row_idx, column=1).value = gen_id
        sheet.cell(row=row_idx, column=2).value = node_id
        sheet.cell(row=row_idx, column=3).value = gen_type
        sheet.cell(row=row_idx, column=4).value = 'Pg, [MW]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_pg[gen_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        row_idx = row_idx + 1

        # Active Power net
        if generator.is_curtaillable() and network.params.rg_curt:
            sheet.cell(row=row_idx, column=1).value = gen_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = gen_type
            sheet.cell(row=row_idx, column=4).value = 'Pg_net, [MW]'
            sheet.cell(row=row_idx, column=5).value = 'Expected'
            sheet.cell(row=row_idx, column=6).value = expected_pg_net[gen_id]
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            row_idx = row_idx + 1

        # Reactive Power
        sheet.cell(row=row_idx, column=1).value = gen_id
        sheet.cell(row=row_idx, column=2).value = node_id
        sheet.cell(row=row_idx, column=3).value = gen_type
        sheet.cell(row=row_idx, column=4).value = 'Qg, [MVAr]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_qg[gen_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        row_idx = row_idx + 1

        # Reactive Power net
        if generator.is_curtaillable() and network.params.rg_curt:

            sheet.cell(row=row_idx, column=1).value = gen_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = gen_type
            sheet.cell(row=row_idx, column=4).value = 'Qg_net, [MVAr]'
            sheet.cell(row=row_idx, column=5).value = 'Expected'
            sheet.cell(row=row_idx, column=6).value = expected_qg_net[gen_id]
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            row_idx = row_idx + 1

        # Apparent Power
        if generator.is_curtaillable() and network.params.rg_curt:

            sheet.cell(row=row_idx, column=1).value = gen_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = gen_type
            sheet.cell(row=row_idx, column=4).value = 'Sg, [MVA]'
            sheet.cell(row=row_idx, column=5).value = 'Expected'
            sheet.cell(row=row_idx, column=6).value = expected_sg[gen_id]
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            row_idx = row_idx + 1

            sheet.cell(row=row_idx, column=1).value = gen_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = gen_type
            sheet.cell(row=row_idx, column=4).value = 'Sg_curt, [MVA]'
            sheet.cell(row=row_idx, column=5).value = 'Expected'
            sheet.cell(row=row_idx, column=6).value = expected_sg_curt[gen_id]
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            row_idx = row_idx + 1

            sheet.cell(row=row_idx, column=1).value = gen_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = gen_type
            sheet.cell(row=row_idx, column=4).value = 'Sg_net, [MVA]'
            sheet.cell(row=row_idx, column=5).value = 'Expected'
            sheet.cell(row=row_idx, column=6).value = expected_sg_net[gen_id]
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            row_idx = row_idx + 1


def _write_network_branch_results_to_excel(network, workbook, results, result_type, n=0):

    sheet_name = str()
    aux_string = str()
    if result_type == 'losses':
        sheet_name = 'Branch Losses'
        aux_string = 'P, [MW]'
    elif result_type == 'ratio':
        sheet_name = 'Transformer Ratio'
        aux_string = 'Ratio'

    row_idx = 1
    decimal_style = '0.00'

    sheet = workbook.create_sheet(sheet_name)

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Branch ID'
    sheet.cell(row=row_idx, column=2).value = 'From Node ID'
    sheet.cell(row=row_idx, column=3).value = 'To Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=6).value = n
    row_idx = row_idx + 1

    expected_values = dict()
    for branch in network.branches:
        expected_values[branch.branch_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for branch in network.branches:

            branch_id = branch.branch_id

            if not(result_type == 'ratio' and not branch.is_transformer):

                value = results['scenarios'][s_o]['branches'][result_type][branch_id]

                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = aux_string
                sheet.cell(row=row_idx, column=5).value = s_o
                sheet.cell(row=row_idx, column=6).value = value
                sheet.cell(row=row_idx, column=6).number_format = decimal_style
                expected_values[branch_id] += value * omega_s
                row_idx = row_idx + 1

    for branch in network.branches:

        branch_id = branch.branch_id

        if not (result_type == 'ratio' and not branch.is_transformer):

            sheet.cell(row=row_idx, column=1).value = branch_id
            sheet.cell(row=row_idx, column=2).value = branch.fbus
            sheet.cell(row=row_idx, column=3).value = branch.tbus
            sheet.cell(row=row_idx, column=4).value = aux_string
            sheet.cell(row=row_idx, column=5).value = 'Expected'
            sheet.cell(row=row_idx, column=6).value = expected_values[branch_id]
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            row_idx = row_idx + 1


def _write_network_branch_loading_results_to_excel(network, workbook, results, n=0):

    sheet = workbook.create_sheet('Branch Loading')

    row_idx = 1
    perc_style = '0.00%'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Branch ID'
    sheet.cell(row=row_idx, column=2).value = 'From Node ID'
    sheet.cell(row=row_idx, column=3).value = 'To Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=6).value = n
    row_idx = row_idx + 1

    expected_values = {'flow_ij': {}}
    for branch in network.branches:
        expected_values['flow_ij'][branch.branch_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for branch in network.branches:

            value = results['scenarios'][s_o]['branches']['branch_flow']['flow_ij_perc'][branch.branch_id]

            # flow ij, [%]
            sheet.cell(row=row_idx, column=1).value = branch.branch_id
            sheet.cell(row=row_idx, column=2).value = branch.fbus
            sheet.cell(row=row_idx, column=3).value = branch.tbus
            sheet.cell(row=row_idx, column=4).value = 'Flow_ij, [%]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = value
            sheet.cell(row=row_idx, column=6).number_format = perc_style
            if value > 1.00 + VIOLATION_TOLERANCE:
                sheet.cell(row=row_idx, column=6).fill = violation_fill
            expected_values['flow_ij'][branch.branch_id] += value * omega_s
            row_idx = row_idx + 1

    for branch in network.branches:

        value = expected_values['flow_ij'][branch.branch_id]

        # flow ij, [%]
        sheet.cell(row=row_idx, column=1).value = branch.branch_id
        sheet.cell(row=row_idx, column=2).value = branch.fbus
        sheet.cell(row=row_idx, column=3).value = branch.tbus
        sheet.cell(row=row_idx, column=4).value = 'Flow_ij, [%]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = value
        sheet.cell(row=row_idx, column=6).number_format = perc_style
        if value > 1.00 + VIOLATION_TOLERANCE:
            sheet.cell(row=row_idx, column=6).fill = violation_fill
        row_idx = row_idx + 1


def _write_network_branch_power_flow_results_to_excel(network, workbook, results, n=0):

    sheet = workbook.create_sheet('Power Flows')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'BranchID'
    sheet.cell(row=row_idx, column=2).value = 'From Node ID'
    sheet.cell(row=row_idx, column=3).value = 'To Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=6).value = n
    row_idx = row_idx + 1

    expected_values = {'pij': {}, 'pji': {}, 'qij': {}, 'qji': {}, 'sij': {}, 'sji': {}}
    for branch in network.branches:
        branch_id = branch.branch_id
        expected_values['pij'][branch_id] = 0.00
        expected_values['pji'][branch_id] = 0.00
        expected_values['qij'][branch_id] = 0.00
        expected_values['qji'][branch_id] = 0.00
        expected_values['sij'][branch_id] = 0.00
        expected_values['sji'][branch_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for branch in network.branches:

            branch_id = branch.branch_id
            rating = branch.rate
            if rating == 0.0:
                rating = BRANCH_UNKNOWN_RATING

            pij = results['scenarios'][s_o]['branches']['power_flow']['pij'][branch_id]
            pji = results['scenarios'][s_o]['branches']['power_flow']['pji'][branch_id]
            qij = results['scenarios'][s_o]['branches']['power_flow']['qij'][branch_id]
            qji = results['scenarios'][s_o]['branches']['power_flow']['qji'][branch_id]
            sij = results['scenarios'][s_o]['branches']['power_flow']['sij'][branch_id]
            sji = results['scenarios'][s_o]['branches']['power_flow']['sji'][branch_id]

            # Pij, [MW]
            sheet.cell(row=row_idx, column=1).value = branch_id
            sheet.cell(row=row_idx, column=2).value = branch.fbus
            sheet.cell(row=row_idx, column=3).value = branch.tbus
            sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = pij
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            expected_values['pij'][branch_id] += pij * omega_s
            row_idx = row_idx + 1

            # Pji, [MW]
            sheet.cell(row=row_idx, column=1).value = branch_id
            sheet.cell(row=row_idx, column=2).value = branch.tbus
            sheet.cell(row=row_idx, column=3).value = branch.fbus
            sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = pji
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            expected_values['pji'][branch_id] += pji * omega_s
            row_idx = row_idx + 1

            # Qij, [MVAr]
            sheet.cell(row=row_idx, column=1).value = branch_id
            sheet.cell(row=row_idx, column=2).value = branch.fbus
            sheet.cell(row=row_idx, column=3).value = branch.tbus
            sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = qij
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            expected_values['qij'][branch_id] += qij * omega_s
            row_idx = row_idx + 1

            # Qji, [MW]
            sheet.cell(row=row_idx, column=1).value = branch_id
            sheet.cell(row=row_idx, column=2).value = branch.tbus
            sheet.cell(row=row_idx, column=3).value = branch.fbus
            sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = qji
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            expected_values['qji'][branch_id] += qji * omega_s
            row_idx = row_idx + 1

            # Sij, [MVA]
            sheet.cell(row=row_idx, column=1).value = branch_id
            sheet.cell(row=row_idx, column=2).value = branch.fbus
            sheet.cell(row=row_idx, column=3).value = branch.tbus
            sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = sij
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            expected_values['sij'][branch_id] += sij * omega_s
            row_idx = row_idx + 1

            # Sij, [%]
            sheet.cell(row=row_idx, column=1).value = branch_id
            sheet.cell(row=row_idx, column=2).value = branch.fbus
            sheet.cell(row=row_idx, column=3).value = branch.tbus
            sheet.cell(row=row_idx, column=4).value = 'S, [%]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = sij / rating
            sheet.cell(row=row_idx, column=6).number_format = perc_style
            row_idx = row_idx + 1

            # Sji, [MW]
            sheet.cell(row=row_idx, column=1).value = branch_id
            sheet.cell(row=row_idx, column=2).value = branch.tbus
            sheet.cell(row=row_idx, column=3).value = branch.fbus
            sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = sji
            sheet.cell(row=row_idx, column=6).number_format = decimal_style
            expected_values['sji'][branch_id]+= sji * omega_s
            row_idx = row_idx + 1

            # Sji, [%]
            sheet.cell(row=row_idx, column=1).value = branch_id
            sheet.cell(row=row_idx, column=2).value = branch.tbus
            sheet.cell(row=row_idx, column=3).value = branch.fbus
            sheet.cell(row=row_idx, column=4).value = 'S, [%]'
            sheet.cell(row=row_idx, column=5).value = s_o
            sheet.cell(row=row_idx, column=6).value = sji / rating
            sheet.cell(row=row_idx, column=6).number_format = perc_style
            row_idx = row_idx + 1

    for branch in network.branches:

        branch_id = branch.branch_id
        rating = branch.rate
        if rating == 0.0:
            rating = BRANCH_UNKNOWN_RATING

        # Pij, [MW]
        sheet.cell(row=row_idx, column=1).value = branch_id
        sheet.cell(row=row_idx, column=2).value = branch.fbus
        sheet.cell(row=row_idx, column=3).value = branch.tbus
        sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_values['pij'][branch_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        row_idx = row_idx + 1

        # Pji, [MW]
        sheet.cell(row=row_idx, column=1).value = branch_id
        sheet.cell(row=row_idx, column=2).value = branch.tbus
        sheet.cell(row=row_idx, column=3).value = branch.fbus
        sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_values['pji'][branch_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        row_idx = row_idx + 1

        # Qij, [MVAr]
        sheet.cell(row=row_idx, column=1).value = branch_id
        sheet.cell(row=row_idx, column=2).value = branch.fbus
        sheet.cell(row=row_idx, column=3).value = branch.tbus
        sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_values['qij'][branch_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        row_idx = row_idx + 1

        # Qji, [MVAr]
        sheet.cell(row=row_idx, column=1).value = branch_id
        sheet.cell(row=row_idx, column=2).value = branch.tbus
        sheet.cell(row=row_idx, column=3).value = branch.fbus
        sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_values['qji'][branch_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        row_idx = row_idx + 1

        # Sij, [MVA]
        sheet.cell(row=row_idx, column=1).value = branch_id
        sheet.cell(row=row_idx, column=2).value = branch.fbus
        sheet.cell(row=row_idx, column=3).value = branch.tbus
        sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_values['sij'][branch_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        row_idx = row_idx + 1

        # Sij, [%]
        sheet.cell(row=row_idx, column=1).value = branch_id
        sheet.cell(row=row_idx, column=2).value = branch.fbus
        sheet.cell(row=row_idx, column=3).value = branch.tbus
        sheet.cell(row=row_idx, column=4).value = 'S, [%]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_values['sij'][branch_id] / rating
        sheet.cell(row=row_idx, column=6).number_format = perc_style
        row_idx = row_idx + 1

        # Sji, [MVA]
        sheet.cell(row=row_idx, column=1).value = branch_id
        sheet.cell(row=row_idx, column=2).value = branch.tbus
        sheet.cell(row=row_idx, column=3).value = branch.fbus
        sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_values['sji'][branch_id]
        sheet.cell(row=row_idx, column=6).number_format = decimal_style
        row_idx = row_idx + 1

        # Sji, [%]
        sheet.cell(row=row_idx, column=1).value = branch_id
        sheet.cell(row=row_idx, column=2).value = branch.tbus
        sheet.cell(row=row_idx, column=3).value = branch.fbus
        sheet.cell(row=row_idx, column=4).value = 'S, [%]'
        sheet.cell(row=row_idx, column=5).value = 'Expected'
        sheet.cell(row=row_idx, column=6).value = expected_values['sji'][branch_id] / rating
        sheet.cell(row=row_idx, column=6).number_format = perc_style
        row_idx = row_idx + 1


def _write_network_energy_storage_results_to_excel(network, workbook, results, n=0):

    sheet = workbook.create_sheet('Energy Storage')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'ESS ID'
    sheet.cell(row=row_idx, column=2).value = 'Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Quantity'
    sheet.cell(row=row_idx, column=4).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=5).value = n
    row_idx = row_idx + 1

    expected_p = dict()
    expected_q = dict()
    expected_s = dict()
    expected_soc = dict()
    expected_soc_perc = dict()
    for energy_storage in network.energy_storages:
        es_id = energy_storage.es_id
        expected_p[es_id] = 0.00
        expected_q[es_id] = 0.00
        expected_s[es_id] = 0.00
        expected_soc[es_id] = 0.00
        expected_soc_perc[es_id] = 0.00

    for s_o in results['scenarios']:

        omega_s = network.prob_operation_scenarios[s_o]

        for energy_storage in network.energy_storages:

            es_id = energy_storage.es_id
            node_id = energy_storage.bus

            pc = results['scenarios'][s_o]['energy_storages']['p'][es_id]
            qc = results['scenarios'][s_o]['energy_storages']['q'][es_id]
            sc = results['scenarios'][s_o]['energy_storages']['s'][es_id]
            soc = results['scenarios'][s_o]['energy_storages']['soc'][es_id]
            soc_perc = results['scenarios'][s_o]['energy_storages']['soc_percent'][es_id]

            # - Active Power
            sheet.cell(row=row_idx, column=1).value = es_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = pc
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if pc != 'N/A':
                expected_p[es_id] += pc * omega_s
            else:
                expected_p[es_id] = 'N/A'
            row_idx = row_idx + 1

            # - Reactive Power
            sheet.cell(row=row_idx, column=1).value = es_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = qc
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if qc != 'N/A':
                expected_q[es_id] += qc * omega_s
            else:
                expected_q[es_id] = 'N/A'
            row_idx = row_idx + 1

            # - Apparent Power
            sheet.cell(row=row_idx, column=1).value = es_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'S, [MVA]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = sc
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if sc != 'N/A':
                expected_s[es_id] += sc * omega_s
            else:
                expected_s[es_id] = 'N/A'
            row_idx = row_idx + 1

            # - SoC, [MWh]
            sheet.cell(row=row_idx, column=1).value = es_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'SoC, [MWh]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = soc
            sheet.cell(row=row_idx, column=5).number_format = decimal_style
            if soc != 'N/A':
                expected_soc[es_id] += soc * omega_s
            else:
                expected_soc[es_id] = 'N/A'
            row_idx = row_idx + 1

            # - SoC, [%]
            sheet.cell(row=row_idx, column=1).value = es_id
            sheet.cell(row=row_idx, column=2).value = node_id
            sheet.cell(row=row_idx, column=3).value = 'SoC, [%]'
            sheet.cell(row=row_idx, column=4).value = s_o
            sheet.cell(row=row_idx, column=5).value = soc_perc
            sheet.cell(row=row_idx, column=5).number_format = perc_style
            if soc_perc != 'N/A':
                expected_soc_perc[es_id] += soc_perc * omega_s
            else:
                expected_soc_perc[es_id] = 'N/A'
            row_idx = row_idx + 1

    for energy_storage in network.energy_storages:

        es_id = energy_storage.es_id
        node_id = energy_storage.bus

        # - Active Power
        sheet.cell(row=row_idx, column=1).value = es_id
        sheet.cell(row=row_idx, column=2).value = node_id
        sheet.cell(row=row_idx, column=3).value = 'P, [MW]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_p[es_id]
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx = row_idx + 1

        # - Reactive Power
        sheet.cell(row=row_idx, column=1).value = es_id
        sheet.cell(row=row_idx, column=2).value = node_id
        sheet.cell(row=row_idx, column=3).value = 'Q, [MVAr]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_q[es_id]
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx = row_idx + 1

        # - Apparent Power
        sheet.cell(row=row_idx, column=1).value = es_id
        sheet.cell(row=row_idx, column=2).value = node_id
        sheet.cell(row=row_idx, column=3).value = 'S, [MVA]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_s[es_id]
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx = row_idx + 1

        # - SoC, [MWh]
        sheet.cell(row=row_idx, column=1).value = es_id
        sheet.cell(row=row_idx, column=2).value = node_id
        sheet.cell(row=row_idx, column=3).value = 'SoC, [MWh]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_soc[es_id]
        sheet.cell(row=row_idx, column=5).number_format = decimal_style
        row_idx = row_idx + 1

        # - SoC, [%]
        sheet.cell(row=row_idx, column=1).value = es_id
        sheet.cell(row=row_idx, column=2).value = node_id
        sheet.cell(row=row_idx, column=3).value = 'SoC, [%]'
        sheet.cell(row=row_idx, column=4).value = 'Expected'
        sheet.cell(row=row_idx, column=5).value = expected_soc_perc[es_id]
        sheet.cell(row=row_idx, column=5).number_format = perc_style
        row_idx = row_idx + 1


def _write_relaxation_slacks_scenarios_results_to_excel(network, workbook, results, n=0):

    sheet = workbook.create_sheet('Relaxation Slacks, Operation')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Resource ID'
    sheet.cell(row=row_idx, column=2).value = 'Quantity'
    sheet.cell(row=row_idx, column=3).value = 'Operation Scenario'
    sheet.cell(row=row_idx, column=4).value = n
    row_idx = row_idx + 1

    for s_o in results['scenarios']:

        # Voltage slacks
        if network.params.slacks.grid_operation.voltage:

            for node in network.nodes:

                node_id = node.bus_i
                slack_e = results['scenarios'][s_o]['relaxation_slacks']['voltage']['e'][node_id]
                slack_f = results['scenarios'][s_o]['relaxation_slacks']['voltage']['f'][node_id]

                # - slack_e
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'Voltage, e'
                sheet.cell(row=row_idx, column=3).value = s_o
                sheet.cell(row=row_idx, column=4).value = slack_e
                sheet.cell(row=row_idx, column=4).number_format = decimal_style
                row_idx = row_idx + 1

                # - slack_f
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'Voltage, f'
                sheet.cell(row=row_idx, column=3).value = s_o
                sheet.cell(row=row_idx, column=4).value = slack_f
                sheet.cell(row=row_idx, column=4).number_format = decimal_style
                row_idx = row_idx + 1

        # Branch flow slacks
        if network.params.slacks.grid_operation.branch_flow:

            for branch in network.branches:

                branch_id = branch.branch_id
                iij_sqr = results['scenarios'][s_o]['relaxation_slacks']['branch_flow']['flow_ij_sqr'][branch_id]

                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = 'Flow_ij_sqr'
                sheet.cell(row=row_idx, column=3).value = s_o
                sheet.cell(row=row_idx, column=4).value = iij_sqr
                sheet.cell(row=row_idx, column=4).number_format = decimal_style
                row_idx = row_idx + 1

        # Node balance
        if network.params.slacks.node_balance:

            for node in network.nodes:

                node_id = node.bus_i
                slack_p = results['scenarios'][s_o]['relaxation_slacks']['node_balance']['p'][node_id]
                slack_q = results['scenarios'][s_o]['relaxation_slacks']['node_balance']['q'][node_id]

                # - slack_p
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'Node balance, p'
                sheet.cell(row=row_idx, column=3).value = s_o
                sheet.cell(row=row_idx, column=4).value = slack_p
                sheet.cell(row=row_idx, column=4).number_format = decimal_style
                row_idx = row_idx + 1

                # - slack_q
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'Node balance, q'
                sheet.cell(row=row_idx, column=3).value = s_o
                sheet.cell(row=row_idx, column=4).value = slack_q
                sheet.cell(row=row_idx, column=4).number_format = decimal_style
                row_idx = row_idx + 1

        # Shared ESS
        for shared_energy_storage in network.shared_energy_storages:

            node_id = shared_energy_storage.bus

            # - Complementarity
            if network.params.slacks.shared_ess.complementarity:

                comp = results['scenarios'][s_o]['relaxation_slacks']['shared_energy_storages']['comp'][node_id]

                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'Shared Energy Storage, comp'
                sheet.cell(row=row_idx, column=3).value = s_o
                sheet.cell(row=row_idx, column=4).value = comp
                sheet.cell(row=row_idx, column=4).number_format = decimal_style
                row_idx = row_idx + 1

        # ESS
        if network.params.es_reg:

            for energy_storage in network.energy_storages:

                es_id = energy_storage.es_id

                # - Complementarity
                if network.params.slacks.ess.complementarity:

                    comp = results['scenarios'][s_o]['relaxation_slacks']['energy_storages']['comp'][es_id]

                    sheet.cell(row=row_idx, column=1).value = es_id
                    sheet.cell(row=row_idx, column=2).value = 'Energy Storage, comp'
                    sheet.cell(row=row_idx, column=3).value = s_o
                    sheet.cell(row=row_idx, column=4).value = comp
                    sheet.cell(row=row_idx, column=4).number_format = decimal_style
                    row_idx = row_idx + 1


# ======================================================================================================================
#   Other (aux) functions
# ======================================================================================================================
def _perform_network_check(network):

    n_bus = len(network.nodes)
    if n_bus == 0:
        print(f'[ERROR] Reading network {network.name}. No nodes imported.')
        exit(ERROR_NETWORK_FILE)

    n_branch = len(network.branches)
    if n_branch == 0:
        print(f'[ERROR] Reading network {network.name}. No branches imported.')
        exit(ERROR_NETWORK_FILE)


def _pre_process_network(network):

    processed_nodes = []
    for node in network.nodes:
        if node.type != BUS_ISOLATED:
            processed_nodes.append(node)

    processed_gens = []
    for gen in network.generators:
        node_type = network.get_node_type(gen.bus)
        if node_type != BUS_ISOLATED:
            processed_gens.append(gen)

    processed_branches = []
    for branch in network.branches:

        if not branch.is_connected():  # If branch is disconnected for all days and periods, remove
            continue

        if branch.pre_processed:
            continue

        fbus, tbus = branch.fbus, branch.tbus
        fnode_type = network.get_node_type(fbus)
        tnode_type = network.get_node_type(tbus)
        if fnode_type == BUS_ISOLATED or tnode_type == BUS_ISOLATED:
            branch.pre_processed = True
            continue

        parallel_branches = [branch for branch in network.branches if ((branch.fbus == fbus and branch.tbus == tbus) or (branch.fbus == tbus and branch.tbus == fbus))]
        connected_parallel_branches = [branch for branch in parallel_branches if branch.is_connected()]
        if len(connected_parallel_branches) > 1:
            processed_branch = connected_parallel_branches[0]
            r_eq, x_eq, g_eq, b_eq = _pre_process_parallel_branches(connected_parallel_branches)
            processed_branch.r = r_eq
            processed_branch.x = x_eq
            processed_branch.g_sh = g_eq
            processed_branch.b_sh = b_eq
            processed_branch.rate = sum([branch.rate for branch in connected_parallel_branches])
            processed_branch.ratio = branch.ratio
            processed_branch.pre_processed = True
            for branch_parallel in parallel_branches:
                branch_parallel.pre_processed = True
            processed_branches.append(processed_branch)
        else:
            for branch_parallel in parallel_branches:
                branch_parallel.pre_processed = True
            for branch_parallel in connected_parallel_branches:
                processed_branches.append(branch_parallel)

    network.nodes = processed_nodes
    network.generators = processed_gens
    network.branches = processed_branches
    for branch in network.branches:
        branch.pre_processed = False


def _pre_process_parallel_branches(branches):
    branch_impedances = [complex(branch.r, branch.x) for branch in branches]
    branch_shunt_admittance = [complex(branch.g_sh, branch.b_sh) for branch in branches]
    z_eq = 1/sum([(1/impedance) for impedance in branch_impedances])
    ysh_eq = sum([admittance for admittance in branch_shunt_admittance])
    return abs(z_eq.real), abs(z_eq.imag), ysh_eq.real, ysh_eq.imag


def _get_branch_power_losses(network, params, model, branch_idx, s_o):

    # Active power flow, from i to j and from j to i
    branch = network.branches[branch_idx]
    pij, _ = _get_branch_power_flow(network, params, branch, branch.fbus, branch.tbus, model, s_o)
    pji, _ = _get_branch_power_flow(network, params, branch, branch.tbus, branch.fbus, model, s_o)

    return abs(pij + pji)


def _get_branch_power_flow(network, params, branch, fbus, tbus, model, s_o):

    fbus_idx = network.get_node_idx(fbus)
    tbus_idx = network.get_node_idx(tbus)
    branch_idx = network.get_branch_idx(branch)

    rij = pe.value(model.r[branch_idx, s_o])
    ei = pe.value(model.e_actual[fbus_idx, s_o])
    fi = pe.value(model.f_actual[fbus_idx, s_o])
    ej = pe.value(model.e_actual[tbus_idx, s_o])
    fj = pe.value(model.f_actual[tbus_idx, s_o])

    if branch.fbus == fbus:
        pij = branch.g * (ei ** 2 + fi ** 2) * rij ** 2
        pij -= branch.g * (ei * ej + fi * fj) * rij
        pij -= branch.b * (fi * ej - ei * fj) * rij

        qij = - (branch.b + branch.b_sh * 0.50) * (ei ** 2 + fi ** 2) * rij ** 2
        qij += branch.b * (ei * ej + fi * fj) * rij
        qij -= branch.g * (fi * ej - ei * fj) * rij
    else:
        pij = branch.g * (ei ** 2 + fi ** 2)
        pij -= branch.g * (ei * ej + fi * fj) * rij
        pij -= branch.b * (fi * ej - ei * fj) * rij

        qij = - (branch.b + branch.b_sh * 0.50) * (ei ** 2 + fi ** 2)
        qij += branch.b * (ei * ej + fi * fj) * rij
        qij -= branch.g * (fi * ej - ei * fj) * rij

    return pij * network.baseMVA, qij * network.baseMVA


def _get_info_from_results(results, info_string):
    i = str(results).lower().find(info_string.lower()) + len(info_string)
    value = ''
    while str(results)[i] != '\n':
        value = value + str(results)[i]
        i += 1
    return value
